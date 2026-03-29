"""
bridge.py — WebView ↔ Python 연결층

설계 원칙:
  - 호출/반환/상태 연결만 담당 (UI 로직·비즈니스 로직 금지)
  - ScanResult/RunResult 원본은 JS로 절대 넘기지 않음 → to_*_payload() 변환 후 전달
  - JS → Python: 슬롯 호출 + params_json
  - Python → JS: pyqtSignal emit (QWebChannel)
  - 모든 날짜는 YYYY-MM-DD 문자열로 통일

엔진 시그니처 주의사항 (실제 코드 기준):
  - load_all_school_names(roster_xlsx, col_map)  ← work_root 아님
  - get_school_domain(roster_xlsx, school_name, col_map)
  - get_project_dirs(work_root)  ← school_name 없음
  - run_main_engine(scan, work_date, school_start_date, ...)  ← ScanResult 객체 직접 받음
  - run_diff_engine(work_root, school_name, ...)  ← 내부에서 scan 재실행
"""

from __future__ import annotations

import time
import json
import re
import traceback as tb_module
from pathlib import Path

# 디버그 출력용
def _dbg(*args):
    print("[BRIDGE]", *args, flush=True)

from PyQt6.QtCore import QObject, QThread, pyqtSignal, pyqtSlot
from PyQt6.QtWidgets import QApplication, QFileDialog

from engine import (
    inspect_work_root,
    load_all_school_names,
    scan_main_engine,
    run_main_engine,
    scan_diff_engine,
    run_diff_engine,
    get_school_domain,
    get_project_dirs,
    load_notice_templates,
)
from core.roster_log import write_work_result, write_email_sent
from core.events import CoreEvent, RowMark

from core.common import derive_grade_year_map


# ── app_config 직접 읽기/쓰기 (별도 모듈 없음) ──────────────

CONFIG_PATH = Path("app_config.json")

DEFAULT_APP_CONFIG = {
    "work_root": "",
    "roster_log_path": "",
    "worker_name": "",
    "school_start_date": "",
    "work_date": "",
    "last_school": "",
    "roster_col_map": {
        "sheet": "",
        "header_row": 0,
        "data_start": 0,
        "col_school": 0,
        "col_email_arr": 0,
        "col_email_snt": 0,
        "col_worker": 0,
        "col_freshmen": 0,
        "col_transfer": 0,
        "col_withdraw": 0,
        "col_teacher": 0,
        "col_seq": 0,
    },
}

def _load_app_config() -> dict:
    if CONFIG_PATH.exists():
        try:
            return json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
        except Exception:
            pass
    return dict(DEFAULT_APP_CONFIG)

def _save_app_config(config: dict) -> None:
    CONFIG_PATH.write_text(
        json.dumps(config, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


# ── work_history 직접 읽기/쓰기 ──────────────────────────────

def _work_history_path(school_year: int) -> Path:
    return Path(f"work_history_{school_year}.json")

def _load_work_history(school_year: int) -> dict:
    path = _work_history_path(school_year)
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}

def _save_work_history(school_year: int, school_name: str, entry: dict) -> None:
    history = _load_work_history(school_year)
    history[school_name] = entry
    _work_history_path(school_year).write_text(
        json.dumps(history, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


# ──────────────────────────────────────────────
# 공통 응답 헬퍼
# ──────────────────────────────────────────────

def ok_response(data: dict) -> str:
    return json.dumps({"ok": True, "data": data}, ensure_ascii=False)

def error_response(message: str) -> str:
    return json.dumps({"ok": False, "error": message}, ensure_ascii=False)

def async_ok(task: str, data: dict) -> str:
    return json.dumps({"ok": True, "task": task, "data": data}, ensure_ascii=False)

def async_error(task: str, message: str, traceback: str = "") -> str:
    return json.dumps(
        {"ok": False, "task": task, "error": message, "traceback": traceback},
        ensure_ascii=False,
    )


# ──────────────────────────────────────────────
# Result -> Payload 변환
# (ScanResult/PipelineResult 원본은 JS로 절대 노출하지 않는다)
# ──────────────────────────────────────────────

def _parse_log_entry(raw: str) -> dict:
    """'[INFO] ...' 또는 '[HH:MM:SS] [INFO] ...' 형식 로그 -> {level, message} dict"""
    import re as _re
    s = str(raw)
    # 타임스탬프 prefix 제거: [22:11:05] 형태
    s = _re.sub(r'^\[\d{2}:\d{2}:\d{2}\]\s*', '', s)
    if s.startswith("[ERROR]"):
        return {"level": "error", "message": s[7:].strip()}
    if s.startswith("[WARN]"):
        return {"level": "warn",  "message": s[6:].strip()}
    if s.startswith("[DEBUG]"):
        return {"level": "debug", "message": s[7:].strip()}
    if s.startswith("[INFO]"):
        return {"level": "info",  "message": s[6:].strip()}
    if s.startswith("[DONE]"):
        return {"level": "info",  "message": s[6:].strip()}
    if s.startswith("[TIMER]"):
        return {"level": "info",  "message": s.strip()}
    return {"level": "info", "message": s.strip()}

def _logs_from_result(result) -> list:
    return [_parse_log_entry(l) for l in (getattr(result, "logs", None) or [])]


def _status_from_events(events: list) -> dict:
    errs  = [e for e in events if e.level == "error"]
    holds = [e for e in events if e.level == "hold"]
    warns = [e for e in events if e.level == "warn"]
    if errs:    level, badge = "error", {"type": "err",  "text": "오류"}
    elif holds: level, badge = "hold",  {"type": "hold", "text": "보류"}
    elif warns: level, badge = "warn",  {"type": "warn", "text": "경고"}
    else:       level, badge = "ok",    {"type": "ok",   "text": "완료"}
    all_msgs = errs + holds + warns
    if errs:    summary = f"오류 {len(errs)}건이 있습니다."
    elif holds: summary = f"보류 {len(holds)}건이 있습니다."
    elif warns: summary = f"경고 {len(warns)}건이 있습니다."
    else:       summary = "완료"
    return {
        "level": level, "badge": badge,
        "messages": [{"level": e.level, "text": e.message} for e in all_msgs],
        "summary_text": summary,
        "detail_messages": [e.message for e in all_msgs],
        "action_text": "",
        "row_marks": {"warn_rows": [], "error_rows": [], "issue_rows": []},
    }


def _serialize_event(e: CoreEvent) -> dict:
    return {
        "code": e.code, "level": e.level, "message": e.message,
        "detail": e.detail, "file_key": e.file_key,
        "row": e.row, "field_name": e.field_name, "blocking": e.blocking,
    }


def _serialize_row_mark(m: RowMark) -> dict:
    return {"file_key": m.file_key, "row": m.row, "level": m.level, "code": m.code}


def _meta_get(meta, key, default=None):
    if meta is None:
        return default
    if isinstance(meta, dict):
        return meta.get(key, default)
    return getattr(meta, key, default)


def _as_abs_issue_rows(issue_rows, data_start_row):
    out = []
    base = int(data_start_row or 1)
    for r in list(issue_rows or []):
        try:
            r_i = int(r)
        except Exception:
            continue
        if r_i >= base:
            out.append(r_i)
        else:
            out.append(base + r_i)
    # stable unique
    seen = set()
    uniq = []
    for r in out:
        if r not in seen:
            seen.add(r)
            uniq.append(r)
    return uniq


def _build_status_bundle(*, logs=None, warning_text='', issue_rows=None, running=False, ok=True, default_ok_text='완료', action_text=''):
    logs = list(logs or [])
    warn_msgs = []
    err_msgs = []
    for l in logs:
        level = str(l.get('level', 'info'))
        msg = str(l.get('message', '')).strip()
        if not msg:
            continue
        if level == 'error':
            if msg not in err_msgs:
                err_msgs.append(msg)
        elif level == 'warn':
            if msg not in warn_msgs:
                warn_msgs.append(msg)
    for msg in str(warning_text or '').splitlines():
        msg = msg.strip()
        if msg and msg not in warn_msgs:
            warn_msgs.append(msg)

    if running:
        level = 'running'; badge_type = 'running'; badge_text = '실행 중'
    elif err_msgs or not ok:
        level = 'error'; badge_type = 'err'; badge_text = '오류'
    elif warn_msgs or list(issue_rows or []):
        level = 'warn'; badge_type = 'warn'; badge_text = '경고'
    else:
        level = 'ok'; badge_type = 'ok'; badge_text = default_ok_text

    detail_messages = err_msgs if level == 'error' else warn_msgs
    if level == 'error':
        summary_text = f"오류 {len(err_msgs)}건이 있습니다." if err_msgs else '오류가 있습니다.'
    elif level == 'warn':
        summary_text = f"경고 {len(warn_msgs)}건이 있습니다." if warn_msgs else '경고가 있습니다.'
    elif running:
        summary_text = '실행 중입니다.'
    else:
        summary_text = default_ok_text

    return {
        'level': level,
        'badge': {'type': badge_type, 'text': badge_text},
        'messages': ([{'level': 'error', 'text': m} for m in err_msgs] +
                     [{'level': 'warn', 'text': m} for m in warn_msgs]),
        'summary_text': summary_text,
        'detail_messages': detail_messages,
        'action_text': str(action_text or '').strip(),
        'row_marks': {
            'warn_rows': list(issue_rows or []),
            'error_rows': [],
            'issue_rows': list(issue_rows or []),
        },
    }


def _normalize_scan_item(meta, kind_label):
    if meta is None:
        return None
    data_start_row = _meta_get(meta, 'data_start_row', 2)
    warning = _meta_get(meta, 'warning', '')
    issue_rows = _as_abs_issue_rows(_meta_get(meta, 'issue_rows', []) or [], data_start_row)
    logs = []
    severity = _meta_get(meta, 'severity', 'ok')
    if severity == 'error' and warning:
        logs = [{'level': 'error', 'message': warning}]
    item = {
        'kind': kind_label,
        'file_name': _meta_get(meta, 'file_name', ''),
        'file_path': _meta_get(meta, 'file_path', ''),
        'sheet_name': _meta_get(meta, 'sheet_name', ''),
        'header_row': _meta_get(meta, 'header_row', 1),
        'data_start_row': data_start_row,
        'row_count': _meta_get(meta, 'row_count', 0),
        'warning': warning,
        'issue_rows': issue_rows,
        'extra_grades': list(_meta_get(meta, 'extra_grades', []) or []),
        'severity': severity,
    }
    item['status'] = _build_status_bundle(logs=logs, warning_text=(warning if severity != 'error' else ''), issue_rows=issue_rows, ok=(severity != 'error'), default_ok_text='확인 완료')
    return item


def _normalize_compare_item(compare_file, compare_layout):
    if not compare_file:
        return None
    compare_layout = compare_layout or {}
    issue_rows = _as_abs_issue_rows(_meta_get(compare_layout, 'issue_rows', []) or [], _meta_get(compare_layout, 'data_start_row', 2))
    warning = _meta_get(compare_layout, 'warning', '')
    item = {
        'kind': '재학생',
        'file_name': getattr(compare_file, 'name', str(compare_file)),
        'file_path': str(compare_file),
        'sheet_name': _meta_get(compare_layout, 'sheet_name', ''),
        'header_row': _meta_get(compare_layout, 'header_row', 1),
        'data_start_row': _meta_get(compare_layout, 'data_start_row', 2),
        'row_count': _meta_get(compare_layout, 'row_count', 0),
        'warning': warning,
        'issue_rows': issue_rows,
        'extra_grades': [],
        'severity': _meta_get(compare_layout, 'severity', 'warn' if warning or issue_rows else 'ok'),
    }
    item['status'] = _build_status_bundle(logs=[], warning_text=warning, issue_rows=issue_rows, ok=(item['severity'] != 'error'), default_ok_text='확인 완료')
    return item


def to_scan_payload(result) -> dict:
    """ScanResult / DiffScanResult -> JS 전달용 요약 dict"""
    logs      = _logs_from_result(result)
    events    = list(getattr(result, "events",    None) or [])
    row_marks = list(getattr(result, "row_marks", None) or [])
    warnings  = [l for l in logs if l["level"] in ("warn", "error")]

    items = [
        i for i in [
            _normalize_scan_item(getattr(result, 'freshmen', None), '신입생'),
            _normalize_scan_item(getattr(result, 'transfer_in', None), '전입생'),
            _normalize_scan_item(getattr(result, 'transfer_out', None), '전출생'),
            _normalize_scan_item(getattr(result, 'teachers', None), '교직원'),
            _normalize_compare_item(getattr(result, 'compare_file', None), getattr(result, 'compare_layout', None)),
        ] if i is not None
    ]

    roster_basis_date = ""
    rbd = getattr(result, "roster_basis_date", None)
    if rbd is not None:
        roster_basis_date = rbd.isoformat() if hasattr(rbd, "isoformat") else str(rbd)

    grade_year_map = {}
    roster_info = getattr(result, "roster_info", None)
    if roster_info is not None:
        target_grades = set()
        prefix_mode = getattr(roster_info, "prefix_mode_by_roster_grade", {}) or {}
        shift = int(getattr(roster_info, "ref_grade_shift", 0) or 0)
        for g_roster in prefix_mode.keys():
            try:
                g_cur = int(g_roster) - shift
            except Exception:
                continue
            if g_cur > 0:
                target_grades.add(g_cur)
        freshmen_meta = getattr(result, "freshmen", None) or {}
        extra_grades = _meta_get(freshmen_meta, 'extra_grades', []) or []
        for g in extra_grades:
            try:
                g_i = int(g)
            except Exception:
                continue
            if g_i > 0:
                target_grades.add(g_i)
        target_grades.update({1,2,3,4,5,6})
        year_int = int(getattr(result, "year_int", 0) or 0)
        derived = derive_grade_year_map(target_grades=sorted(target_grades), input_year=year_int, roster_info=roster_info)
        grade_year_map = {int(k): int(v) for k, v in derived.items()}

    overall_issue_rows = []
    for item in items:
        for r in item.get('issue_rows', []) or []:
            if r not in overall_issue_rows:
                overall_issue_rows.append(r)
    if events:
        status = _status_from_events(events)
        _warn_rows  = [m.row for m in row_marks if m.level in ("warn", "dup")]
        _error_rows = [m.row for m in row_marks if m.level == "error"]
        status["row_marks"] = {"warn_rows": _warn_rows, "error_rows": _error_rows, "issue_rows": _warn_rows + _error_rows}
    else:
        status = _build_status_bundle(logs=logs, issue_rows=overall_issue_rows, ok=bool(getattr(result, 'ok', False)), default_ok_text='스캔 완료')

    return {
        "ok":                     bool(getattr(result, "ok", False)),
        "school_profile_mode":    str(getattr(result, "school_profile_mode", "single") or "single"),
        "school_kind_needs_choice": bool(getattr(result, "school_kind_needs_choice", False)),
        "grade_rule_max_grade":   int(getattr(result, "grade_rule_max_grade", 6) or 6),
        "can_execute":            bool(getattr(result, "can_execute", False)),
        "can_execute_after_input":bool(getattr(result, "can_execute_after_input", False)),
        "missing_fields":         list(getattr(result, "missing_fields", []) or []),
        "needs_open_date":        bool(getattr(result, "needs_open_date", False)),
        "need_roster":            bool(getattr(result, "need_roster", False)),
        "roster_date_mismatch":   bool(getattr(result, "roster_date_mismatch", False)),
        "roster_basis_date":      roster_basis_date,
        "roster_path": str(getattr(result, "roster_path", None) or "") or None,
        "has_school_kind_warn":   False,
        "grade_year_map":         grade_year_map,
        "items":    items,
        "warnings":  warnings,
        "logs":      logs,
        "status":    status,
        "events":    [_serialize_event(e)    for e in events],
        "row_marks": [_serialize_row_mark(m) for m in row_marks],
    }


def to_run_payload(result) -> dict:
    """PipelineResult -> JS 전달용 요약 dict (Path -> str 전수 변환)"""
    logs      = _logs_from_result(result)
    events    = list(getattr(result, "events",    None) or [])
    row_marks = list(getattr(result, "row_marks", None) or [])
    warnings  = [l for l in logs if l["level"] in ("warn", "error")]
    audit = getattr(result, "audit_summary", {}) or {}
    in_cnt = audit.get("input_counts", {})
    transfer_in_hold = int(getattr(result, 'transfer_in_hold', 0) or 0)
    transfer_out_hold = int(getattr(result, 'transfer_out_hold', 0) or 0)
    transfer_out_auto_skip = int(getattr(result, 'transfer_out_auto_skip', 0) or 0)
    real_hold = transfer_in_hold + transfer_out_hold - transfer_out_auto_skip
    issue_rows = list(getattr(result, 'notice_dup_rows', []) or []) + list(getattr(result, 'notice_teacher_dup_rows', []) or [])
    action_text = ''
    if any('헤더를 찾을 수 없습니다.' in str(l.get('message', '')) for l in logs):
        action_text = '헤더행과 열 이름을 확인해 주세요.'
    status_logs = list(logs)
    if real_hold > 0:
        status_logs.append({'level':'warn','message':f'확인 필요 건이 {real_hold}건 있습니다.'})
    if events:
        status = _status_from_events(events)
    else:
        status = _build_status_bundle(
            logs=status_logs,
            issue_rows=issue_rows,
            ok=bool(getattr(result, 'ok', False)),
            default_ok_text='완료',
            action_text=action_text,
        )
    return {
        "ok": bool(getattr(result, "ok", False)),
        "output_files": [{"name": p.name, "path": str(p)} for p in (getattr(result, "outputs", None) or [])],
        "freshmen_count":          int(in_cnt.get("freshmen", 0)),
        "teacher_count":           int(in_cnt.get("teacher",  0)),
        "transfer_in_done":        int(getattr(result, "transfer_in_done",       0)),
        "transfer_in_hold":        int(getattr(result, "transfer_in_hold",       0)),
        "transfer_out_done":       int(getattr(result, "transfer_out_done",      0)),
        "transfer_out_hold":       int(getattr(result, "transfer_out_hold",      0)),
        "transfer_out_auto_skip":  int(getattr(result, "transfer_out_auto_skip", 0)),
        "notice_dup_rows":          list(getattr(result, "notice_dup_rows", []) or []),
        "notice_teacher_dup_rows":  list(getattr(result, "notice_teacher_dup_rows", []) or []),
        "warnings":  warnings,
        "logs":      logs,
        "status":    status,
        "events":    [_serialize_event(e)    for e in events],
        "row_marks": [_serialize_row_mark(m) for m in row_marks],
    }


def to_diff_run_payload(result) -> dict:
    """DiffPipelineResult -> JS 전달용 요약 dict"""
    logs      = _logs_from_result(result)
    events    = list(getattr(result, "events",    None) or [])
    row_marks = list(getattr(result, "row_marks", None) or [])
    warnings  = [l for l in logs if l["level"] in ("warn", "error")]

    if events:
        status = _status_from_events(events)
    else:
        status = _build_status_bundle(
            logs=logs, issue_rows=[],
            ok=bool(getattr(result, 'ok', False)),
            default_ok_text='완료',
        )

    return {
        "ok": bool(getattr(result, "ok", False)),
        "output_files": [{"name": p.name, "path": str(p)} for p in (getattr(result, "outputs", None) or [])],
        "compare_only_count": int(getattr(result, "compare_only_count", 0)),
        "roster_only_count":  int(getattr(result, "roster_only_count",  0)),
        "matched_count":      int(getattr(result, "matched_count",      0)),
        "unresolved_count":   int(getattr(result, "unresolved_count",   0)),
        "transfer_in_done":   int(getattr(result, "transfer_in_done",   0)),
        "transfer_in_hold":   int(getattr(result, "transfer_in_hold",   0)),
        "transfer_out_done":  int(getattr(result, "transfer_out_done",  0)),
        "transfer_out_hold":  int(getattr(result, "transfer_out_hold",  0)),
        "roster_only_rows":   list(getattr(result, "roster_only_rows", []) or []),
        "matched_rows":       list(getattr(result, "matched_rows", []) or []),
        "compare_only_rows":  list(getattr(result, "compare_only_rows", []) or []),
        "unresolved_rows":    list(getattr(result, "unresolved_rows", []) or []),
        "warnings":  warnings,
        "logs":      logs,
        "status":    status,
        "events":    [_serialize_event(e)    for e in events],
        "row_marks": [_serialize_row_mark(m) for m in row_marks],
    }


def _validate_date(date_str: str) -> bool:
    return bool(re.fullmatch(r"\d{4}-\d{2}-\d{2}", date_str or ""))


# ──────────────────────────────────────────────
# Workers
# ──────────────────────────────────────────────

class ScanWorker(QObject):
    finished = pyqtSignal(str)
    failed   = pyqtSignal(str)

    def __init__(self, params: dict):
        super().__init__()
        self._params = params
        self.scan_result = None

    def run(self):
        try:
            _dbg("ScanWorker.run start", {
                "school_name": self._params.get("school_name"),
                "work_root": self._params.get("work_root"),
                "roster_xlsx": self._params.get("roster_xlsx"),
                "work_date": self._params.get("work_date"),
                "school_start_date": self._params.get("school_start_date"),
                "roster_basis_date": self._params.get("roster_basis_date"),
            })

            started = time.time()

            result = scan_main_engine(
                work_root=self._params["work_root"],
                school_name=self._params["school_name"],
                school_start_date=self._params["school_start_date"],
                work_date=self._params["work_date"],
                roster_basis_date=self._params.get("roster_basis_date"),
                roster_xlsx=self._params.get("roster_xlsx") or None,
                col_map=self._params.get("col_map"),
            )

            _dbg("scan_main_engine returned", time.time() - started)

            self.scan_result = result
            payload = async_ok("scan_main", to_scan_payload(result))
            _dbg("ScanWorker before finished emit")
            self.finished.emit(payload)

        except Exception as e:
            _dbg("ScanWorker error", str(e))
            self.scan_result = None
            self.failed.emit(async_error("scan_main", str(e), tb_module.format_exc()))


class RunWorker(QObject):
    """run_main_engine은 ScanResult 객체를 직접 받는다."""
    finished = pyqtSignal(str)
    failed   = pyqtSignal(str)

    def __init__(self, scan_result, work_date: str, school_start_date: str,
                 layout_overrides=None, school_kind_override=None):
        super().__init__()
        self._scan = scan_result
        self._work_date = work_date
        self._school_start_date = school_start_date
        self._layout_overrides = layout_overrides
        self._school_kind_override = school_kind_override

    def run(self):
        try:
            result = run_main_engine(
                scan=self._scan,
                work_date=self._work_date,
                school_start_date=self._school_start_date,
                layout_overrides=self._layout_overrides,
                school_kind_override=self._school_kind_override,
            )
            self.finished.emit(async_ok("run_main", to_run_payload(result)))
        except Exception as e:
            self.failed.emit(async_error("run_main", str(e), tb_module.format_exc()))


class DiffScanWorker(QObject):
    finished = pyqtSignal(str)
    failed   = pyqtSignal(str)

    def __init__(self, params: dict):
        super().__init__()
        self._params = params

    def run(self):
        try:
            result = scan_diff_engine(
                work_root=self._params["work_root"],
                school_name=self._params["school_name"],
                target_year=self._params.get("target_year"),
                school_start_date=self._params["school_start_date"],
                work_date=self._params["work_date"],
                roster_basis_date=self._params.get("roster_basis_date"),
                roster_xlsx=self._params.get("roster_xlsx") or None,
                col_map=self._params.get("col_map"),
            )
            self.finished.emit(async_ok("diff_scan", to_scan_payload(result)))
        except Exception as e:
            self.failed.emit(async_error("diff_scan", str(e), tb_module.format_exc()))


class DiffRunWorker(QObject):
    """run_diff_engine은 내부에서 scan을 재실행하므로 scan 객체 불필요."""
    finished = pyqtSignal(str)
    failed   = pyqtSignal(str)

    def __init__(self, params: dict):
        super().__init__()
        self._params = params

    def run(self):
        try:
            result = run_diff_engine(
                work_root=self._params["work_root"],
                school_name=self._params["school_name"],
                target_year=self._params.get("target_year"),
                school_start_date=self._params["school_start_date"],
                work_date=self._params["work_date"],
                roster_basis_date=self._params.get("roster_basis_date"),
                roster_xlsx=self._params.get("roster_xlsx") or None,
                col_map=self._params.get("col_map"),
                layout_overrides=self._params.get("layout_overrides"),
            )
            self.finished.emit(async_ok("diff_run", to_diff_run_payload(result)))
        except Exception as e:
            self.failed.emit(async_error("diff_run", str(e), tb_module.format_exc()))


class PreviewWorker(QObject):
    """
    실제 시트의 1행부터 보여주는 미리보기 워커.
    시작행 이전은 JS에서 회색 처리하고, issue_rows는 경고 행으로 표시한다.
    """
    finished = pyqtSignal(str)
    failed   = pyqtSignal(str)

    MAX_PREVIEW_ROWS = 300

    def __init__(self, params: dict):
        super().__init__()
        self._params = params

    def run(self):
        kind = self._params.get("kind", "")
        try:
            from core.common import safe_load_workbook as _safe_wb
            from pathlib import Path as _Path

            file_path    = self._params["file_path"]
            sheet_name   = self._params.get("sheet_name", "")
            header_row   = int(self._params.get("header_row", 1))
            data_start   = int(self._params.get("data_start_row", 2))
            start_row    = int(self._params.get("start_row", 1) or 1)
            issue_rows   = list(self._params.get("issue_rows", []) or [])

            wb = _safe_wb(_Path(file_path), data_only=True, read_only=True)
            try:
                ws = (wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.worksheets[0])
                actual_sheet = getattr(ws, "title", None) or sheet_name

                hdr = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
                header_values = [str(c) if c is not None else "" for c in (hdr[0] if hdr else [])]

                max_cols = len(header_values)
                rows = []
                displayed_count = 0
                last_data_row = start_row - 1
                blank_streak = 0
                MAX_BLANK_STREAK = 10
                # 출력 파일(run_output)은 하나라도 값 있으면 데이터 행
                # 입력 파일은 헤더 열 과반수 기준 (no 열 등 자동번호만 있는 행 제외)
                header_col_count = max(len(header_values), 1)
                is_output = (kind == "run_output")
                for row_idx, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=start_row):
                    vals = ["" if v is None else str(v) for v in row]
                    max_cols = max(max_cols, len(vals))
                    if is_output:
                        has_data = any(v.strip() for v in vals)
                    else:
                        header_vals = vals[:header_col_count]
                        filled = sum(1 for v in header_vals if v.strip())
                        has_data = filled > (header_col_count / 2)
                    if has_data:
                        last_data_row = row_idx
                        blank_streak = 0
                    else:
                        blank_streak += 1
                        if blank_streak >= MAX_BLANK_STREAK:
                            break
                    if displayed_count < self.MAX_PREVIEW_ROWS:
                        rows.append(vals)
                        displayed_count += 1
                # 실제 데이터 행 기준으로 자르기 — 뒤쪽 빈/반빈 행 제거
                if last_data_row >= start_row:
                    rows = rows[:last_data_row - start_row + 1]
                    displayed_count = len(rows)
                actual_count = max(0, last_data_row - start_row + 1)
                max_row = last_data_row

                columns = header_values if header_values else [""] * max_cols
                if len(columns) < max_cols:
                    columns += [""] * (max_cols - len(columns))
                rows = [r + [""] * (max_cols - len(r)) for r in rows]
                row_marks = {
                    'warn_rows': _as_abs_issue_rows(issue_rows, data_start),
                    'error_rows': [],
                    'issue_rows': _as_abs_issue_rows(issue_rows, data_start),
                    'muted_rows': list(range(start_row, max(data_start, start_row) )),
                }
            finally:
                wb.close()

            self.finished.emit(json.dumps({
                "ok": True, "kind": kind,
                "columns": columns, "rows": rows,
                "displayed_count": displayed_count,
                "actual_count": actual_count,
                "total_count": actual_count,
                "max_row": max_row,
                "truncated": bool(actual_count > displayed_count),
                "source_file": Path(file_path).name,
                "sheet_name": actual_sheet,
                "header_row": header_row,
                "data_start_row": data_start,
                "start_row": start_row,
                "issue_rows": _as_abs_issue_rows(issue_rows, data_start),
                "row_marks": row_marks,
            }, ensure_ascii=False))

        except Exception as e:
            tb = tb_module.format_exc()
            _dbg("PreviewWorker error:", str(e), "\n", tb)
            self.failed.emit(json.dumps({
                "ok": False, "kind": kind,
                "error": str(e), "traceback": tb,
            }, ensure_ascii=False))


# ──────────────────────────────────────────────
# Bridge
# ──────────────────────────────────────────────

class Bridge(QObject):
    """
    WebView ↔ Python 연결 객체.
    QWebChannel에 등록하여 JS에서 직접 호출한다.
    """

    # Python -> JS 시그널
    scanFinished     = pyqtSignal(str)
    scanFailed       = pyqtSignal(str)
    runFinished      = pyqtSignal(str)
    runFailed        = pyqtSignal(str)
    diffScanFinished = pyqtSignal(str)
    diffScanFailed   = pyqtSignal(str)
    diffRunFinished  = pyqtSignal(str)
    diffRunFailed    = pyqtSignal(str)
    previewLoaded    = pyqtSignal(str)
    previewFailed    = pyqtSignal(str)

    def __init__(self, parent=None):
        super().__init__(parent)

        # 중복 실행 방지 플래그
        self._is_scanning      = False
        self._is_running       = False
        self._is_diff_scanning = False
        self._is_diff_running  = False
        self._is_previewing    = False

        # 원본 결과 보관 (JS로 직접 노출 금지)
        # run_main_engine이 ScanResult를 직접 받으므로 Bridge에서 보관
        self._last_scan_result = None   # ScanResult (run_main_engine 인자용)
        self._last_run_result  = None   # PipelineResult
        self._current_output_files: list = []
        self._school_name_set: set = set()

        # Worker / Thread 참조 (GC 방지)
        self._scan_thread = None;       self._scan_worker = None
        self._run_thread  = None;       self._run_worker  = None
        self._diff_scan_thread = None;  self._diff_scan_worker = None
        self._diff_run_thread  = None;  self._diff_run_worker  = None
        self._preview_thread   = None;  self._preview_worker   = None

    def _start_worker(self, worker, thread, on_finished, on_failed):
        """Worker/Thread 연결 공통 처리"""
        worker.moveToThread(thread)
        thread.started.connect(worker.run)
        worker.finished.connect(on_finished)
        worker.failed.connect(on_failed)
        worker.finished.connect(thread.quit)
        worker.failed.connect(thread.quit)
        thread.finished.connect(thread.deleteLater)
        thread.start()

    # ──────────────────────────────────────────
    # A. 조회 계열 (동기)
    # ──────────────────────────────────────────

    @pyqtSlot(str, result=str)
    def inspectWorkRoot(self, work_root: str) -> str:
        try:
            result = inspect_work_root(work_root)
            serializable = {
                k: str(v) if isinstance(v, Path) else v
                for k, v in result.items()
            }
            return ok_response(serializable)
        except Exception as e:
            return error_response(str(e))

    @pyqtSlot(str, str, result=str)
    def loadSchoolNames(self, roster_xlsx: str, col_map_json: str) -> str:
        """
        roster_xlsx: 명단 파일 경로 (work_root 아님)
        col_map_json: JSON 문자열 또는 "{}"
        """
        try:
            col_map = json.loads(col_map_json) if col_map_json else {}
        except Exception:
            col_map = {}
        try:
            names = load_all_school_names(
                roster_xlsx=Path(roster_xlsx) if roster_xlsx else None,
                col_map=col_map or None,
            )
            self._school_name_set = set(names)
            return ok_response({"school_names": names})
        except Exception as e:
            return error_response(str(e))

    @pyqtSlot(str, str, str, result=str)
    def getSchoolDomain(self, roster_xlsx: str, school_name: str, col_map_json: str) -> str:
        try:
            col_map = json.loads(col_map_json) if col_map_json else {}
        except Exception:
            col_map = {}
        try:
            domain = get_school_domain(
                roster_xlsx=Path(roster_xlsx) if roster_xlsx else None,
                school_name=school_name,
                col_map=col_map or None,
            )
            return ok_response({"domain": domain or ""})
        except Exception as e:
            return error_response(str(e))

    @pyqtSlot(str, result=str)
    def getProjectDirs(self, work_root: str) -> str:
        """get_project_dirs(work_root) — school_name 파라미터 없음"""
        try:
            dirs = get_project_dirs(Path(work_root))
            return ok_response({"dirs": {k: str(v) for k, v in dirs.items()}})
        except Exception as e:
            return error_response(str(e))

    @pyqtSlot(str, result=str)
    def loadNoticeTemplates(self, work_root: str) -> str:
        try:
            templates = load_notice_templates(Path(work_root))
            return ok_response({"templates": templates})
        except Exception as e:
            return error_response(str(e))

    @pyqtSlot(result=str)
    def loadAppConfig(self) -> str:
        try:
            return ok_response({"config": _load_app_config()})
        except Exception as e:
            return error_response(str(e))

    @pyqtSlot(str, result=str)
    def loadWorkHistory(self, school_year: str) -> str:
        """
        school_year: 연도 문자열 (예: "2025")
        반환: { history: { school_name: { last_date, worker, counts } } }
        """
        try:
            year = int(school_year)
        except (ValueError, TypeError):
            return error_response("잘못된 학년도 형식입니다")
        try:
            history = _load_work_history(year)
            return ok_response({"history": history})
        except Exception as e:
            return error_response(str(e))

    @pyqtSlot(str, str, str, result=str)
    def saveWorkHistory(self, school_year: str, school_name: str, entry_json: str) -> str:
        """
        school_year: 연도 문자열
        school_name: 학교명
        entry_json: { last_date, worker, counts } JSON 문자열
        """
        try:
            year  = int(school_year)
            entry = json.loads(entry_json)
        except Exception:
            return error_response("잘못된 파라미터 형식입니다")
        try:
            _save_work_history(year, school_name, entry)
            return ok_response({})
        except Exception as e:
            return error_response(str(e))

    # ──────────────────────────────────────────
    # B. 저장 계열 (동기)
    # ──────────────────────────────────────────

    @pyqtSlot(str, result=str)
    def saveAppConfig(self, config_json: str) -> str:
        try:
            config = json.loads(config_json)
        except Exception:
            return error_response("잘못된 파라미터 형식입니다")
        try:
            _save_app_config(config)
            return ok_response({})
        except Exception as e:
            return error_response(str(e))

    @pyqtSlot(str, result=str)
    def writeWorkResult(self, params_json: str) -> str:
        """
        params: {
          xlsx_path, school_name, worker,
          kind_flags: {"신입생": bool, ...},
          email_arrived_date: "YYYY-MM-DD" | "",
          col_map: {...},
          seq_no: int | null
        }
        """
        try:
            p = json.loads(params_json)
        except Exception:
            return error_response("잘못된 파라미터 형식입니다")
        try:
            from datetime import datetime as _dt
            arr_date = None
            raw = p.get("email_arrived_date", "")
            if raw and _validate_date(raw):
                arr_date = _dt.strptime(raw, "%Y-%m-%d").date()

            ok, msg = write_work_result(
                xlsx_path=Path(p["xlsx_path"]),
                school_name=p["school_name"],
                worker=p.get("worker", ""),
                kind_flags=p.get("kind_flags", {}),
                email_arrived_date=arr_date,
                col_map=p.get("col_map"),
                seq_no=p.get("seq_no"),
            )
            return ok_response({"message": msg}) if ok else error_response(msg)
        except Exception as e:
            return error_response(str(e))

    @pyqtSlot(str, result=str)
    def writeEmailSent(self, params_json: str) -> str:
        """
        params: {
          xlsx_path, school_name,
          sent_date: "YYYY-MM-DD" | "",
          col_map: {...}
        }
        """
        try:
            p = json.loads(params_json)
        except Exception:
            return error_response("잘못된 파라미터 형식입니다")
        try:
            from datetime import datetime as _dt
            sent_date = None
            raw = p.get("sent_date", "")
            if raw and _validate_date(raw):
                sent_date = _dt.strptime(raw, "%Y-%m-%d").date()

            ok, msg = write_email_sent(
                xlsx_path=Path(p["xlsx_path"]),
                school_name=p["school_name"],
                sent_date=sent_date,
                col_map=p.get("col_map"),
            )
            return ok_response({"message": msg}) if ok else error_response(msg)
        except Exception as e:
            return error_response(str(e))

    # ──────────────────────────────────────────
    # C. 비동기 시작 계열
    # ──────────────────────────────────────────

    @pyqtSlot(str, result=str)
    def startScanMain(self, params_json: str) -> str:
        _dbg("startScanMain called", params_json)
        """
        params: {
          work_root, school_name,
          school_start_date (YYYY-MM-DD), work_date (YYYY-MM-DD),
          roster_xlsx (optional), roster_basis_date (optional),
          col_map (optional)
        }
        완료 -> scanFinished(payload) / 실패 -> scanFailed(payload)
        """
        try:
            params = json.loads(params_json)
        except Exception:
            return error_response("잘못된 파라미터 형식입니다")

        if not params.get("work_root"):
            return error_response("작업 폴더가 없습니다")
        if not params.get("school_name"):
            return error_response("학교가 선택되지 않았습니다")
        if not _validate_date(params.get("school_start_date", "")):
            return error_response("개학일 형식이 올바르지 않습니다 (YYYY-MM-DD)")
        if not _validate_date(params.get("work_date", "")):
            return error_response("작업일 형식이 올바르지 않습니다 (YYYY-MM-DD)")
        if self._is_scanning:
            return error_response("이미 스캔이 진행 중입니다")

        self._is_scanning = True
        self._last_scan_result = None

        worker = ScanWorker(params)
        thread = QThread()
        self._scan_worker = worker
        self._scan_thread = thread
        _dbg("startScanMain starting worker thread")
        self._start_worker(worker, thread, self._on_scan_finished, self._on_scan_failed)
        return ok_response({})

    def _on_scan_finished(self, payload: str):
        _dbg("_on_scan_finished entered")
        self._is_scanning = False
        # Worker의 scan_result 원본을 Bridge에 보관 (run_main_engine 인자용)
        if self._scan_worker is not None:
            self._last_scan_result = getattr(self._scan_worker, "scan_result", None)
        self.scanFinished.emit(payload)
        _dbg("_on_scan_finished emitted")

    def _on_scan_failed(self, payload: str):
        _dbg("_on_scan_failed entered")
        self._is_scanning = False
        self._last_scan_result = None
        self.scanFailed.emit(payload)
        _dbg("_on_scan_failed emitted")

    # ── Run ────────────────────────────────────

    @pyqtSlot(str, result=str)
    def startRunMain(self, params_json: str) -> str:
        """
        run_main_engine은 ScanResult 객체를 직접 받는다.
        Bridge._last_scan_result (scanFinished 이후 자동 저장됨) 를 사용.

        params: {
          work_date (YYYY-MM-DD), school_start_date (YYYY-MM-DD),
          layout_overrides (optional), school_kind_override (optional)
        }
        완료 -> runFinished(payload) / 실패 -> runFailed(payload)
        """
        try:
            params = json.loads(params_json)
        except Exception:
            return error_response("잘못된 파라미터 형식입니다")

        if self._last_scan_result is None:
            return error_response("스캔 결과가 없습니다. 먼저 스캔을 실행해 주세요.")
        if not getattr(self._last_scan_result, "ok", False):
            return error_response("스캔이 실패 상태입니다. 스캔을 다시 실행해 주세요.")
        if not _validate_date(params.get("work_date", "")):
            return error_response("작업일 형식이 올바르지 않습니다 (YYYY-MM-DD)")
        if not _validate_date(params.get("school_start_date", "")):
            return error_response("개학일 형식이 올바르지 않습니다 (YYYY-MM-DD)")
        if self._is_running:
            return error_response("이미 실행이 진행 중입니다")

        self._is_running = True

        worker = RunWorker(
            scan_result=self._last_scan_result,
            work_date=params["work_date"],
            school_start_date=params["school_start_date"],
            layout_overrides=params.get("layout_overrides"),
            school_kind_override=params.get("school_kind_override"),
        )
        thread = QThread()
        self._run_worker = worker
        self._run_thread = thread
        self._start_worker(worker, thread, self._on_run_finished, self._on_run_failed)
        return ok_response({})

    def _on_run_finished(self, payload: str):
        self._is_running = False
        self.runFinished.emit(payload)

    def _on_run_failed(self, payload: str):
        self._is_running = False
        self.runFailed.emit(payload)

    # ── Diff Scan ──────────────────────────────

    @pyqtSlot(str, result=str)
    def startScanDiff(self, params_json: str) -> str:
        """
        params: {
          work_root, school_name, target_year (int),
          school_start_date (YYYY-MM-DD), work_date (YYYY-MM-DD),
          roster_xlsx (optional), roster_basis_date (optional), col_map (optional)
        }
        """
        try:
            params = json.loads(params_json)
        except Exception:
            return error_response("잘못된 파라미터 형식입니다")

        if not params.get("work_root"):
            return error_response("작업 폴더가 없습니다")
        if not params.get("school_name"):
            return error_response("학교가 선택되지 않았습니다")
        if not _validate_date(params.get("school_start_date", "")):
            return error_response("개학일 형식이 올바르지 않습니다 (YYYY-MM-DD)")
        if not _validate_date(params.get("work_date", "")):
            return error_response("작업일 형식이 올바르지 않습니다 (YYYY-MM-DD)")
        if self._is_diff_scanning:
            return error_response("이미 명단비교 스캔이 진행 중입니다")

        self._is_diff_scanning = True

        worker = DiffScanWorker(params)
        thread = QThread()
        self._diff_scan_worker = worker
        self._diff_scan_thread = thread
        self._start_worker(worker, thread, self._on_diff_scan_finished, self._on_diff_scan_failed)
        return ok_response({})

    def _on_diff_scan_finished(self, payload: str):
        self._is_diff_scanning = False
        self.diffScanFinished.emit(payload)

    def _on_diff_scan_failed(self, payload: str):
        self._is_diff_scanning = False
        self.diffScanFailed.emit(payload)

    # ── Diff Run ───────────────────────────────

    @pyqtSlot(str, result=str)
    def startRunDiff(self, params_json: str) -> str:
        """
        run_diff_engine은 내부에서 scan을 재실행 — scan 객체 불필요.
        params: {
          work_root, school_name, target_year (int),
          school_start_date (YYYY-MM-DD), work_date (YYYY-MM-DD),
          roster_basis_date (optional)
        }
        """
        try:
            params = json.loads(params_json)
        except Exception:
            return error_response("잘못된 파라미터 형식입니다")

        if not params.get("work_root"):
            return error_response("작업 폴더가 없습니다")
        if not params.get("school_name"):
            return error_response("학교가 선택되지 않았습니다")
        if self._is_diff_running:
            return error_response("이미 명단비교 실행이 진행 중입니다")

        self._is_diff_running = True

        worker = DiffRunWorker(params)
        thread = QThread()
        self._diff_run_worker = worker
        self._diff_run_thread = thread
        self._start_worker(worker, thread, self._on_diff_run_finished, self._on_diff_run_failed)
        return ok_response({})

    def _on_diff_run_finished(self, payload: str):
        self._is_diff_running = False
        self.diffRunFinished.emit(payload)

    def _on_diff_run_failed(self, payload: str):
        self._is_diff_running = False
        self.diffRunFailed.emit(payload)

    # ── Preview ────────────────────────────────

    @pyqtSlot(str, result=str)
    def startPreview(self, params_json: str) -> str:
        """
        params: {
          kind: "freshmen"|"transfer_in"|"transfer_out"|"teachers"|"roster",
          file_path, sheet_name,
          header_row (int), data_start_row (int)
        }
        완료 -> previewLoaded(payload) / 실패 -> previewFailed(payload)
        """
        try:
            params = json.loads(params_json)
        except Exception:
            return error_response("잘못된 파라미터 형식입니다")

        if not params.get("kind"):
            return error_response("미리보기 종류가 지정되지 않았습니다")
        if not params.get("file_path"):
            return error_response("파일 경로가 없습니다")
        if self._is_previewing:
            return error_response("이미 미리보기가 로딩 중입니다")

        self._is_previewing = True

        worker = PreviewWorker(params)
        thread = QThread()
        self._preview_worker = worker
        self._preview_thread = thread
        self._start_worker(worker, thread, self._on_preview_loaded, self._on_preview_failed)
        return ok_response({})

    def _on_preview_loaded(self, payload: str):
        self._is_previewing = False
        self.previewLoaded.emit(payload)

    def _on_preview_failed(self, payload: str):
        self._is_previewing = False
        self.previewFailed.emit(payload)

    # ──────────────────────────────────────────
    # D. OS 연동 계열
    # ──────────────────────────────────────────

    @pyqtSlot(result=str)
    def pickWorkFolder(self) -> str:
        path = QFileDialog.getExistingDirectory(None, "작업 폴더 선택")
        return ok_response({"path": path or ""})

    @pyqtSlot(result=str)
    def pickRosterLogFile(self) -> str:
        path, _ = QFileDialog.getOpenFileName(
            None, "명단 파일 선택", "", "Excel 파일 (*.xlsx)"
        )
        return ok_response({"path": path or ""})

    @pyqtSlot(str, str, int, result=str)
    def readXlsxMeta(self, xlsx_path: str, sheet_name: str, header_row: int) -> str:
        """
        열 매핑 다이얼로그용 — xlsx 파일 시트 목록 + 지정 시트의 미리보기 반환.

        xlsx_path:  파일 경로
        sheet_name: 빈 문자열이면 첫 시트
        header_row: 헤더 행 번호 (1-based)

        반환:
          { sheets: [...], headers: [...], rows: [[...], ...] }
        """
        try:
            from openpyxl import load_workbook as _load_wb
            wb = _load_wb(str(xlsx_path), read_only=True, data_only=True)

            sheets = wb.sheetnames
            target = sheet_name if (sheet_name and sheet_name in sheets) else sheets[0]
            ws = wb[target]

            h_idx = max(0, header_row - 1)   # 0-based
            rows_raw = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                rows_raw.append([str(v) if v is not None else "" for v in row])
                if i >= h_idx + 15:           # 헤더 + 데이터 15행까지
                    break
            wb.close()

            headers = rows_raw[h_idx] if h_idx < len(rows_raw) else []
            preview  = rows_raw[h_idx + 1: h_idx + 11]  # 데이터 10행

            return ok_response({
                "sheets":  sheets,
                "sheet":   target,
                "headers": headers,
                "rows":    preview,
            })
        except Exception as e:
            return error_response(str(e))

    @pyqtSlot(str, result=str)
    def openFile(self, path: str) -> str:
        import subprocess, sys, os
        try:
            if sys.platform == "win32":
                os.startfile(path)
            elif sys.platform == "darwin":
                subprocess.run(["open", path])
            else:
                subprocess.run(["xdg-open", path])
            return ok_response({})
        except Exception as e:
            return error_response(str(e))

    @pyqtSlot(str, result=str)
    def openFolder(self, path: str) -> str:
        import subprocess, sys, os
        try:
            target = str(Path(path).parent) if Path(path).is_file() else path
            if sys.platform == "win32":
                os.startfile(target)
            elif sys.platform == "darwin":
                subprocess.run(["open", target])
            else:
                subprocess.run(["xdg-open", target])
            return ok_response({})
        except Exception as e:
            return error_response(str(e))

    @pyqtSlot(str, result=str)
    def copyToClipboard(self, text: str) -> str:
        try:
            QApplication.clipboard().setText(text)
            return ok_response({})
        except Exception as e:
            return error_response(str(e))
