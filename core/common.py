# core/common.py
"""
메인 / diff 파이프라인에서 함께 사용하는 공통 유틸 모듈.

책임 범위:
  - 경로 / 폴더 확인
  - 워크북 안전 로드
  - xlsx 형식 검증
  - 헤더 / 데이터 시작 행 감지 보조
  - 셀 값 정규화 및 공통 파싱 보조
  - 명부 시트 로드 등 공통 처리
  - 입력 파일 종류별 헤더 슬롯 상수 (scan/run 공용)
  - RosterInfo 데이터 구조 정의

특정 실행 로직(run)이나 스캔 로직(scan)에 종속되지 않는 공통 기능만 둔다.
HEADER_SLOTS 상수는 scan_main(헤더 감지)과 run_main(데이터 읽기) 양쪽에서
모두 사용하므로 여기에 정의한다.

공개 API:
  [데이터 구조]
  RosterInfo

  [헤더 슬롯 상수]
  FRESHMEN_HEADER_SLOTS, TRANSFER_HEADER_SLOTS,
  WITHDRAW_HEADER_SLOTS, TEACHER_HEADER_SLOTS

  [경로 / 파일]
  get_project_dirs, ensure_xlsx_only, safe_load_workbook
  get_first_sheet_with_warning, warn_if_multi_sheet

  [헤더 감지]
  normalize_header_cell, header_map
  HANGUL_RE, EN_RE
  _build_header_slot_map, _detect_header_row_generic

  [이름 정규화]
  normalize_name, normalize_name_key

  [명부]
  load_roster_sheet, parse_roster_year_from_filename

  [동명이인 suffix]
  english_casefold_key, dedup_suffix_letters, split_korean_name_suffix
  apply_suffix_for_duplicates
  resolve_transfer_name_conflicts
  _strip_korean_suffix_for_notice, notice_name_key

  [도메인 / 학교 종류]
  parse_class_str, extract_id_prefix4
  school_kind_from_name, load_notice_templates
"""
from __future__ import annotations

import re
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from datetime import datetime
from io import BytesIO

from openpyxl import load_workbook

from core.utils import normalize_text, text_contains

from collections import defaultdict, Counter
from dataclasses import dataclass, field

# =========================
# RosterInfo 데이터 구조
# =========================
@dataclass
class RosterInfo:
    """
    학생명부 분석 결과를 담는 구조체.

    scan_main.analyze_roster_once()에서 생성되고,
    run_main의 전입 ID 생성(build_transfer_ids)과
    전출 매칭(build_withdraw_outputs)에서 소비된다.

    필드:
      roster_time:
        명부의 학년도 추정값. 아이디 패턴으로 자동 감지하되,
        최종값은 명부 기준일과 개학일 비교로 결정된다.
          "this_year" — 올해 학년도 명부 (개학일 이후 기준일)
          "last_year" — 작년 학년도 명부 (개학일 이전 기준일)
          "manual"    — UI에서 사용자가 학년도를 직접 지정한 경우
          "unknown"   — 판정 불가

      ref_grade_shift:
        명부의 학년 번호와 현재 실제 학년 번호 사이의 차이.
          0  — 올해 명부: 명부 학년 = 현재 학년 (개학 이후)
          -1 — 작년 명부: 명부 학년 = 현재 학년 - 1 (개학 이전)
        사용 예: 현재 3학년 전입생의 ID prefix를 찾을 때
          g_roster = g_current + shift → prefix_mode[g_roster]로 조회

      prefix_mode_by_roster_grade:
        {명부학년(int): 최빈 ID prefix 4자리(int)}
        명부의 학년별로 학생 아이디 앞 4자리(입학년도)의 최빈값.
        전입생 ID 생성 시 "이 학년 학생들은 보통 YYYY년에 입학했다"는
        기준값으로 사용한다.
        키는 명부 기준 학년이므로 현재 학년과 다를 수 있다(ref_grade_shift 참고).
        예: {1: 2024, 2: 2023, 3: 2022, ...}

      name_count_by_roster_grade:
        {명부학년(int): Counter({정규화이름: 등장횟수})}
        명부의 학년별 이름 빈도. 주로 동명이인 여부 확인에 사용.

      roster_names_by_grade:
        {명부학년(int): [이름(str), ...]}
        명부의 학년별 이름 목록 (suffix 포함).
        전입생 동명이인 suffix 결정 시 기존 점유 suffix 계산에 사용.
    """
    roster_time: str = "unknown"
    ref_grade_shift: int = 0
    prefix_mode_by_roster_grade: Dict[int, int] = field(default_factory=dict)
    name_count_by_roster_grade: Dict[int, Any] = field(default_factory=dict)
    roster_names_by_grade: Dict[int, List[str]] = field(default_factory=dict)




# =========================
# 입력 파일 헤더 슬롯 상수
# =========================
# scan_main(헤더 자동 감지)과 run_main(데이터 읽기) 양쪽에서 공유하므로
# 어느 한쪽에 두지 않고 common에 정의한다.
# 슬롯 이름(key)은 파이프라인 전체에서 열을 참조하는 표준 식별자이므로
# 변경 시 scan_main / run_main 양쪽의 slot_cols.get(...) 호출도 함께 확인할 것.

FRESHMEN_HEADER_SLOTS: Dict[str, List[str]] = {
    # 슬롯: [매칭할 헤더 키워드 목록] (부분 일치)
    "no":    ["no", "번호"],
    "grade": ["학년"],
    "class": ["반", "학급"],
    "num":   ["번호", "번"],        # 출석번호 (no와 별개)
    "name":  ["성명", "이름", "학생이름"],
}

TRANSFER_HEADER_SLOTS: Dict[str, List[str]] = {
    "no":     ["no", "번호"],
    "grade":  ["학년"],
    "class":  ["반", "학급"],
    "number": ["번호", "번", "출석번호"],
    "name":   ["성명", "이름"],
    "remark": ["비고", "메모", "특이사항"],
}

WITHDRAW_HEADER_SLOTS: Dict[str, List[str]] = {
    "no":     ["no", "번호"],
    "grade":  ["학년"],
    "class":  ["반", "학급"],
    "name":   ["성명", "이름"],
    "remark": ["비고", "메모", "특이사항"],
}

TEACHER_HEADER_SLOTS: Dict[str, List[str]] = {
    "no":       ["no", "번호"],
    "position": ["직위", "담당", "직위담당", "직책"],
    "name": [
        "성명", "이름", "성함",
        "교사명", "교사이름", "교사성명",
        "교원명", "교원이름", "교원성명",
        "교직원명", "교직원이름", "교직원성명",
        "선생님이름", "선생님성명", "선생님명",
        "담당자명", "담당자이름",
    ],
    "learn": ["학습용id신청", "학습용id", "학습용", "학습용아이디"],
    "admin": ["관리용id신청", "관리용id", "관리용", "관리용아이디"],
}


# =========================
# 경로 / 파일 / 워크북
# =========================
def get_project_dirs(work_root: Path) -> Dict[str, Path]:
    """
    work_root/
      resources/  (이름에 'resources' 포함된 폴더 1개)
        DB/, templates/, notices/
      A초등학교/, B중학교/, ...
    """
    work_root = work_root.resolve()

    candidates = [
        p for p in work_root.iterdir()
        if p.is_dir() and "resources" in p.name.lower()
    ]

    if len(candidates) == 0:
        resources_root = work_root / "resources"
    elif len(candidates) == 1:
        resources_root = candidates[0]
    else:
        names = [p.name for p in candidates]
        raise ValueError(
            f"[ERROR] 작업 폴더 내에 'resources' 폴더가 여러 개 있습니다: {names}"
        )

    return {
        "WORK_ROOT":      work_root,
        "RESOURCES_ROOT": resources_root,
        "TEMPLATES":      resources_root / "templates",
        "NOTICES":        resources_root / "notices",
        "SCHOOL_ROOT":    work_root,
    }


def ensure_xlsx_only(p: Path) -> None:
    """파일 확장자가 .xlsx가 아니면 ValueError를 발생시킨다."""
    if p.suffix.lower() != ".xlsx":
        raise ValueError(f"[ERROR] .xlsx 파일이 아닙니다: {p.name}")


def safe_load_workbook(xlsx_path: Path, data_only: bool = True, read_only: bool = False):
    """
    openpyxl로 워크북을 안전하게 로드한다.
    read_only=True가 필요한 경우 명시적으로 전달할 것.
    ws.cell() 랜덤 접근이 필요한 곳에서는 반드시 read_only=False(기본값) 사용.

    두 가지 알려진 openpyxl 버그를 자동으로 우회한다:
      1. docProps/custom.xml에 name 속성 없는 property가 있는 경우 (TypeError)
         → zip을 직접 열어 빈 name 속성 property를 제거한 뒤 재로드
      2. 스타일 인덱스 오류 (IndexError)
         → read_only=True 모드로 재시도 (스타일 파싱 건너뜀)
    """
    try:
        return load_workbook(xlsx_path, data_only=data_only, read_only=read_only)
    except TypeError as e:
        msg = str(e)
        if "openpyxl.packaging.custom" not in msg or "NoneType" not in msg:
            raise

        buffer = BytesIO()
        with zipfile.ZipFile(xlsx_path, "r") as zin, zipfile.ZipFile(
            buffer, "w", compression=zipfile.ZIP_DEFLATED
        ) as zout:
            for item in zin.infolist():
                if item.filename == "docProps/custom.xml":
                    root = ET.fromstring(zin.read(item.filename))
                    ns  = "http://schemas.openxmlformats.org/officeDocument/2006/custom-properties"
                    tag = f"{{{ns}}}property"
                    for prop in list(root.findall(tag)):
                        name = prop.get("name")
                        if name is None or str(name).strip() == "":
                            root.remove(prop)
                    new_xml = ET.tostring(root, encoding="utf-8", xml_declaration=True)
                    zout.writestr(item, new_xml)
                else:
                    zout.writestr(item, zin.read(item.filename))

        buffer.seek(0)
        return load_workbook(buffer, data_only=data_only, read_only=read_only)

    except IndexError:
        if read_only:
            raise
        # 스타일 인덱스 오류 — read_only 모드로 재시도 (스타일 파싱 건너뜀)
        return load_workbook(xlsx_path, data_only=data_only, read_only=True)


def get_first_sheet_with_warning(
    wb,
    file_name: str,
    logs: Optional[List[str]] = None,
):
    sheets = wb.worksheets
    if not sheets:
        raise ValueError(f"[ERROR] 파일에 시트가 없습니다: {file_name}")

    if len(sheets) > 1 and logs is not None:
        logs.append(
            f"[WARN] '{file_name}' 파일에 시트가 {len(sheets)}개 있습니다. "
            f"첫 번째 시트('{sheets[0].title}')만 사용합니다."
        )

    return sheets[0]


def warn_if_multi_sheet(
    xlsx_path: Optional[Path],
    logs: List[str],
    label: str = "",
) -> None:
    if xlsx_path is None:
        return
    try:
        wb = safe_load_workbook(xlsx_path, data_only=True)
        try:
            get_first_sheet_with_warning(wb, xlsx_path.name, logs)
        finally:
            wb.close()
    except Exception:
        pass


# =========================
# 헤더 감지
# =========================
def normalize_header_cell(val: Any) -> str:
    """
    헤더 셀 값을 비교용 정규화 문자열로 변환한다.
    공백(nbsp 포함) 제거, 점(.) 제거, 소문자 변환.
    header_map / _build_header_slot_map 내부에서 키 생성 시 사용한다.
    """
    if val is None:
        s = ""
    else:
        s = str(val)
    s = s.replace("\u00A0", " ")
    s = re.sub(r"\s+", "", s)
    s = s.replace(".", "")
    s = s.lower()
    return s


def header_map(ws, header_row: int = 1) -> Dict[str, int]:
    """헤더 행을 읽어 {정규화된 헤더명: 열번호} dict를 반환한다."""
    mapping: Dict[str, int] = {}
    for cell in ws[header_row]:
        if cell.value is None:
            continue
        key = str(cell.value)
        key = key.replace("\u00A0", " ")
        key = re.sub(r"\s+", "", key)
        key = key.replace(".", "")
        key = key.lower()
        mapping[key] = cell.column
    return mapping


def _build_header_slot_map(
    ws,
    header_row: int,
    slots: Dict[str, List[str]],
) -> Dict[str, int]:
    """
    slots 정의를 기준으로 헤더 행에서 각 slot의 컬럼 위치를 찾아 반환.
    반환: {slot_name: column_index}
    """
    hm = header_map(ws, header_row)

    norm_to_col: Dict[str, int] = {}
    for raw_key, col in hm.items():
        norm_key = normalize_header_cell(raw_key)
        if norm_key:
            norm_to_col[norm_key] = col

    result: Dict[str, int] = {}

    for slot, patterns in slots.items():
        for pat in patterns:
            pat_norm = normalize_header_cell(pat)
            if not pat_norm:
                continue
            for header_norm, col in norm_to_col.items():
                if pat_norm in header_norm:
                    result[slot] = col
                    break
            if slot in result:
                break

    return result


def _detect_header_row_generic(
    ws,
    slots: Dict[str, List[str]],
    max_search_row: int = 15,
    max_col: int = 10,
    min_match_slots: int = 3,
) -> int:
    """
    한 행에서 slots 중 min_match_slots 개 이상이 매칭되면 헤더 행으로 판정.
    """
    best_row:   Optional[int] = None
    best_score: int = 0

    for row in ws.iter_rows(min_row=1, max_row=max_search_row):
        row_idx = row[0].row
        vals = [normalize_header_cell(c.value) for c in row[:max_col]]

        matched_slots = set()
        for slot, patterns in slots.items():
            for pat in patterns:
                pat_norm = normalize_header_cell(pat)
                if not pat_norm:
                    continue
                if any(pat_norm in v for v in vals if v):
                    matched_slots.add(slot)
                    break

        score = len(matched_slots)
        if score > best_score:
            best_score = score
            best_row   = row_idx

    if best_row is None or best_score < min_match_slots:
        raise ValueError(
            "[ERROR] 파일에서 헤더를 찾을 수 없습니다."
        )

    return best_row


# =========================
# 이름 정규화
# =========================
HANGUL_RE = re.compile(r"[가-힣]")
EN_RE     = re.compile(r"[A-Za-z]")


def normalize_compare_name(raw: Any) -> str:
    """
    비교(diff) 모드용 이름 표시값 정리.
    - 원문을 최대한 유지한다.
    - 앞뒤 공백 제거
    - 중간 연속 공백만 1칸으로 정리
    - 한글/영문/대문자 suffix는 그대로 둔다.
    """
    if raw is None:
        return ""
    s = str(raw).replace("　", " ").replace(" ", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def normalize_compare_name_key(raw: Any) -> str:
    """
    비교(diff) 모드용 이름 키.
    - 원문 기반 비교를 우선한다.
    - 공백만 제거하고 나머지 문자는 유지한다.
    - 영문은 casefold 처리만 하여 대소문자 차이만 무시한다.
    """
    s = normalize_compare_name(raw)
    if not s:
        return ""
    return re.sub(r"\s+", "", s).casefold()


def get_compare_name_warnings(raw: Any) -> List[str]:
    """
    비교(diff) 모드용 이름 형식 경고.
    자동 수정은 하지 않고, 원문 기준 비교 후 사용자에게 확인만 요청한다.
    """
    s = normalize_compare_name(raw)
    if not s:
        return []

    warnings: List[str] = []
    has_ko = bool(HANGUL_RE.search(s))
    has_en = bool(EN_RE.search(s))

    if has_ko and has_en:
        warnings.append("이름 형식 확인 필요 — 한글과 영문이 함께 있습니다. 원문 기준으로 비교합니다.")

    if re.search(r"\s", s):
        warnings.append("이름 형식 확인 필요 — 이름에 공백이 있습니다. 공백만 정리해 비교합니다.")

    if has_ko and re.search(r"[A-Z]+$", s):
        warnings.append("이름 형식 확인 필요 — 이름 끝 영문 구분자가 있습니다. 원문 기준으로 비교합니다.")

    return warnings


def normalize_name(raw: Any) -> str:
    """
    학생/교사 이름을 표시용으로 정규화한다.
    숫자·특수문자 제거, 한글은 공백 제거, 영어는 단어별 첫 글자 대문자 처리.
    동명이인 suffix(A, B, C...)가 붙어 있으면 그대로 유지된다.
    """
    if raw is None:
        return ""
    s = str(raw).strip()
    s = re.sub(r"[^A-Za-z가-힣\s]", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    if not s:
        return ""

    has_ko = bool(HANGUL_RE.search(s))
    has_en = bool(EN_RE.search(s))

    if has_ko and not has_en:
        return s.replace(" ", "")

    if has_en and not has_ko:
        parts = [p for p in s.split(" ") if p]
        parts = [p.lower().capitalize() for p in parts]
        return "".join(parts)

    if has_ko and has_en:
        def _fix_en(m: re.Match) -> str:
            tok = m.group(0).lower()
            return tok[0].upper() + tok[1:] if tok else tok
        s2 = re.sub(r"[A-Za-z]+", _fix_en, s)
        return s2.replace(" ", "")

    return ""


def normalize_name_key(raw: Any) -> str:
    """
    이름을 비교(매칭)용 키로 변환한다.
    숫자·특수문자·공백 제거 후 casefold(대소문자 무시).
    동명이인 suffix(A, B, C...)도 포함한 채로 키를 만든다.
    suffix 제거가 필요한 경우는 _strip_korean_suffix_for_notice() 또는
    split_korean_name_suffix()를 별도로 사용한다.
    """
    if raw is None:
        return ""
    s = str(raw).strip()
    s = re.sub(r"[^A-Za-z가-힣\s]", "", s)
    s = re.sub(r"\s+", "", s)
    return s.casefold()


# =========================
# DB / 로스터
# =========================
# DB / 명부
# =========================
def parse_roster_year_from_filename(roster_path: Path) -> Optional[int]:
    """
    학생명부 파일명에서 학년도(4자리 연도)를 추출한다.
    '2024학년도_학생명부.xlsx' → 2024
    파일명에 연도가 없으면 None을 반환한다.
    """
    stem = roster_path.stem
    s = stem.replace("\u00A0", " ")
    s = re.sub(r"\s+", " ", s).strip()

    m = re.search(r"(\d{4})\s*학\s*년도", s)
    if m:
        return int(m.group(1))

    m2 = re.search(r"(19\d{2}|20\d{2})", s)
    if m2:
        return int(m2.group(1))

    return None


def load_roster_sheet(dirs: Dict[str, Path], school_name: str):
    """
    학교 폴더에서 학생명부(.xlsx, 파일명에 '학생명부' 포함)를 찾아
    (wb, ws, roster_path, roster_year) 튜플로 반환한다.

    탐색 규칙:
    - dirs["SCHOOL_ROOT"] 하위에서 school_name이 포함된 폴더를 찾는다.
    - 해당 폴더 안에서 파일명에 '학생명부'가 포함된 .xlsx 파일을 찾는다.
    - 여러 개면 최근 수정일 순으로 첫 번째를 사용한다.

    반환: (Workbook, Worksheet, Path, Optional[int])
      마지막 int는 파일명 파싱으로 추출한 학년도 (없으면 None)
    """
    root_dir = dirs["SCHOOL_ROOT"]

    kw = (school_name or "").strip()
    if not kw:
        raise ValueError("[ERROR] 학교명을 입력해 주세요.")

    matches = [
        p for p in root_dir.iterdir()
        if p.is_dir() and text_contains(p.name, kw)
    ]

    if not matches:
        raise ValueError(
            f"[ERROR] 학생명부를 찾을 학교 폴더를 찾을 수 없습니다. "
            f"(작업 폴더 내 '{school_name}' 폴더가 없습니다)"
        )

    if len(matches) > 1:
        raise ValueError(
            "[ERROR] 학생명부를 찾을 학교 폴더 후보가 여러 개입니다: "
            + ", ".join(p.name for p in matches)
        )

    school_root = matches[0]

    import unicodedata as _ud
    candidates: List[Path] = [
        p for p in school_root.iterdir()
        if p.is_file()
        and p.suffix.lower() == ".xlsx"
        and "학생명부" in _ud.normalize('NFC', p.stem)
        and not p.name.startswith("~$")
    ]
    if not candidates:
        raise ValueError("[ERROR] 학생명부 파일을 찾을 수 없습니다. 파일명에 '학생명부'가 포함되어야 합니다.")

    candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    roster_path = candidates[0]

    wb = safe_load_workbook(roster_path, data_only=True)
    ws = get_first_sheet_with_warning(wb, roster_path.name)
    roster_year = parse_roster_year_from_filename(roster_path)

    return wb, ws, roster_path, roster_year


# =========================
# 동명이인 suffix 헬퍼
# =========================

def english_casefold_key(name: str) -> str:
    """영문 이름 비교용 키: casefold(대소문자 무시)."""
    if name is None:
        return ""
    return str(name).strip().casefold()


def dedup_suffix_letters(n: int) -> str:
    """
    동명이인 구분을 위한 suffix 문자열을 1-based 인덱스로 생성한다.
    1 → "A", 2 → "B", ..., 26 → "Z", 27 → "AA", 28 → "AB", ...
    """
    if n <= 0:
        return ""
    out = ""
    while n > 0:
        n -= 1
        out = chr(ord("A") + (n % 26)) + out
        n //= 26
    return out

def split_korean_name_suffix(raw_name: Any) -> Tuple[str, int]:
    """
    한글 이름 기준:
    - 김지우   -> ("김지우", 0)
    - 김지우A  -> ("김지우", 1)
    - 김지우B  -> ("김지우", 2)
    - 김지우AA -> ("김지우", 27)

    영어 이름은 suffix 처리 대상 아님.

    이름을 base + suffix로 분리한다.

    예:
        김지우 → ("김지우", "")
        김지우A → ("김지우", "A")

    규칙:
    - suffix는 이름 끝의 영문 대문자(A-Z)만 인정
    - 나머지는 모두 base로 처리

    """
    name = normalize_name(raw_name)
    if not name:
        return "", 0

    if not HANGUL_RE.search(name):
        return name, 0

    m = re.match(r"^(.*?)([A-Z]+)?$", name)
    if not m:
        return name, 0

    base = (m.group(1) or "").strip()
    sfx = (m.group(2) or "").strip()

    if not sfx:
        return base, 0

    idx = 0
    for ch in sfx:
        idx = idx * 26 + (ord(ch) - ord("A") + 1)
    return base, idx



def resolve_transfer_name_conflicts(
    transfer_rows: List[Dict[str, Any]],
    roster_info: "RosterInfo",
) -> List[Dict[str, Any]]:
    """
    전입생 동명이인 처리: 명부(기존 학생) + 전입생을 함께 고려하여
    각 전입생에게 최종 이름(suffix 포함)과 하이라이트 필요 여부를 반환한다.

    반환 리스트의 각 항목:
      name_out        — suffix가 반영된 최종 이름 (예: 김지우B)
      dup_with_roster — 명부에 동명이인이 있으면 True
      needs_highlight — 등록 파일에서 노란 하이라이트 표시가 필요하면 True
                        (명부 중복 또는 전입생끼리 중복인 경우)

    suffix 부여 규칙:
      - 명부에 이미 같은 이름이 있으면 명부의 마지막 suffix 다음부터 이어받음
          명부: 김지우, 김지우A → 전입생 첫 명: 김지우B
      - 명부에 없고 전입생끼리만 중복이면 A부터 시작
          전입생: 김지우, 김지우 → 김지우A, 김지우B
      - 중복 없으면 suffix 없음

    ref_grade_shift:
      명부학년 = 현재학년 + shift 로 변환하여 명부에서 같은 학년을 찾는다.
      (RosterInfo.ref_grade_shift 참고)
    """
    shift = int(roster_info.ref_grade_shift or 0)
    roster_names_by_grade = roster_info.roster_names_by_grade or {}

    used_suffixes = defaultdict(set)
    total_transfer = defaultdict(int)

    current_grades = {
        int(r["grade"])
        for r in transfer_rows
        if r.get("grade") is not None
    }

    # 명부 suffix 점유 상태 적재
    for g_cur in current_grades:
        g_roster = g_cur + shift
        for roster_name in roster_names_by_grade.get(g_roster, []):
            base_name, suffix_idx = split_korean_name_suffix(roster_name)
            key = (g_cur, normalize_name_key(base_name))
            used_suffixes[key].add(suffix_idx)

    # 전입생 내부 중복 수
    for row in transfer_rows:
        key = (row.get("grade"), normalize_name_key(row.get("name", "")))
        total_transfer[key] += 1

    assigned_count = defaultdict(int)
    results = []

    for row in transfer_rows:
        g_cur = row.get("grade")
        original_name = normalize_name(row.get("name", ""))
        key = (g_cur, normalize_name_key(original_name))

        assigned_count[key] += 1
        dup_total = total_transfer[key]
        occupied = used_suffixes[key]

        if occupied:
            next_idx = max(occupied) + assigned_count[key]
            name_out = original_name + dedup_suffix_letters(next_idx)
            dup_with_roster = True
        elif dup_total <= 1:
            name_out = original_name
            dup_with_roster = False
        else:
            name_out = original_name + dedup_suffix_letters(assigned_count[key])
            dup_with_roster = False

        results.append({
            "name_out": name_out,
            "dup_with_roster": dup_with_roster,
            "needs_highlight": dup_with_roster or dup_total >= 2,
        })

    return results


def apply_suffix_for_duplicates(names: List[str]) -> List[str]:
    from collections import Counter
    total: Dict[str, int] = {}
    for nm in names:
        key = english_casefold_key(nm)
        total[key] = total.get(key, 0) + 1

    seen: Dict[str, int] = {}
    out: List[str] = []
    for nm in names:
        key = english_casefold_key(nm)
        if total.get(key, 0) <= 1:
            out.append(nm)
            continue
        seen[key] = seen.get(key, 0) + 1
        out.append(nm + dedup_suffix_letters(seen[key]))
    return out



def _strip_korean_suffix_for_notice(raw_name: Any) -> str:
    """
    안내파일 동명이인 판정용:
    '김서현A' → '김서현' (한글+대문자 suffix 제거)
    영어 이름은 그대로 유지.
    """
    if raw_name is None:
        return ""
    s = str(raw_name).strip()
    if not s:
        return ""
    has_ko = bool(HANGUL_RE.search(s))
    if has_ko and re.search(r"[A-Z]$", s) and len(s) >= 3:
        s = re.sub(r"[A-Z]+$", "", s).strip()
    return s


def notice_name_key(raw_name: Any) -> str:
    """
    안내파일 동명이인 판정용 최종 키:
    한글+A/B/C suffix 제거 후 normalize_name_key 적용.
    """
    base = _strip_korean_suffix_for_notice(raw_name)
    return normalize_name_key(base)


# =========================
# Roster 파싱 헬퍼
# =========================

def parse_class_str(s: Any) -> Optional[Tuple[int, str]]:
    if s is None:
        return None
    m = re.match(r"^\s*(\d+)\s*-\s*(.+?)\s*$", str(s))
    if not m:
        return None
    return int(m.group(1)), m.group(2).strip()


def extract_id_prefix4(uid: Any) -> Optional[int]:
    if uid is None:
        return None
    s = str(uid).strip()
    if len(s) >= 4 and s[:4].isdigit():
        return int(s[:4])
    return None


# =========================
# Domain 헬퍼
# =========================

def school_profile_from_name(school_name: str) -> Dict[str, Any]:
    s = (school_name or "").strip()
    if not s:
        return {"mode": "unknown", "default_kind": "", "default_prefix": "", "grade_rule_max_grade": 6, "grade_ranges": []}

    if s.endswith("분교"):
        return {
            "mode": "needs_user_choice",
            "default_kind": "",
            "default_prefix": "",
            "grade_rule_max_grade": 6,
            "grade_ranges": [],
        }

    if s.endswith("초중"):
        return {
            "mode": "mixed",
            "default_kind": "",
            "default_prefix": "",
            "grade_rule_max_grade": 9,
            "grade_ranges": [
                (1, 6, "초등부", "초"),
                (7, 9, "중등부", "중"),
            ],
        }

    last = s[-1]
    if last == "초":
        return {
            "mode": "single",
            "default_kind": "초등부",
            "default_prefix": "초",
            "grade_rule_max_grade": 6,
            "grade_ranges": [(1, 6, "초등부", "초")],
        }
    if last == "중":
        return {
            "mode": "single",
            "default_kind": "중등부",
            "default_prefix": "중",
            "grade_rule_max_grade": 3,
            "grade_ranges": [(1, 3, "중등부", "중")],
        }
    if last == "고":
        return {
            "mode": "single",
            "default_kind": "고등부",
            "default_prefix": "고",
            "grade_rule_max_grade": 3,
            "grade_ranges": [(1, 3, "고등부", "고")],
        }

    return {"mode": "unknown", "default_kind": "", "default_prefix": "", "grade_rule_max_grade": 6, "grade_ranges": []}


def apply_school_kind_override(profile: Dict[str, Any], school_kind_override: Optional[str]) -> Dict[str, Any]:
    if not school_kind_override:
        return dict(profile or {})
    override_map = {
        "초등부": ("초등부", "초", 6),
        "중등부": ("중등부", "중", 3),
        "고등부": ("고등부", "고", 3),
        "기타(빈칸)": ("", "", 6),
    }
    kind_full, kind_prefix, max_grade = override_map.get(school_kind_override, ("", "", 6))
    return {
        "mode": "override",
        "default_kind": kind_full,
        "default_prefix": kind_prefix,
        "grade_rule_max_grade": max_grade,
        "grade_ranges": [(1, max_grade, kind_full, kind_prefix)] if kind_full or kind_prefix else [],
    }


def resolve_school_kind_by_grade(profile: Dict[str, Any], grade: Any) -> Tuple[str, str]:
    try:
        grade_i = int(grade)
    except Exception:
        grade_i = None

    for start, end, kind_full, kind_prefix in list((profile or {}).get("grade_ranges", []) or []):
        if grade_i is not None and start <= grade_i <= end:
            return kind_full, kind_prefix

    return (profile or {}).get("default_kind", ""), (profile or {}).get("default_prefix", "")


def school_kind_from_name(school_name: str) -> Tuple[str, str]:
    profile = school_profile_from_name(school_name)
    return profile.get("default_kind", ""), profile.get("default_prefix", "")


# =========================
# 학년별 학년도(ID prefix) 파생
# =========================
def derive_grade_year_map(
    target_grades,
    input_year: int,
    roster_info=None,
) -> Dict[int, int]:
    """
    target_grades(현재학년 기준) 각각에 대해 ID prefix(입학년도 4자리)를 반환.

    우선순위:
      1) roster_info.prefix_mode_by_roster_grade + ref_grade_shift 로 명부 직접값
      2) 없으면 input_year - (g - 1) 로 역산

    반환: {현재학년(int): prefix4(int)}
    """
    result: Dict[int, int] = {}

    shift: int = 0
    if roster_info is not None:
        raw_prefix = (
            getattr(roster_info, "prefix_mode_by_roster_grade", None)
            if not isinstance(roster_info, dict)
            else roster_info.get("prefix_mode_by_roster_grade", {})
        ) or {}
        shift = int((
            getattr(roster_info, "ref_grade_shift", 0)
            if not isinstance(roster_info, dict)
            else roster_info.get("ref_grade_shift", 0)
        ) or 0)
        for g_roster, pref in raw_prefix.items():
            try:
                g_cur = int(g_roster) - shift
                result[g_cur] = int(pref)
            except Exception:
                continue

    # 명부 직접값이 있는 학년 목록 (앵커로 사용)
    known_grades = sorted(result.keys())

    for g in target_grades:
        try:
            g_i = int(g)
        except Exception:
            continue
        if g_i <= 0:
            continue
        if g_i in result:
            continue

        if known_grades:
            # 가장 가까운 명부 학년을 앵커로 삼아 상대 계산
            # 예: 앵커 4학년=2025 → 3학년=2025+1=2026, 5학년=2025-1=2024
            anchor_g = min(known_grades, key=lambda x: abs(x - g_i))
            anchor_pref = result[anchor_g]
            result[g_i] = anchor_pref + (anchor_g - g_i)
        else:
            # 명부가 전혀 없을 때만 절대 역산 (1학년=input_year 기준)
            result[g_i] = (input_year - (g_i - 1)) if input_year else 0

    return result


def load_notice_templates(work_root: Path) -> Dict[str, str]:
    dirs = get_project_dirs(work_root)
    notice_dir = dirs["NOTICES"]

    if not notice_dir.exists():
        return {}

    result: Dict[str, str] = {}
    for p in notice_dir.glob("*.txt"):
        if not p.is_file():
            continue
        try:
            text = p.read_text(encoding="utf-8")
        except UnicodeDecodeError:
            text = p.read_text(encoding="utf-8-sig")
        result[p.stem.strip()] = text.strip()

    return result