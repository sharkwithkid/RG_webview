# core/pipeline.py
from __future__ import annotations

import re
import zipfile
import xml.etree.ElementTree as ET
from dataclasses import dataclass, field
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any, Sequence
from collections import Counter, defaultdict

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border
from openpyxl.worksheet.views import Selection
from pyxlsb import open_workbook as open_xlsb_workbook

from core.utils import normalize_text, text_contains, text_eq



# =========================
# 공식 API (UI / PyQt에서만 사용)
# =========================

__all__ = [
    # dataclasses
    "PipelineResult",
    "ScanResult",
    "LogEntry",
    "PipelineError",

    # high-level API
    "get_project_dirs",
    "scan_work_root",
    "scan_pipeline",
    "execute_pipeline",
    "detect_input_layout",

    # DB / 도메인 조회
    "search_schools_in_db",
    "get_school_domain_from_db",
    "domain_missing_message",

    # 기타 공용 도우미 (UI에서 쓰는 것만)
    "NOTICE_ORDER",
]


# =========================
# 에러 / 로그 타입
# =========================

@dataclass
class LogEntry:
    level: str                    # "INFO" | "WARN" | "ERROR" ...
    code: str                     # "FRESHMEN_MISSING_REQUIRED_COLUMNS" ...
    message: str = ""             # 개발자용 짧은 설명 (UI에서는 안 써도 됨)
    context: Dict[str, Any] = field(default_factory=dict)
    timestamp: datetime = field(default_factory=datetime.now)


class PipelineError(Exception):
    """
    - 파이프라인 내부에서 나는 '의도된' 도메인 오류는 전부 이걸로 던짐.
    - UI 에서는 e.code 를 보고 한글 문구를 매핑.
    """
    def __init__(
        self,
        code: str,
        message: str = "",
        context: Optional[Dict[str, Any]] = None,
    ) -> None:
        self.code = code
        self.message = message or code
        self.context = context or {}
        super().__init__(self.message)



@dataclass
class PipelineResult:
    ok: bool
    outputs: List[Path]
    logs: List[LogEntry]

    transfer_in_done: int = 0
    transfer_in_hold: int = 0
    transfer_out_done: int = 0
    transfer_out_hold: int = 0
    transfer_out_auto_skip: int = 0


@dataclass
class ScanResult:
    # 기본 상태
    ok: bool = False
    logs: List[LogEntry] = field(default_factory=list)

    # 학교/연도 정보
    school_name: str = ""
    year_str: str = ""
    year_int: int = 0

    # 경로들
    project_root: Path = Path(".")
    input_dir: Path = Path(".")
    output_dir: Path = Path(".")
    template_register: Optional[Path] = None
    template_notice: Optional[Path] = None
    db_path: Optional[Path] = None

    # 인풋 파일
    freshmen_file: Optional[Path] = None
    teacher_file: Optional[Path] = None
    transfer_file: Optional[Path] = None
    withdraw_file: Optional[Path] = None

    # 학생명부 관련
    need_roster: bool = False              # 전입/전출 중 하나라도 있으면 True
    roster_path: Optional[Path] = None
    roster_year: Optional[int] = None
    roster_info: Optional[Dict[str, Any]] = None
    roster_basis_date: Optional[date] = None  # 학생명부 기준일(파일 수정일 or 사용자가 수정한 값)

    # UI 플래그
    needs_open_date: bool = False          # 전출 있으면 True → 개학일 필요
    missing_fields: List[str] = field(default_factory=list)
    can_execute: bool = False
    can_execute_after_input: bool = False


FRESHMEN_KEYWORDS = ["신입생", "신입"]
TEACHER_KEYWORDS  = ["교사", "교원"]
TRANSFER_KEYWORDS = ["전입생", "전입"]
WITHDRAW_KEYWORDS = ["전출생", "전출"]



NOTICE_ORDER = [
    "신규등록 - 메일",
    "신규등록 - 문자",
    "교직원 등록 - 메일",
    "반이동 - 메일",
    "반이동 - 메일 (신입생, 교직원 등록 & 반이동)",
    "반이동 - 문자",
    "2-6학년 명단 보내 온 경우 - 메일",
    "2-6학년 반편성 자료 재요청 - 문자",
]


HANGUL_RE = re.compile(r"[가-힣]")
EN_RE = re.compile(r"[A-Za-z]")


# 슬롯별 헤더 키워드 (느슨한 매칭)
FRESHMEN_HEADER_SLOTS = {
    "no":    ["no", "번호"],
    "grade": ["학년"],
    "class": ["반", "학급"],
    "num":   ["번호", "번"],
    "name":  ["성명", "이름", "학생이름"],
}

TRANSFER_HEADER_SLOTS = {
    "no":    ["no", "번호"],
    "grade": ["학년"],
    "class": ["반", "학급"],
    "number":["번호", "번", "출석번호"],
    "name":  ["성명", "이름"],
    "remark":["비고", "메모", "특이사항"],
}

WITHDRAW_HEADER_SLOTS = {
    "no":    ["no", "번호"],
    "grade": ["학년"],
    "class": ["반", "학급"],
    "name":  ["성명", "이름"],
    "remark":["비고", "메모", "특이사항"],
}

TEACHER_HEADER_SLOTS = {
    "no":      ["no", "번호"],
    "position":["직위", "담당", "직위담당"],
    "name":    ["성명", "이름", "선생님이름", "교사명", "교원명"],
    "learn":   ["학습용id신청", "학습용id", "학습용", "학습용아이디"],
    "admin":   ["관리용id신청", "관리용id", "관리용", "관리용아이디"],
}


EXAMPLE_NAMES_RAW = ["홍길동", "이순신", "유관순", "임꺽정"]
EXAMPLE_NAMES_NORM = {normalize_text(n) for n in EXAMPLE_NAMES_RAW}
EXAMPLE_KEYWORDS = ["예시"]  # 행 안 어느 셀이라도 '예시' 포함되면 예시로 처리


FILL_TRANSFER = PatternFill("solid", fgColor="F8CBAD")  # 옅은 주황
FILL_DUP      = PatternFill("solid", fgColor="FFFF00")  # 노랑
FILL_GREY     = PatternFill("solid", fgColor="D9D9D9")  # 회색




# ========== L0: infra / excel utils ==========

def _ensure_xlsx_only(p: Path) -> None:
    if p.suffix.lower() != ".xlsx":
        raise ValueError(f"[오류] 파일 형식이 .xlsx가 아닙니다: {p.name}")

def _backup_if_exists(out_path: Path) -> Optional[Path]:
    """기존 파일이 있으면 작업/_backup으로 이동."""
    out_path = Path(out_path)
    if not out_path.exists():
        return None
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_dir = out_path.parent / "_backup"
    backup_dir.mkdir(parents=True, exist_ok=True)
    dest = backup_dir / f"{out_path.stem}_{ts}{out_path.suffix}"
    out_path.replace(dest)
    return dest

def _safe_load_workbook(xlsx_path: Path, data_only: bool = True):
    try:
        return load_workbook(xlsx_path, data_only=data_only)
    except TypeError as e:
        msg = str(e)
        if "openpyxl.packaging.custom" not in msg or "NoneType" not in msg:
            raise

        buffer = BytesIO()
        with zipfile.ZipFile(xlsx_path, "r") as zin, zipfile.ZipFile(
            buffer, "w", compression=zipfile.ZIP_DEFLATED
        ) as zout:
            for item in zin.infolist():
                if item.filename == "docProps/custom.xml":
                    root = ET.fromstring(zin.read(item.filename))
                    ns = "http://schemas.openxmlformats.org/officeDocument/2006/custom-properties"
                    tag = f"{{{ns}}}property"
                    for prop in list(root.findall(tag)):
                        name = prop.get("name")
                        if name is None or str(name).strip() == "":
                            root.remove(prop)
                    new_xml = ET.tostring(root, encoding="utf-8", xml_declaration=True)
                    zout.writestr(item, new_xml)
                else:
                    zout.writestr(item, zin.read(item.filename))

        buffer.seek(0)
        return load_workbook(buffer, data_only=data_only)

    except IndexError as e:
        # 스타일 인덱스 꼬여서 나는 openpyxl 버그 회피용
        # 템플릿 저장에 쓰일 일 있는 케이스(data_only=False)는 그대로 올려보내고,
        # 인풋 읽기용(data_only=True)일 때만 read_only 모드로 다시 시도
        if not data_only:
            raise
        return load_workbook(xlsx_path, data_only=data_only, read_only=True)

def _header_map(ws, header_row: int = 1):
    mapping = {}
    for cell in ws[header_row]:
        if cell.value is None:
            continue
        key = str(cell.value)
        key = key.replace("\u00A0", " ")
        key = re.sub(r"\s+", "", key)
        key = key.replace(".", "")
        mapping[key] = cell.column
    return mapping

def _write_text_cell(ws, row: int, col: int, value: Any):
    """
    값은 그대로 문자열로 넣고, 셀 타입/서식은 텍스트로 강제.
    - 3-1, 01, 010-1234 같은 것들 날짜/숫자로 안 바뀌게 막기 위함.
    """
    cell = ws.cell(row=row, column=col)
    cell.value = "" if value is None else str(value)
    cell.data_type = "s"
    cell.number_format = "@"
    return cell

def _find_last_data_row(ws, key_col: int, start_row: int) -> int:
    last = start_row - 1
    for r in range(start_row, ws.max_row + 1):
        v = ws.cell(row=r, column=key_col).value
        if v is not None and str(v).strip() != "":
            last = r
    return last

def _clear_sheet_rows(ws, start_row=2):
    if ws.max_row >= start_row:
        ws.delete_rows(start_row, ws.max_row - start_row + 1)

def _move_sheet_after(wb, sheet_name: str, after_name: str):
    if sheet_name not in wb.sheetnames or after_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]
    wb._sheets.remove(ws)
    idx = wb.sheetnames.index(after_name)
    wb._sheets.insert(idx + 1, ws)

def _delete_rows_below(ws, last_keep_row: int):
    if ws.max_row > last_keep_row:
        ws.delete_rows(last_keep_row + 1, ws.max_row - last_keep_row)

def _clear_format_workbook_from_row(wb, start_row: int = 2):
    """
    모든 시트에서:
    - start_row부터 실제 데이터가 있는 마지막 행까지 스캔
    - 그 아래 행들에 대해서만 서식(fill, border) 제거
    """
    for ws in wb.worksheets:
        last_data_row = 0
        max_row = ws.max_row
        max_col = ws.max_column or 1

        # 실제 데이터 마지막 행 찾기
        for r in range(start_row, max_row + 1):
            row_has_value = False
            for c in range(1, max_col + 1):
                v = ws.cell(row=r, column=c).value
                if v is not None and str(v).strip() != "":
                    row_has_value = True
                    break
            if row_has_value:
                last_data_row = r

        if last_data_row == 0:
            continue

        # 마지막 데이터 행 아래부터 서식 제거
        for r in range(last_data_row + 1, max_row + 1):
            for c in range(1, max_col + 1):
                cell = ws.cell(r, c)
                cell.fill = PatternFill(fill_type=None)
                cell.border = Border()

def _reset_view_to_a1(wb):
    """
    - 모든 시트: 화면은 A1, 커서는 A2
    - 모든 시트: 1행 고정(freeze_panes = A2)
    - 모든 시트: 그룹 선택(tabSelected) 해제
    - 통합문서: 첫 번째 시트만 선택 + 활성
    """
    # 1) 공통 뷰/고정 설정
    for ws in wb.worksheets:
        sv = ws.sheet_view

        # 화면/커서
        sv.topLeftCell = "A1"
        sv.activeCell = "A2"
        sv.selection = [Selection(activeCell="A2", sqref="A2")]

        # 1행 고정
        ws.freeze_panes = "A2"

        # 시트 그룹 선택 풀기
        if hasattr(sv, "tabSelected"):
            sv.tabSelected = False

    # 2) 첫 번째 시트만 선택 + 활성
    first_ws = wb.worksheets[0]
    if hasattr(first_ws.sheet_view, "tabSelected"):
        first_ws.sheet_view.tabSelected = True

    wb.active = 0

    # 3) 통합문서 뷰도 첫 시트 기준으로 통일
    if getattr(wb, "views", None):
        views = wb.views
        if views:
            views[0].activeTab = 0
            views[0].firstSheet = 0





# ========== L1: domain utils (names / headers / examples) ==========

def _normalize_name(raw: str) -> str:
    if raw is None:
        return ""
    s = str(raw).strip()
    s = re.sub(r"[^A-Za-z가-힣\s]", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    if not s:
        return ""

    has_ko = bool(HANGUL_RE.search(s))
    has_en = bool(EN_RE.search(s))

    if has_ko and not has_en:
        return s.replace(" ", "")

    if has_en and not has_ko:
        parts = [p for p in s.split(" ") if p]
        parts = [p.lower().capitalize() for p in parts]
        return "".join(parts)

    if has_ko and has_en:
        def _fix_en(m: re.Match) -> str:
            tok = m.group(0).lower()
            return tok[0].upper() + tok[1:] if tok else tok
        s2 = re.sub(r"[A-Za-z]+", _fix_en, s)
        return s2.replace(" ", "")

    return ""

def _normalize_name_key(raw: str) -> str:
    if raw is None:
        return ""
    s = str(raw).strip()
    s = re.sub(r"[^A-Za-z가-힣\s]", "", s)
    s = re.sub(r"\s+", "", s)
    return s.casefold()

def _english_casefold_key(name: str) -> str:
    if name is None:
        return ""
    return str(name).strip().casefold()

def _dedup_suffix_letters(n: int) -> str:
    if n <= 0:
        return ""
    out = ""
    while n > 0:
        n -= 1
        out = chr(ord("A") + (n % 26)) + out
        n //= 26
    return out

def _apply_suffix_for_duplicates(names: List[str]) -> List[str]:
    total = {}
    for nm in names:
        key = _english_casefold_key(nm)
        total[key] = total.get(key, 0) + 1

    seen = {}
    out = []
    for nm in names:
        key = _english_casefold_key(nm)
        if total.get(key, 0) <= 1:
            out.append(nm)
            continue
        seen[key] = seen.get(key, 0) + 1
        out.append(nm + _dedup_suffix_letters(seen[key]))
    return out

def _strip_korean_suffix_for_notice(raw_name: Any) -> str:
    """
    안내파일 동명이인 판정용:
    - '김서현A', '김서현B' 같이 '한글+대문자' 패턴이면 뒤의 대문자만 떼고 본 이름만 남김
    - 영어 이름(James, Anna 등)은 그대로 유지 (맨 끝이 대문자여도 한글이 없으면 안 깎음)
    """
    if raw_name is None:
        return ""
    s = str(raw_name).strip()
    if not s:
        return ""

    has_ko = bool(HANGUL_RE.search(s))

    # 한글이 있고, 끝이 대문자이고, 길이가 어느 정도 이상이면 suffix로 간주
    if has_ko and re.search(r"[A-Z]$", s) and len(s) >= 3:
        # 맨 끝 연속 대문자만 제거 (예: '김서현A', '김서현AB')
        s = re.sub(r"[A-Z]+$", "", s).strip()

    return s

def _notice_name_key(raw_name: Any) -> str:
    """
    안내파일 동명이인 판정용 최종 키:
    - 한글+A/B/C suffix 제거 후 normalize_name_key 적용
    """
    base = _strip_korean_suffix_for_notice(raw_name)
    return _normalize_name_key(base)

def _normalize_header_cell(val: Any) -> str:
    """
    엑셀 헤더 셀 정규화:
    - None → ""
    - 공백/줄바꿈/nbsp 제거
    - 마침표(.) 제거
    - 소문자 변환
    """
    if val is None:
        s = ""
    else:
        s = str(val)

    # nbsp → 일반 공백
    s = s.replace("\u00A0", " ")
    # 모든 공백 제거
    s = re.sub(r"\s+", "", s)
    # 마침표 제거
    s = s.replace(".", "")
    # 소문자
    s = s.lower()
    return s

def _build_header_slot_map(ws, header_row: int, slots: Dict[str, List[str]]) -> Dict[str, int]:
    """
    slots 정의(FRESHMEN_HEADER_SLOTS 등)를 기준으로
    실제 엑셀 헤더 행에서 각 slot이 어느 컬럼에 있는지 찾아서
    {slot: col_idx} 형태로 반환.
    """
    # header_map: {헤더텍스트 -> column index}
    hm = _header_map(ws, header_row)

    # 헤더 텍스트를 정규화해서 비교 (공백/마침표 제거, 소문자)
    norm_to_col: Dict[str, int] = {}
    for raw_key, col in hm.items():
        norm_key = _normalize_header_cell(raw_key)
        if norm_key:
            norm_to_col[norm_key] = col

    result: Dict[str, int] = {}

    for slot, patterns in slots.items():
        for pat in patterns:
            pat_norm = _normalize_header_cell(pat)
            if not pat_norm:
                continue
            # 헤더 정규화 문자열 안에 패턴이 포함되면 매칭
            for header_norm, col in norm_to_col.items():
                if pat_norm in header_norm:
                    result[slot] = col
                    break
            if slot in result:
                break

    return result

def _detect_header_row_generic(ws, slots: Dict[str, List[str]],
                               max_search_row: int = 15,
                               max_col: int = 10,
                               min_match_slots: int = 3) -> int:
    """
    slots: {slot_name: [pattern1, pattern2, ...]}
    한 행에서 slot이 몇 개 매칭되는지 보고, min_match_slots 이상이면 헤더 후보로 본다.
    """
    best_row: Optional[int] = None
    best_score: int = 0

    for row in ws.iter_rows(min_row=1, max_row=max_search_row):
        row_idx = row[0].row
        vals = [_normalize_header_cell(c.value) for c in row[:max_col]]

        matched_slots = set()
        for slot, patterns in slots.items():
            for pat in patterns:
                pat_norm = _normalize_header_cell(pat)
                if not pat_norm:
                    continue
                if any(pat_norm in v for v in vals if v):
                    matched_slots.add(slot)
                    break  # 슬롯 당 1회만 카운트

        score = len(matched_slots)
        if score > best_score:
            best_score = score
            best_row = row_idx

    if best_row is None or best_score < min_match_slots:
        raise ValueError(
            "[오류] 인풋 파일에서 헤더를 자동으로 찾지 못했습니다. "
            "헤더 행에 학년/반/이름/NO 등의 키워드가 동시에 3개 이상 있어야 합니다."
        )

    return best_row

def _detect_header_row_freshmen(ws) -> int:
    return _detect_header_row_generic(ws, FRESHMEN_HEADER_SLOTS,
                                      max_search_row=15, max_col=10, min_match_slots=3)

def _detect_header_row_transfer(ws) -> int:
    return _detect_header_row_generic(ws, TRANSFER_HEADER_SLOTS,
                                      max_search_row=15, max_col=10, min_match_slots=3)

def _detect_header_row_withdraw(ws) -> int:
    return _detect_header_row_generic(ws, WITHDRAW_HEADER_SLOTS,
                                      max_search_row=15, max_col=10, min_match_slots=3)

def _detect_header_row_teacher(ws) -> int:
    # 교사는 NO / 이름 / 학년/반 / 신청 컬럼 중 최소 3슬롯 이상
    return _detect_header_row_generic(ws, TEACHER_HEADER_SLOTS,
                                      max_search_row=15, max_col=10, min_match_slots=3)

def _row_is_empty(ws, row: int, max_col: Optional[int] = None) -> bool:
    if max_col is None:
        max_col = ws.max_column or 1
    for c in range(1, max_col + 1):
        v = ws.cell(row=row, column=c).value
        if v is not None and str(v).strip() != "":
            return False
    return True

def _row_has_example_keyword(ws, row: int, max_col: Optional[int] = None) -> bool:
    if max_col is None:
        max_col = ws.max_column or 1
    for c in range(1, max_col + 1):
        v = ws.cell(row=row, column=c).value
        if v is None:
            continue
        s = normalize_text(str(v))
        if not s:
            continue
        for kw in EXAMPLE_KEYWORDS:
            if kw in s:
                return True
    return False

def _cell_is_example_name(value: Any) -> bool:
    if value is None:
        return False
    s = normalize_text(str(value))
    return bool(s) and s in EXAMPLE_NAMES_NORM

def _detect_example_and_data_start(
    ws,
    header_row: int,
    name_col: int,
    max_search_row: Optional[int] = None,
    max_col: Optional[int] = None,
) -> Tuple[List[int], int]:
    """
    헤더 아래에서 예시 행(0개 이상)과 실제 데이터 시작 행을 자동 감지한다.

    - header_row 바로 아래 행부터 스캔
    - 완전 빈 행은 건너뜀
    - '예시' 키워드가 있거나 이름 칸이 예시 이름이면 → 예시 행
    - 그 외 첫 번째 비-예시 행 → 실제 데이터 시작 행
    """
    if max_search_row is None:
        max_search_row = ws.max_row

    example_rows: List[int] = []
    r = header_row + 1

    while r <= max_search_row:
        # 1) 완전 빈 행은 스킵
        if _row_is_empty(ws, r, max_col=max_col):
            r += 1
            continue

        # 2) 행 안에 '예시' 키워드 있으면 예시
        if _row_has_example_keyword(ws, r, max_col=max_col):
            example_rows.append(r)
            r += 1
            continue

        # 3) 이름 칸이 예시 이름이면 예시
        v_name = ws.cell(row=r, column=name_col).value
        if _cell_is_example_name(v_name):
            example_rows.append(r)
            r += 1
            continue

        # 4) 여기까지 안 걸리면 → 실제 데이터 시작
        return example_rows, r

    raise ValueError(
        f"[오류] 데이터 시작 행을 찾지 못했습니다. 헤더({header_row}행) 아래에 예시나 실제 데이터로 보이는 행이 없습니다."
    )

def detect_input_layout(xlsx_path: Path, kind: str) -> Dict[str, Any]:
    """
    UI에서 인풋 파일 구조를 미리 보여줄 때 사용.
    kind: 'freshmen' | 'transfer' | 'withdraw' | 'teacher'
    반환:
      {
        "header_row": int,
        "example_rows": [int, ...],
        "data_start_row": int,
      }
    """
    _ensure_xlsx_only(xlsx_path)
    wb = _safe_load_workbook(xlsx_path, data_only=True)
    ws = wb.worksheets[0]

    kind_norm = (kind or "").strip().lower()

    # 1) 헤더 자동 감지
    if kind_norm == "freshmen":
        header_row = _detect_header_row_freshmen(ws)
        # 🔹 헤더 이름 기준으로 실제 이름 컬럼 찾기
        slot_cols = _build_header_slot_map(ws, header_row, FRESHMEN_HEADER_SLOTS)
        name_col = slot_cols.get("name", 5)  # 못 찾으면 기존 E열 fallback
   
    elif kind_norm == "transfer":
        header_row = _detect_header_row_transfer(ws)
        name_col = 5  # E열: 성명
    elif kind_norm == "withdraw":
        header_row = _detect_header_row_withdraw(ws)
        name_col = 4  # D열: 성명
    elif kind_norm == "teacher":
        header_row = _detect_header_row_teacher(ws)
        name_col = 3  # C열: 선생님 이름
    else:
        raise ValueError(f"[오류] 지원하지 않는 kind 값입니다: {kind}")

    # 2) 예시 행 + 실제 데이터 시작 행 자동 감지
    example_rows, data_start_row = _detect_example_and_data_start(
        ws,
        header_row=header_row,
        name_col=name_col,
    )

    return {
        "header_row": header_row,
        "example_rows": example_rows,
        "data_start_row": data_start_row,
    }

def _normalize_withdraw_class(raw, grade: int) -> str:
    """
    전출 명단 C열(반) 문자열을 통일된 형식으로 정규화:
    - 학년 정보는 무시하고, 문자열에서 '마지막 숫자 묶음'을 반 번호로 사용한다.
      예)
        '1-10'          -> grade-10반
        '1학년10반'     -> grade-10반
        '1 학년 10 반'  -> grade-10반
        '10반' / '10'   -> grade-10반
        '3-5반'         -> grade-5반
    """
    if raw is None:
        return ""

    s = str(raw).strip()
    if not s:
        return ""

    # 공백/전각 공백 정리
    s = s.replace("\u3000", " ").replace("\u00A0", " ")
    s = re.sub(r"\s+", "", s)

    # 문자열 안의 숫자 덩어리들 전부 추출
    nums = re.findall(r"\d+", s)
    if not nums:
        # 숫자가 전혀 없으면 그냥 원본 반환 (최소한 이상한 값이라는 걸 눈으로 보게)
        return s

    # "마지막 숫자 묶음"을 반 번호로 사용
    class_no = int(nums[-1])

    return f"{grade}-{class_no}반"





# ========== L2: input readers (신입/전입/전출/교사) ==========

def _read_freshmen_rows(
    xlsx_path: Path,
    header_row: Optional[int] = None,
    data_start_row: Optional[int] = None,
) -> List[Dict]:
    """
    신입생 파일을 읽어서
    [{"grade": 학년(int), "class": 반(str), "number": 번호(str), "name": 이름(str)}, ...]
    형태로 반환.

    - 학년 / 반 / 이름은 필수
    - 번호는 없어도 됨(빈 문자열로 처리)
    - 컬럼 위치는 고정(B,C,D,E)이 아니라 헤더 이름 기준으로 탐색
    """
    _ensure_xlsx_only(xlsx_path)
    wb = _safe_load_workbook(xlsx_path, data_only=True)
    ws = wb.worksheets[0]

    # 1) 헤더 자동 감지 (필요 시)
    if header_row is None:
        header_row = _detect_header_row_freshmen(ws)

    # 1-1) 헤더에서 실제 컬럼 위치 찾기
    slot_cols = _build_header_slot_map(ws, header_row, FRESHMEN_HEADER_SLOTS)
    col_grade = slot_cols.get("grade")
    col_class = slot_cols.get("class")
    # 번호는 num(번호/번) 우선, 없으면 no(번호)라도 사용
    col_num = slot_cols.get("num") or slot_cols.get("no")
    col_name = slot_cols.get("name")

    missing = []
    if col_grade is None:
        missing.append("학년")
    if col_class is None:
        missing.append("반")
    if col_name is None:
        missing.append("성명/이름")

    # 학년/반/이름은 필수, 번호는 없어도 됨
    if missing:
        raise ValueError(
            "[오류] 신입생 파일 헤더에서 "
            + ", ".join(missing)
            + " 열을 찾지 못했습니다. '학년', '반', '이름' 헤더를 추가하거나 수정해 주세요."
        )

    # 2) 예시/데이터 시작 행 자동 감지 (사용자가 data_start_row 직접 준 경우 우선)
    if data_start_row is None:
        _, data_start_row = _detect_example_and_data_start(
            ws,
            header_row=header_row,
            name_col=col_name,
        )

    out: List[Dict[str, Any]] = []
    row = data_start_row
    while True:
        grade = ws.cell(row=row, column=col_grade).value
        cls   = ws.cell(row=row, column=col_class).value
        num   = ws.cell(row=row, column=col_num).value if col_num is not None else None
        name  = ws.cell(row=row, column=col_name).value

        # 1) 행 전체가 비어 있으면 종료
        if all(v is None or str(v).strip() == "" for v in [grade, cls, num, name]):
            break

        # 2) 필수값(학년, 반, 이름) 체크
        if any(v is None or str(v).strip() == "" for v in [grade, cls, name]):
            raise ValueError(
                f"[오류] 신입생 파일 {row}행에서 학년/반/이름 중 비어 있는 값이 있습니다."
            )

        # 3) 학년에서 숫자만 추출
        grade_s = str(grade).strip()
        m = re.search(r"\d+", grade_s)
        if not m:
            raise ValueError(
                f"[오류] 신입생 파일 {row}행 학년에서 숫자를 찾지 못했습니다: {grade_s!r}"
            )
        grade_i = int(m.group(0))

        cls_s = str(cls).strip()
        num_s = "" if (num is None or str(num).strip() == "") else str(num).strip()
        name_n = _normalize_name(name)
        if not name_n:
            raise ValueError(
                f"[오류] 신입생 파일 {row}행 이름 정규화 결과가 비어 있습니다."
            )

        out.append(
            {
                "grade": grade_i,
                "class": cls_s,
                "number": num_s,
                "name": name_n,
            }
        )
        row += 1

    def _safe_int(x: str):
        try:
            return (0, int(x))
        except Exception:
            return (1, x)

    out.sort(
        key=lambda r: (
            r["grade"],
            _safe_int(r["class"]),
            _safe_int(r["number"]),
        )
    )
    return out

def _read_transfer_rows(
    xlsx_path: Path,
    header_row: Optional[int] = None,
    data_start_row: Optional[int] = None,
) -> List[Dict]:
    """
    전입생 엑셀에서 학년/반/번호/이름만 읽어온다.
    - ID는 전혀 사용하지 않는다 (슬롯도 두지 않음).
    - 헤더 매핑은 TRANSFER_HEADER_SLOTS만 사용.
    """
    _ensure_xlsx_only(xlsx_path)
    wb = _safe_load_workbook(xlsx_path, data_only=True)
    ws = wb.worksheets[0]

    # 1) 헤더 행 탐지
    if header_row is None:
        header_row = _detect_header_row_transfer(ws)

    # 2) 헤더 → 컬럼 매핑 (슬롯은 TRANSFER_HEADER_SLOTS에 정의된 것만 사용)
    slot_cols = _build_header_slot_map(ws, header_row, TRANSFER_HEADER_SLOTS)

    col_grade = slot_cols.get("grade")
    col_class = slot_cols.get("class")
    col_num   = slot_cols.get("num")   # 번호는 있으면 사용, 없어도 됨
    col_name  = slot_cols.get("name")

    missing = []
    if col_grade is None:
        missing.append("학년")
    if col_class is None:
        missing.append("반")
    if col_name is None:
        missing.append("이름")

    # 번호는 필수 아님
    if missing:
        raise ValueError(
            "[오류] 전입생 파일 헤더에서 "
            + ", ".join(missing)
            + " 열을 찾지 못했습니다. 헤더명을 확인해 주세요."
        )

    # 3) 데이터 시작 행
    if data_start_row is None:
        _, data_start_row = _detect_example_and_data_start(
            ws,
            header_row=header_row,
            name_col=col_name,
        )

    out: List[Dict[str, Any]] = []
    row = data_start_row

    while True:
        grade = ws.cell(row=row, column=col_grade).value
        cls   = ws.cell(row=row, column=col_class).value
        num   = ws.cell(row=row, column=col_num).value if col_num is not None else None
        name  = ws.cell(row=row, column=col_name).value

        # 완전 빈 줄이면 종료 (학년/반/번호/이름 전부 비어 있으면)
        if all(
            v is None or str(v).strip() == ""
            for v in [grade, cls, num, name]
        ):
            break

        # 필수값 체크: 학년/반/이름만 필수
        if any(
            v is None or str(v).strip() == ""
            for v in [grade, cls, name]
        ):
            raise ValueError(
                f"[오류] 전입생 파일 {row}행에서 학년/반/이름 중 비어 있는 값이 있습니다."
            )

        # 학년 숫자 추출
        grade_s = str(grade).strip()
        m = re.search(r"\d+", grade_s)
        if not m:
            raise ValueError(
                f"[오류] 전입생 파일 {row}행 학년에서 숫자를 찾지 못했습니다: {grade_s!r}"
            )
        grade_i = int(m.group(0))

        cls_s = str(cls).strip()
        num_s = "" if (num is None or str(num).strip() == "") else str(num).strip()
        name_n = _normalize_name(name)

        out.append(
            {
                "grade": grade_i,
                "class": cls_s,
                "number": num_s,
                "name": name_n,
                # ⚠️ ID는 전입에서 절대 쓰지 않는다 -> 키 자체를 만들지 않음
            }
        )
        row += 1

    # 필요하면 여기서 grade / class / number 기준 정렬 가능
    return out

def _read_teacher_rows(
    xlsx_path: Path,
    header_row: Optional[int] = None,
    data_start_row: Optional[int] = None,
) -> List[Dict]:
    """
    교사 파일 읽기.
    헤더 이름을 기준으로 직위/이름/학습용ID신청/관리용ID신청 컬럼을 찾는다.
    """
    _ensure_xlsx_only(xlsx_path)
    wb = _safe_load_workbook(xlsx_path, data_only=True)
    ws = wb.worksheets[0]

    # 1) 헤더 자동 감지
    if header_row is None:
        header_row = _detect_header_row_teacher(ws)

    # 2) 헤더 → 컬럼 매핑
    slot_cols = _build_header_slot_map(ws, header_row, TEACHER_HEADER_SLOTS)
    col_pos   = slot_cols.get("position")  # 직위/담당 (없어도 됨)
    col_name  = slot_cols.get("name")      # 이름 (필수)
    col_learn = slot_cols.get("learn")     # 학습용 ID 신청 (없으면 False)
    col_admin = slot_cols.get("admin")     # 관리용 ID 신청 (없으면 False)

    if col_name is None:
        raise ValueError(
            "[오류] 교사 파일 헤더에서 이름 열을 찾지 못했습니다. '성명' 또는 '이름' 헤더를 확인해 주세요."
        )

    # 3) 예시/데이터 시작 행 자동 감지
    if data_start_row is None:
        _, data_start_row = _detect_example_and_data_start(
            ws,
            header_row=header_row,
            name_col=col_name,
        )

    out: List[Dict[str, Any]] = []
    row = data_start_row
    while True:
        # 현재 행 값들 읽기
        def _get(col_idx: Optional[int]):
            if col_idx is None:
                return None
            return ws.cell(row=row, column=col_idx).value

        pos    = _get(col_pos)
        name   = _get(col_name)
        v_learn = _get(col_learn)
        v_admin = _get(col_admin)

        # 완전 빈 줄이면 종료
        if all(
            v is None or str(v).strip() == ""
            for v in [pos, name, v_learn, v_admin]
        ):
            break

        # 이름 없으면 그 행은 건너뜀
        if name is None or str(name).strip() == "":
            row += 1
            continue

        name_n = _normalize_name(name)
        if not name_n:
            row += 1
            continue

        learn_apply = False
        admin_apply = False
        if col_learn is not None:
            learn_apply = not (v_learn is None or str(v_learn).strip() == "")
        if col_admin is not None:
            admin_apply = not (v_admin is None or str(v_admin).strip() == "")

        out.append(
            {
                "position": "" if pos is None else str(pos).strip(),
                "name": name_n,
                "learn_apply": learn_apply,
                "admin_apply": admin_apply,
            }
        )
        row += 1

    return out

def _read_withdraw_rows(
    xlsx_path: Path,
    header_row: Optional[int] = None,
    data_start_row: Optional[int] = None,
) -> List[Dict]:
    """
    전출생 파일 읽기.
    헤더 이름을 기준으로 학년/반/이름 컬럼을 찾는다.
    """
    _ensure_xlsx_only(xlsx_path)
    wb = _safe_load_workbook(xlsx_path, data_only=True)
    ws = wb.worksheets[0]

    # 1) 헤더 자동 감지
    if header_row is None:
        header_row = _detect_header_row_withdraw(ws)

    # 2) 헤더 → 컬럼 매핑
    slot_cols = _build_header_slot_map(ws, header_row, WITHDRAW_HEADER_SLOTS)
    col_grade = slot_cols.get("grade")
    col_class = slot_cols.get("class")
    col_name  = slot_cols.get("name")

    missing = []
    if col_grade is None:
        missing.append("학년")
    if col_class is None:
        missing.append("반")
    if col_name is None:
        missing.append("성명/이름")

    if missing:
        raise ValueError(
            "[오류] 전출생 파일 헤더에서 "
            + ", ".join(missing)
            + " 열을 찾지 못했습니다. 헤더명을 확인해 주세요."
        )

    # 3) 예시/데이터 시작 행 자동 감지
    if data_start_row is None:
        _, data_start_row = _detect_example_and_data_start(
            ws,
            header_row=header_row,
            name_col=col_name,
        )

    out: List[Dict[str, Any]] = []
    row = data_start_row
    while True:
        grade = ws.cell(row=row, column=col_grade).value
        cls   = ws.cell(row=row, column=col_class).value
        name  = ws.cell(row=row, column=col_name).value

        vals = [grade, cls, name]
        # 완전 빈 줄이면 종료
        if all(v is None or str(v).strip() == "" for v in vals):
            break
        # 일부만 비어 있으면 오류
        if any(v is None or str(v).strip() == "" for v in vals):
            raise ValueError(
                f"[오류] 전출생 파일 {row}행에 학년/반/이름 중 비어 있는 값이 있습니다."
            )

        # 학년에서 숫자만 추출 (1, 2학년, "3" 다 커버)
        grade_s = str(grade).strip()
        m = re.search(r"\d+", grade_s)
        if not m:
            raise ValueError(
                f"[오류] 전출생 파일 {row}행 학년에서 숫자를 찾지 못했습니다: {grade_s!r}"
            )
        grade_i = int(m.group(0))

        cls_s = _normalize_withdraw_class(cls, grade_i)
        if not cls_s:
            raise ValueError(
                f"[오류] 전출생 파일 {row}행 반 정규화 결과가 비어 있습니다."
            )

        name_n = _normalize_name(name)
        if not name_n:
            raise ValueError(
                f"[오류] 전출생 파일 {row}행 이름 정규화 결과가 비어 있습니다."
            )

        out.append({"grade": grade_i, "class": cls_s, "name": name_n})
        row += 1

    return out





# ========== L3: roster / transfer / withdraw core logic ==========

def _parse_roster_year_from_filename(roster_path: Path) -> Optional[int]:
    stem = roster_path.stem
    s = stem.replace("\u00A0", " ")
    s = re.sub(r"\s+", " ", s).strip()

    m = re.search(r"(\d{4})\s*학\s*년도", s)
    if m:
        return int(m.group(1))

    m2 = re.search(r"(19\d{2}|20\d{2})", s)
    if m2:
        return int(m2.group(1))

    return None

def _load_roster_sheet(dirs: Dict[str, Path], school_name: str):
    """
    학생명부(.xlsx, 파일명에 '학생명부' 포함)를 학교 폴더에서 찾아서
    - 첫 번째 시트를 openpyxl 워크시트로 반환
    - 파일 경로
    - 파일명 기준 추정 학년도 (없으면 None)
    를 돌려준다.
    """
    root_dir = dirs["SCHOOL_ROOT"]

    kw = (school_name or "").strip()
    if not kw:
        raise ValueError("[오류] 학교명이 비어 있어 학생명부 폴더를 찾을 수 없습니다.")

    # 🔹 학교 폴더를 포함 매칭으로 찾기
    matches = [
        p
        for p in root_dir.iterdir()
        if p.is_dir() and text_contains(p.name, kw)
    ]

    if not matches:
        raise ValueError(
            f"[오류] 학생명부를 찾을 학교 폴더를 찾지 못했습니다. "
            f"(작업 폴더 내 '{school_name}' 포함 폴더 없음)"
        )

    if len(matches) > 1:
        raise ValueError(
            f"[오류] 학생명부를 찾을 학교 폴더 후보가 여러 개입니다: "
            + ", ".join(p.name for p in matches)
        )

    school_root = matches[0]

    candidates: List[Path] = [
        p
        for p in school_root.iterdir()
        if p.is_file()
        and p.suffix.lower() == ".xlsx"
        and "학생명부" in p.stem
        and not p.name.startswith("~$")
    ]
    if not candidates:
        raise ValueError("[오류] 학생명부(.xlsx, 파일명에 '학생명부') 파일을 찾지 못했습니다.")

    # 가장 최근 수정 파일 사용
    candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    roster_path = candidates[0]

    wb = _safe_load_workbook(roster_path, data_only=True)
    ws = wb.worksheets[0]
    roster_year = _parse_roster_year_from_filename(roster_path)

    return ws, roster_path, roster_year

def _parse_class_str(s: str) -> Optional[Tuple[int, str]]:
    if s is None:
        return None
    m = re.match(r"^\s*(\d+)\s*-\s*(.+?)\s*$", str(s))
    if not m:
        return None
    return int(m.group(1)), m.group(2).strip()

def _extract_id_prefix4(uid: str) -> Optional[int]:
    if uid is None:
        return None
    s = str(uid).strip()
    if len(s) >= 4 and s[:4].isdigit():
        return int(s[:4])
    return None

def _analyze_roster_once(roster_ws, input_year: int) -> Dict:
    hm = _header_map(roster_ws, 1)
    need = ["현재반", "이전반", "학생이름", "아이디"]
    for k in need:
        if k not in hm:
            raise ValueError(f"[오류] 학생명부에 '{k}' 헤더가 없습니다.")

    c_class = hm["현재반"]
    c_name  = hm["학생이름"]
    c_id    = hm["아이디"]

    prefixes_by_grade = defaultdict(list)
    name_counter_by_grade = defaultdict(Counter)
    prefixes_grade1 = []

    for r in range(2, roster_ws.max_row + 1):
        clv = roster_ws.cell(r, c_class).value
        nmv = roster_ws.cell(r, c_name).value
        idv = roster_ws.cell(r, c_id).value
        if clv is None or nmv is None:
            continue

        parsed = _parse_class_str(clv)
        if parsed is None:
            continue
        g, _cls = parsed

        nm = _normalize_name(nmv)
        if not nm:
            continue
        name_counter_by_grade[g][nm] += 1

        p4 = _extract_id_prefix4(idv)
        if p4 is not None:
            prefixes_by_grade[g].append(p4)
            if g == 1:
                prefixes_grade1.append(p4)

    prefix_mode_by_grade = {}
    for g, arr in prefixes_by_grade.items():
        if arr:
            prefix_mode_by_grade[g] = Counter(arr).most_common(1)[0][0]

    roster_time = "unknown"
    ref_shift = 0
    if prefixes_grade1:
        mode1 = Counter(prefixes_grade1).most_common(1)[0][0]
        if mode1 == input_year:
            roster_time = "this_year"
            ref_shift = 0
        elif mode1 == input_year - 1:
            roster_time = "last_year"
            ref_shift = -1
        else:
            roster_time = "unknown"
            ref_shift = 0

    return {
        "roster_time": roster_time,
        "ref_grade_shift": ref_shift,
        "prefix_mode_by_roster_grade": prefix_mode_by_grade,
        "name_count_by_roster_grade": name_counter_by_grade,
    }

def _build_transfer_ids(
    transfer_rows: List[Dict],
    roster_info: Dict,
    input_year: int,
) -> Tuple[List[Dict], List[Dict], Dict[int, int]]:
    shift = roster_info["ref_grade_shift"]
    prefix_mode = roster_info["prefix_mode_by_roster_grade"]
    name_counts = roster_info["name_count_by_roster_grade"]

    done: List[Dict] = []
    hold: List[Dict] = []
    final_prefix_by_current_grade: Dict[int, int] = {}
    seen_in_transfer_by_grade = defaultdict(Counter)

    grade1_rows = [tr for tr in transfer_rows if tr["grade"] == 1]
    if grade1_rows:
        g1_names = [tr["name"] for tr in grade1_rows]
        g1_names_sfx = _apply_suffix_for_duplicates(g1_names)
        for tr, nm_sfx in zip(grade1_rows, g1_names_sfx):
            uid = f"{input_year}{nm_sfx}"
            done.append({**tr, "id": uid})

    other_rows = [tr for tr in transfer_rows if tr["grade"] != 1]

    for tr in other_rows:
        g_cur = tr["grade"]
        g_roster = g_cur + shift

        pref = prefix_mode.get(g_roster)
        if pref is None:
            hold.append({**tr, "보류사유": f"명부 학년({g_roster})에서 ID prefix 최빈값 산출 불가"})
            continue

        final_prefix_by_current_grade[g_cur] = pref

        nm = tr["name"]
        base_cnt = name_counts.get(g_roster, Counter()).get(nm, 0)

        seen_in_transfer_by_grade[g_cur][nm] += 1
        add_seq = seen_in_transfer_by_grade[g_cur][nm]

        need_suffix = (base_cnt > 0)
        suffix = _dedup_suffix_letters(add_seq) if need_suffix else ""

        uid = f"{pref}{nm}{suffix}"

        is_dup_with_roster = base_cnt > 0  # 🔸 명부 기준 동명이인 여부

        done.append({**tr, "id": uid})

    def _safe_int(x: str):
        try:
            return (0, int(x))
        except Exception:
            return (1, str(x))

    done.sort(key=lambda r: (r["grade"], _safe_int(r["class"]), _safe_int(r["number"]), r["name"]))
    hold.sort(key=lambda r: (r["grade"], _safe_int(r["class"]), _safe_int(r["number"]), r["name"]))

    return done, hold, final_prefix_by_current_grade

def _build_withdraw_outputs(
    roster_ws,
    withdraw_rows: List[Dict],
    school_start_date: date,
    work_date: date,
    roster_info: Optional[Dict] = None,
) -> Tuple[List[Dict], List[Dict]]:
    """
    학생명부 + 전출 명단 기반 퇴원/보류 리스트 생성.
    - 퇴원일자: 작업일 < 개학일 → 개학일, 그 외에는 작업일 기준
    """
    done: List[Dict] = []
    hold: List[Dict] = []

    # 퇴원일자(파일 전체 공통)
    eff = school_start_date if work_date < school_start_date else work_date

    hm = _header_map(roster_ws, 1)
    need = ["현재반", "이전반", "학생이름", "아이디"]
    for k in need:
        if k not in hm:
            raise ValueError(f"[오류] 학생명부에 '{k}' 헤더가 없습니다.")

    col_now   = hm["현재반"]
    col_prev  = hm["이전반"]
    col_name  = hm["학생이름"]
    col_id    = hm["아이디"]

    # scan에서 넘겨준 학년도 정보는 로그/참고용으로만 사용
    roster_time = (roster_info or {}).get("roster_time", "this_year")

    # 인덱스들
    roster_map: Dict[str, List[Dict]] = {}
    roster_by_grade_name: Dict[str, List[Dict]] = {}
    seen_grade_name_ids = set()  # (grade, name_key, id_str)

    def _strip_name_suffix(name_key: str) -> str:
        """
        이름 키에서 뒤쪽 A/B/C 같은 알파벳 suffix 제거.
        예: '홍길동A' -> '홍길동', '홍길동' -> '홍길동'
        """
        if not name_key:
            return ""
        return re.sub(r"[A-Za-z]+$", "", name_key)

    def _find_suffix_candidates_for_grade(grade: int, base_name_key: str) -> List[Dict]:
        if grade is None or grade <= 0 or not base_name_key:
            return []

        prefix = f"{grade}|"
        out: List[Dict] = []

        for k, rows in roster_by_grade_name.items():
            if not k.startswith(prefix):
                continue
            _, nk = k.split("|", 1)
            if _strip_name_suffix(nk) == base_name_key:
                out.extend(rows)

        return out


    def _index_class_map(class_val, now_class_val, name_key: str, idv, name_disp: str):
        """반+이름 완전 매칭용 인덱스 (현재반/이전반 둘 다 매칭, 현재반은 따로 보관)"""
        if class_val is None:
            return
        c = str(class_val).strip()
        if not c:
            return

        c_now = "" if now_class_val is None else str(now_class_val).strip()

        key1 = f"{c}|{name_key}"
        roster_map.setdefault(key1, []).append(
            {
                "class": c,         # 매칭에 사용된 반(현재/이전 둘 중 하나)
                "now_class": c_now, # 학생명부 A열 현재반 (퇴원반명으로 쓸 값)
                "name_key": name_key,
                "name_disp": name_disp,
                "id": "" if idv is None else str(idv).strip(),
            }
        )

    def _index_grade_map(class_val, now_class_val, name_key: str, idv, name_disp: str):
        """학년+이름 fallback용 인덱스 (전출용 학년은 현재반 기준으로 잡는다)"""
        if class_val is None:
            return
        c = str(class_val).strip()
        if not c:
            return

        parsed = _parse_class_str(c)
        if parsed is None:
            return
        g = parsed[0]

        id_str = "" if idv is None else str(idv).strip()
        dedup_key = (g, name_key, id_str)
        if dedup_key in seen_grade_name_ids:
            return
        seen_grade_name_ids.add(dedup_key)

        c_now = "" if now_class_val is None else str(now_class_val).strip()

        key2 = f"{g}|{name_key}"
        roster_by_grade_name.setdefault(key2, []).append(
            {
                "class": c,
                "now_class": c_now,
                "name_key": name_key,
                "name_disp": name_disp,
                "id": id_str,
                "grade": g,
            }
        )

    # 학생명부 인덱스 생성
    for r in range(2, roster_ws.max_row + 1):
        nmv = roster_ws.cell(r, col_name).value
        if nmv is None:
            continue
        name_disp = _normalize_name(nmv)
        name_key  = _normalize_name_key(nmv)
        if not name_key:
            continue

        idv  = roster_ws.cell(r, col_id).value
        nowv = roster_ws.cell(r, col_now).value   # A열: 현재반
        prevv = roster_ws.cell(r, col_prev).value # 이전반

        now_class_val = nowv
        _index_class_map(nowv,  now_class_val, name_key, idv, name_disp)
        _index_class_map(prevv, now_class_val, name_key, idv, name_disp)

        base_class_val = nowv or prevv
        _index_grade_map(base_class_val, now_class_val, name_key, idv, name_disp)

    # 전출 행 처리
    for w in withdraw_rows:
        g_cur = w["grade"]
        w_name_disp = w["name"]
        w_name_key  = _normalize_name_key(w_name_disp)
        if not w_name_key:
            hold.append(
                {
                    "학년": g_cur,
                    "반": w["class"],
                    "성명": w_name_disp,
                    "보류사유": "성명 정규화(키) 결과가 비어 있음",
                }
            )
            continue

        # 전출 명단의 반(C열)은 이미 normalize_withdraw_class로 통일된 상태라고 가정
        w_class_full = w["class"]
        key = f"{w_class_full}|{w_name_key}"
        matches = roster_map.get(key, [])

        if len(matches) == 0:
            # 같은 학년/다음 학년에서 이름만 일치하는 후보 찾아보기 (exact key 기준)
            cand0 = roster_by_grade_name.get(f"{g_cur}|{w_name_key}", [])
            cand1 = roster_by_grade_name.get(f"{g_cur+1}|{w_name_key}", [])
            cand = cand0 + cand1
            if len(cand) == 1:
                # 이름 키(w_name_key) 기준으로 유일하게 찾은 경우
                matches = cand
            else:
                # 🔹 동명이인(A/B/C suffix) 처리: 이름에서 suffix 제거 후 다시 탐색
                base_name_key = _strip_name_suffix(w_name_key)

                # 1) 현재 학년 기준
                suffix_cand0 = _find_suffix_candidates_for_grade(g_cur, base_name_key)
                if len(suffix_cand0) == 1:
                    # 학년+이름 기준으로 유일 → 자동 매칭
                    matches = suffix_cand0
                elif len(suffix_cand0) >= 2:
                    # 같은 학년 안에서 동명이인 여러 명 → 보류
                    hold.append(
                        {
                            "학년": g_cur,
                            "반": w["class"],
                            "성명": w_name_disp,
                            "보류사유": (
                                "보류: 학생명부에서 동명이인(A,B,C 등)으로 구분된 이름 – "
                                "자동 매칭하지 않고 수동 확인이 필요합니다."
                            ),
                        }
                    )
                    continue
                else:
                    # 2) g+1 학년 기준 (명부 시점과 전출 명단 학년 차이 보정용)
                    suffix_cand1 = _find_suffix_candidates_for_grade(g_cur + 1, base_name_key)
                    if len(suffix_cand1) == 1:
                        matches = suffix_cand1
                    elif len(suffix_cand1) >= 2:
                        hold.append(
                            {
                                "학년": g_cur,
                                "반": w["class"],
                                "성명": w_name_disp,
                                "보류사유": (
                                    "보류: 학생명부에서 동명이인(A,B,C 등)으로 구분된 이름 – "
                                    "자동 매칭하지 않고 수동 확인이 필요합니다."
                                ),
                            }
                        )
                        continue
                    else:
                        # suffix 기준으로도 후보 없음 → 기존 cand 기반 사유 사용
                        if len(cand) == 0:
                            reason = (
                                "자동 제외: 학생명부에 존재하지 않는 학생 – "
                                "서버 미등록/학년 불일치 등으로 추정되며 퇴원 처리 대상에서 제외했습니다. "
                                "(반 매칭 실패, g 또는 g+1 탐색)"
                            )
                        else:
                            reason = (
                                f"보류: 학년+이름 후보가 2건 이상({len(cand)}건) – 수동 확인 필요. "
                                "(반 매칭 실패, g 또는 g+1 탐색)"
                            )
                        hold.append(
                            {
                                "학년": g_cur,
                                "반": w["class"],
                                "성명": w_name_disp,
                                "보류사유": reason,
                            }
                        )
                        continue

        if len(matches) > 1:
            hold.append(
                {
                    "학년": g_cur,
                    "반": w["class"],
                    "성명": w_name_disp,
                    "보류사유": f"중복 매칭({len(matches)}건)",
                }
            )
            continue

        m = matches[0]

        # 학생명부 A열(현재반) 기준으로 퇴원반명 결정
        now_class = m.get("now_class") or m.get("class") or w.get("class")
        withdraw_class = now_class

        done.append(
            {
                "퇴원반명": withdraw_class,
                "학생이름": w_name_disp,
                "아이디": m["id"],
                "퇴원일자": eff,
            }
        )

    return done, hold





# ========== L4: output writers (등록/안내/퇴원) ==========

def _write_withdraw_to_register(wb, done_rows: List[Dict], hold_rows: List[Dict]):
    # 🔹 퇴원 완료 시트: 항상 사용 (없으면 생성)
    ws_done = wb["퇴원"] if "퇴원" in wb.sheetnames else wb.create_sheet("퇴원")

    # 퇴원 완료 정렬 (퇴원반명 → 학생이름 오름차순)
    done_rows = sorted(
        done_rows,
        key=lambda r: (
            str(r.get("퇴원반명", "")).strip(),
            str(r.get("학생이름", "")).strip(),
        ),
    )

    _clear_sheet_rows(ws_done, 2)

    r = 2
    for row in done_rows:
        _write_text_cell(ws_done, r, 1, row["퇴원반명"])
        _write_text_cell(ws_done, r, 2, row["학생이름"])
        _write_text_cell(ws_done, r, 3, row["아이디"])
        ws_done.cell(r, 4).value = row["퇴원일자"]       # 날짜는 date 객체 그대로
        ws_done.cell(r, 4).number_format = "yyyy-mm-dd"
        r += 1

    # 🔹 보류 0명 처리: 시트 아예 만들지 않음 (있으면 삭제)
    ws_hold = None
    if hold_rows:
        # 보류 정렬 (학년 → 반 → 성명)
        hold_rows = sorted(
            hold_rows,
            key=lambda r: (
                str(r.get("학년", "")).strip(),
                str(r.get("반", "")).strip(),
                str(r.get("성명", "")).strip(),
            ),
        )

        ws_hold = wb["퇴원_보류"] if "퇴원_보류" in wb.sheetnames else wb.create_sheet("퇴원_보류")
        # 헤더는 템플릿에 있다고 가정하고, 2행부터만 비우기
        _clear_sheet_rows(ws_hold, 2)

        r = 2
        for row in hold_rows:
            _write_text_cell(ws_hold, r, 1, row.get("학년", ""))
            _write_text_cell(ws_hold, r, 2, row.get("반", ""))
            _write_text_cell(ws_hold, r, 3, row.get("성명", ""))
            _write_text_cell(ws_hold, r, 4, row.get("보류사유", ""))
            r += 1

        _move_sheet_after(wb, "퇴원_보류", "퇴원")
    else:
        # 템플릿에 이미 있는 경우는 삭제
        if "퇴원_보류" in wb.sheetnames:
            wb.remove(wb["퇴원_보류"])

    from openpyxl.styles import Font, Alignment

    def _format_sheet(ws):
        for rr in range(1, ws.max_row + 1):
            for cc in range(1, ws.max_column + 1):
                cell = ws.cell(rr, cc)
                cell.font = Font(size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center")

    _format_sheet(ws_done)
    if ws_hold is not None:
        _format_sheet(ws_hold)

def _school_kind_from_name(school_name: str) -> Tuple[str, str]:
    s = (school_name or "").strip()
    if not s:
        return "", ""
    last = s[-1]
    if last == "초":
        return "초등부", "초"
    if last == "중":
        return "중등부", "중"
    if last == "고":
        return "고등부", "고"
    return "", ""

def _write_transfer_hold_sheet(wb, hold_rows: List[Dict]):
    sheet_name = "전입생_보류"
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)

    ws.delete_rows(1, ws.max_row)

    _write_text_cell(ws, 1, 1, "학년")
    _write_text_cell(ws, 1, 2, "반")
    _write_text_cell(ws, 1, 3, "번호")
    _write_text_cell(ws, 1, 4, "성명")
    _write_text_cell(ws, 1, 5, "보류사유")

    r = 2
    for row in hold_rows:
        _write_text_cell(ws, r, 1, row.get("grade", ""))
        _write_text_cell(ws, r, 2, row.get("class", ""))
        _write_text_cell(ws, r, 3, row.get("number", ""))
        _write_text_cell(ws, r, 4, row.get("name", ""))
        _write_text_cell(ws, r, 5, row.get("보류사유", ""))
        r += 1

def _make_register_class_name(grade_i: int, class_value: Any) -> str:
    """
    등록파일 [학생자료] 수강반 표기 통일:
    - "1-1", "01-01" 같이 '학년-반'이 이미 들어온 경우 → 둘 다 0 제거 → "1-1"
    - "1", "01" 처럼 숫자만 있는 경우 → 학년-반 조합 → "1-1"
    - 그 외 문자열은 일단 grade와 그냥 붙임.
    """
    if class_value is None:
        return ""

    s = str(class_value).strip()
    if not s:
        return ""

    # 1) 이미 "학년-반" 형태인 경우 (01-01, 1-01, 01-1 등)
    m = re.match(r"^\s*0*(\d+)\s*-\s*0*(\d+)\s*$", s)
    if m:
        g = int(m.group(1))
        c = int(m.group(2))
        return f"{g}-{c}"  # → 01-01, 1-01, 01-1 전부 "1-1"

    # 2) 숫자만 있는 경우 ("1", "01", "03" 등) → 같은 학년 안에서 반 번호로 해석
    m2 = re.match(r"^\s*0*(\d+)\s*$", s)
    if m2:
        c = int(m2.group(1))
        return f"{grade_i}-{c}"

    # 3) 그 외 복잡한 문자열이면 일단 있는 그대로 붙인다
    return f"{grade_i}-{s}"

def _fill_register(
    template_path: Path,
    out_path: Path,
    school_name: str,
    year: str,
    freshmen_rows: List[Dict],
    transfer_done_rows: List[Dict],
    teacher_rows: List[Dict],
    transfer_hold_rows: Optional[List[Dict]] = None,
    withdraw_done_rows: Optional[List[Dict]] = None,
    withdraw_hold_rows: Optional[List[Dict]] = None,
) -> None:
    _ensure_xlsx_only(template_path)

    wb = load_workbook(template_path)
    ws_students = wb["학생자료"]
    ws_staff = wb["직원정보"]
    ws_groups = wb["그룹반정보"]

    # =========================
    # [학생자료] 컬럼 매핑
    # =========================
    hm = _header_map(ws_students, 1)
    need = ["No", "학생이름", "ID", "학교구분", "학교", "학년", "수강반"]
    for k in need:
        if k not in hm:
            raise ValueError(f"[오류] 템플릿 [학생자료]에 '{k}' 헤더가 없습니다.")

    col_no = hm["No"]
    col_name = hm["학생이름"]
    col_id = hm["ID"]
    col_kind = hm["학교구분"]
    col_school = hm["학교"]
    col_grade = hm["학년"]
    col_class = hm["수강반"]

    # 기존 데이터 clear
    for r in range(2, ws_students.max_row + 1):
        for c in [col_no, col_name, col_id, col_kind, col_school, col_grade, col_class]:
            ws_students.cell(row=r, column=c).value = None

    kind_full, kind_prefix = _school_kind_from_name(school_name)

    def write_student_row(r: int, no: int, name: str, uid: str, grade_i: int, cls_name: str):
        _write_text_cell(ws_students, r, col_no, no)
        _write_text_cell(ws_students, r, col_name, name)
        _write_text_cell(ws_students, r, col_id, uid)
        _write_text_cell(ws_students, r, col_kind, kind_full if kind_full else "")
        _write_text_cell(ws_students, r, col_school, school_name)
        _write_text_cell(ws_students, r, col_grade, f"{kind_prefix}{grade_i}" if kind_prefix else "")
        _write_text_cell(ws_students, r, col_class, cls_name)

    write_row = 2
    running_no = 1

    # 신입생 ID: 학년도 + 이름(중복 suffix 포함)
    fn_names = [r["name"] for r in freshmen_rows]
    fn_names_sfx = _apply_suffix_for_duplicates(fn_names)
    fn_ids = [f"{year}{nm}" for nm in fn_names_sfx]

    for i, fr in enumerate(freshmen_rows):
        r = write_row + i
        write_student_row(
            r=r,
            no=running_no,
            name=fr["name"],
            uid=fn_ids[i],
            grade_i=fr["grade"],
            cls_name=_make_register_class_name(fr["grade"], fr["class"]),
            )
        running_no += 1
    write_row += len(freshmen_rows)

    # 전입생(완료)
    for tr in transfer_done_rows:
        r = write_row
        write_student_row(
            r=r,
            no=running_no,
            name=tr["name"],
            uid=tr["id"],
            grade_i=tr["grade"],
            cls_name=_make_register_class_name(tr["grade"], tr["class"]),
        )
        running_no += 1
        write_row += 1

    # 선생님(학습용) → 학생자료에 "선생님반"
    teachers_learn = [t for t in teacher_rows if t["learn_apply"]]
    t_names = [t["name"] for t in teachers_learn]
    t_names_sfx = _apply_suffix_for_duplicates(t_names)
    t_ids = [f"{nm}1" for nm in t_names_sfx]

    for j, t in enumerate(teachers_learn):
        r = write_row + j
        write_student_row(
            r=r,
            no=running_no,
            name=t["name"],
            uid=t_ids[j],
            grade_i=1,
            cls_name="선생님반",
        )
        running_no += 1
    write_row += len(teachers_learn)

    # =========================
    # [직원정보]
    # =========================
    hm2 = _header_map(ws_staff, 1)
    hm2_lower = {k.lower(): v for k, v in hm2.items()}

    need2 = ["no", "이름", "아이디", "권한부여"]
    for k in need2:
        if k.lower() not in hm2_lower:
            raise ValueError(f"[오류] 템플릿 [직원정보]에 '{k}' 헤더가 없습니다.")

    col_s_no = hm2_lower["no"]
    col_s_name = hm2_lower["이름"]
    col_s_id = hm2_lower["아이디"]
    col_s_role = hm2_lower["권한부여"]

    for r in range(2, ws_staff.max_row + 1):
        for c in [col_s_no, col_s_name, col_s_id, col_s_role]:
            ws_staff.cell(row=r, column=c).value = None  # 템플릿 클리어는 그냥 둬도 됨

    teachers_admin = [t for t in teacher_rows if t["admin_apply"]]
    a_names = [t["name"] for t in teachers_admin]
    a_names_sfx = _apply_suffix_for_duplicates(a_names)

    staff_write = 2

    for i, t in enumerate(teachers_admin):
        r = staff_write + i
        _write_text_cell(ws_staff, r, col_s_no, i + 1)
        _write_text_cell(ws_staff, r, col_s_name, t["name"])
        _write_text_cell(ws_staff, r, col_s_id, a_names_sfx[i])
        _write_text_cell(ws_staff, r, col_s_role, "선생님")

    # =========================
    # [그룹반정보]
    # =========================
    hm_g = _header_map(ws_groups, 1)
    need_g = ["그룹명", "반명", "수강료", "담임명", "FullMode"]
    for k in need_g:
        if k not in hm_g:
            raise ValueError(f"[오류] 템플릿 [그룹반정보]에 '{k}' 헤더가 없습니다.")

    col_g_group = hm_g["그룹명"]
    col_g_class = hm_g["반명"]
    col_g_fee = hm_g["수강료"]
    col_g_teacher = hm_g["담임명"]
    col_g_full = hm_g["FullMode"]

    for r in range(2, ws_groups.max_row + 1):
        for c in [col_g_group, col_g_class, col_g_fee, col_g_teacher, col_g_full]:
            ws_groups.cell(row=r, column=c).value = None

    class_set = set()
    last_student_row = _find_last_data_row(ws_students, key_col=col_no, start_row=2)
    for r in range(2, last_student_row + 1):
        v = ws_students.cell(row=r, column=col_class).value
        if v is None:
            continue
        s = str(v).strip()
        if s:
            class_set.add(s)

    def parse_grade_class(class_name: str):
        """
        '1-1', '3-10' 형태에서 (학년, 반)을 정수로 추출.
        반이 숫자가 아니면 (학년, None).
        """
        s = str(class_name)

        # 숫자-숫자 (예: 1-10)
        m = re.match(r"^\s*(\d+)\s*-\s*(\d+)\s*$", s)
        if m:
            return int(m.group(1)), int(m.group(2))

        # 숫자-문자 (예: 1-A반)
        m = re.match(r"^\s*(\d+)\s*-\s*(.+)\s*$", s)
        if m:
            return int(m.group(1)), None

        return None, None


    def group_name_from_class(class_name: str) -> str:
        g, _ = parse_grade_class(class_name)
        if g is None:
            return "기타"
        return f"{g}학년"


    def class_sort_key(class_name: str):
        if class_name == "선생님반":
            return (2, 0, 0, "zzz")

        g, c = parse_grade_class(class_name)

        if g is None:
            return (1, 0, 0, str(class_name))

        # 반 번호 숫자 정렬
        class_order = c if c is not None else 9999

        return (0, g, class_order, str(class_name))

    class_list = sorted(class_set, key=class_sort_key)

    start_r = 2
    r = start_r
    for cls_name in class_list:
        _write_text_cell(ws_groups, r, col_g_group, group_name_from_class(cls_name))
        _write_text_cell(ws_groups, r, col_g_class, cls_name)
        ws_groups.cell(r, col_g_fee).value = None
        _write_text_cell(ws_groups, r, col_g_teacher, "선생님")
        _write_text_cell(ws_groups, r, col_g_full, "Y")
        r += 1

    # 전입 보류 시트
    if transfer_hold_rows:
        _write_transfer_hold_sheet(wb, transfer_hold_rows)

    # 전출 완료/보류 시트
    if (withdraw_done_rows is not None) and (withdraw_hold_rows is not None):
        _write_withdraw_to_register(wb, withdraw_done_rows, withdraw_hold_rows)

    # 워크북 전체: 빈 행 아래 서식 제거 + A1로 통일
    _clear_format_workbook_from_row(wb, start_row=2)
    _reset_view_to_a1(wb)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    _backup_if_exists(out_path)
    wb.save(out_path)

def _parse_grade_class_from_register(raw: Any) -> Tuple[Optional[int], str]:
    """
    register의 수강반 컬럼 해석용.

    - "1-3", "01-05" 같은 문자열 → (1, "3"), (1, "05")
    - 엑셀이 날짜로 인식한 값(01-01 → datetime/date)도
      다시 "mm-dd" 형태로 복원해서 처리.
    - 패턴이 안 맞으면 (None, 원본 문자열)을 그대로 돌려줘서
      안내문에는 원래 텍스트가 찍히게 함.
    """
    if raw is None:
        return None, ""

    # 1) 엑셀이 01-01을 날짜로 바꿔버린 경우: datetime/date → "mm-dd"
    if isinstance(raw, (datetime, date)):
        s = raw.strftime("%m-%d")
    else:
        s = str(raw).strip()

    if not s:
        return None, ""

    # "1-3", "01-05", "01-01" 같은 패턴
    m = re.match(r"^\s*0*(\d+)\s*-\s*0*(\d+)\s*$", s)
    if not m:
        # 패턴이 아예 안 맞으면 문자열 그대로 쓰게
        return None, s

    grade = int(m.group(1))           # 앞자리 0 날림 → 01 → 1
    cls_str = str(int(m.group(2)))    # 반도 01 → 1

    return grade, cls_str

def _build_notice_student_sheet(
    ws_notice,
    register_students_ws,
    transfer_ids: set,
    transfer_dup_ids: set,   # 🔸 추가

):
    """
    안내파일 - 학생 ID,PW(학습용)
    헤더 3행: No., 학년, 반, 학생이름, ID, PW
    데이터 4행부터
    """
    hm_r = _header_map(register_students_ws, 1)

    # 🔹 필수 헤더: 예전처럼 No 포함
    need_r = ["No", "학생이름", "ID", "수강반"]
    for k in need_r:
        if k not in hm_r:
            raise ValueError(f"[오류] 등록작업파일 [학생자료]에 '{k}' 헤더가 없습니다.")

    c_r_name = hm_r["학생이름"]
    c_r_id   = hm_r["ID"]
    c_r_cls  = hm_r["수강반"]
    c_r_no   = hm_r["No"]

    header_row = 3
    start_row = 4

    # 0) 기존 데이터 '값'만 지우기 (그대로 유지)
    max_row = ws_notice.max_row
    if max_row >= start_row:
        for r in range(start_row, max_row + 1):
            for c in range(1, 7):  # No~PW
                ws_notice.cell(row=r, column=c).value = None

    # 1) 1차 패스: 안내에 들어갈 학생 목록 + (학년, 이름키) 카운트
    tmp_rows: List[Dict[str, Any]] = []
    name_counter: Counter[tuple] = Counter()

    # 🔵 last_r를 다시 No 기준으로
    last_r = _find_last_data_row(register_students_ws, key_col=c_r_no, start_row=2)
    for r in range(2, last_r + 1):
        nm  = register_students_ws.cell(r, c_r_name).value
        uid = register_students_ws.cell(r, c_r_id).value
        cls = register_students_ws.cell(r, c_r_cls).value

        cls_str = "" if cls is None else str(cls).strip()
        if cls_str == "선생님반":
            continue

        if (nm is None or str(nm).strip() == "") and (uid is None or str(uid).strip() == ""):
            continue

        nm_s  = "" if nm  is None else str(nm).strip()
        uid_s = "" if uid is None else str(uid).strip()
        if not uid_s:
            # ID 없는 행은 안내 파일에서도 제외
            continue

        grade, cls_only = _parse_grade_class_from_register(cls)
        if grade is None:
            g_disp = ""
            cls_disp = cls_str
        else:
            g_disp = grade
            cls_disp = cls_only

        name_key = _notice_name_key(nm_s)
        key = (grade, name_key)

        tmp_rows.append(
            {
                "grade": g_disp,
                "class_disp": cls_disp,
                "name": nm_s,
                "id": uid_s,
                "key": key,
                "is_transfer": uid_s in transfer_ids,
                "is_transfer_dup_with_roster": uid_s in transfer_dup_ids,  # 🔸 추가
            }
        )

        if grade is not None and name_key:
            name_counter[key] += 1

    # 2) 2차 패스: 실제 안내 시트에 쓰면서 색칠
    cur_row = start_row
    running_no = 1

    for rec in tmp_rows:
        key = rec["key"]
        # 1) 기본: 등록파일 내부 (학년, 이름키) 카운트로 동명이인 판정
        dup_flag = name_counter.get(key, 0) >= 2

        # 2) 전입생이고, 명부 기준 동명이인으로 판정된 경우 → 무조건 동명이인 처리
        if rec["is_transfer"] and rec.get("is_transfer_dup_with_roster"):
            dup_flag = True

        # 3) A~F 열 전체를 텍스트로 쓰면서 셀 객체 확보
        cell_no    = _write_text_cell(ws_notice, cur_row, 1, running_no)        # No.
        cell_grade = _write_text_cell(ws_notice, cur_row, 2, rec["grade"])      # 학년
        cell_class = _write_text_cell(ws_notice, cur_row, 3, rec["class_disp"]) # 반
        cell_name  = _write_text_cell(ws_notice, cur_row, 4, rec["name"])       # 이름
        cell_id    = _write_text_cell(ws_notice, cur_row, 5, rec["id"])         # ID

        # PW 컬럼(6열)이 있으면 필요 시 값 넣기 (없으면 None / 빈값)
        cell_pw    = _write_text_cell(ws_notice, cur_row, 6, "1234")

        # 이 행에서 색칠할 대상 셀 전체 (A~F)
        row_cells = [cell_no, cell_grade, cell_class, cell_name, cell_id, cell_pw]

        # 4) 전입생: 행 전체 주황
        if rec["is_transfer"]:
            for cell in row_cells:
                cell.fill = FILL_TRANSFER

        # 5) 동명이인: 행 전체 노랑 (전입+동명이인이면 노랑으로 덮어씀)
        if dup_flag:
            for cell in row_cells:
                cell.fill = FILL_DUP

        running_no += 1
        cur_row += 1

def _build_notice_teacher_sheet(
    ws_notice,
    teacher_rows: List[Dict],
    learn_ids: Optional[List[str]] = None,
    admin_ids: Optional[List[str]] = None,
):
    """
    안내파일 - 선생님ID,PW(관리용,학습용)
    헤더 3행, 데이터 4행부터.
    - No, 직위, 선생님이름: teacher_rows의 position/name 그대로
    - 관리용ID: 등록파일 [직원정보]에서 가져온 ID (fallback: name)
    - 학습용ID: 등록파일 [학생자료] 선생님반에서 가져온 ID (fallback: name+'1')
    - 신청 안 한 칸은 회색 처리
    """
    header_row = 3
    start_row = 4

    # 직위(B열) 컬럼 폭 확장 (긴 직위/담당 명칭 잘리지 않도록)
    try:
        ws_notice.column_dimensions["B"].width = 16.6
    except Exception:
        pass

    # 전체 교사 중 신청자 수
    admin_total = sum(1 for t in teacher_rows if t.get("admin_apply"))
    learn_total = sum(1 for t in teacher_rows if t.get("learn_apply"))

    admin_ids_list = admin_ids or []
    learn_ids_list = learn_ids or []

    # 등록파일에서 가져온 ID 길이가 신청자 수와 맞으면 그대로 사용
    use_admin_from_reg = admin_total > 0 and len(admin_ids_list) >= admin_total
    use_learn_from_reg = learn_total > 0 and len(learn_ids_list) >= learn_total

    idx_admin = 0
    idx_learn = 0

    r_out = start_row
    no = 1
    for t in teacher_rows:
        pos = "" if t.get("position") is None else str(t.get("position")).strip()
        nm  = "" if t.get("name") is None else str(t.get("name")).strip()
        if not nm and not pos and (not t.get("learn_apply")) and (not t.get("admin_apply")):
            continue

        admin_apply = bool(t.get("admin_apply"))
        learn_apply = bool(t.get("learn_apply"))

        # ----- 관리용 ID: 등록파일 우선 -----
        admin_id = ""
        if admin_apply:
            if use_admin_from_reg:
                admin_id = admin_ids_list[idx_admin]
            else:
                admin_id = nm  # fallback: 예전 방식
            idx_admin += 1
        admin_pw = "t1234" if admin_id else ""

        # ----- 학습용 ID: 등록파일 우선 -----
        learn_id = ""
        if learn_apply:
            if use_learn_from_reg:
                learn_id = learn_ids_list[idx_learn]
            else:
                learn_id = f"{nm}1"  # fallback: 예전 방식
            idx_learn += 1
        learn_pw = "1234" if learn_id else ""

        # A: No. / B: 직위 / C: 선생님이름 / D: 구분용 빈 칸
        # E: 관리용 ID / F: PW / G: 구분용 빈 칸 / H: 학습용 ID / I: PW
        _write_text_cell(ws_notice, r_out, 1, no)
        _write_text_cell(ws_notice, r_out, 2, pos)
        _write_text_cell(ws_notice, r_out, 3, nm)
        _write_text_cell(ws_notice, r_out, 5, admin_id)
        _write_text_cell(ws_notice, r_out, 6, admin_pw)
        _write_text_cell(ws_notice, r_out, 8, learn_id)
        _write_text_cell(ws_notice, r_out, 9, learn_pw)

        # 회색 처리(신청 안 한 영역)
        if not admin_apply:
            for c in [5, 6]:
                ws_notice.cell(r_out, c).fill = FILL_GREY

        if not learn_apply:
            for c in [8, 9]:
                ws_notice.cell(r_out, c).fill = FILL_GREY

        no += 1
        r_out += 1

    _delete_rows_below(ws_notice, r_out - 1)

def _build_notice_file(
    template_notice_path: Path,
    out_notice_path: Path,
    out_register_path: Path,
    teacher_file_path: Optional[Path],
    transfer_done_rows: List[Dict],
) -> None:
    _ensure_xlsx_only(template_notice_path)
    _ensure_xlsx_only(out_register_path)

    wb_notice = _safe_load_workbook(template_notice_path, data_only=False)
    wb_reg = load_workbook(out_register_path)

    if "학생자료" not in wb_reg.sheetnames:
        raise PipelineError(
            code="NOTICE_REGISTER_STUDENT_SHEET_MISSING",
            message="등록작업파일에 '학생자료' 시트가 없습니다.",
            context={"register_path": str(out_register_path)},
        )

    ws_reg_students = wb_reg["학생자료"]

    def _norm_sheetname(s: str) -> str:
        s = s or ""
        s = str(s)
        s = s.replace("\u00A0", " ")
        s = re.sub(r"\s+", "", s)
        return s

    def _pick_sheet_by_keywords(wb, keywords: List[str]) -> str:
        keys = [_norm_sheetname(k) for k in keywords]
        for name in wb.sheetnames:
            n = _norm_sheetname(name)
            if all(k in n for k in keys):
                return name
        # 템플릿 구조 문제 → PipelineError로 승격
        raise PipelineError(
            code="NOTICE_TEMPLATE_SHEET_NOT_FOUND",
            message="안내 템플릿에서 필요한 시트를 찾지 못했습니다.",
            context={
                "keywords": keywords,
                "sheetnames": list(wb.sheetnames),
            },
        )
    
    sh_student = _pick_sheet_by_keywords(wb_notice, ["학생", "PW", "학습용"])
    sh_teacher = _pick_sheet_by_keywords(wb_notice, ["선생님", "PW"])

    ws_notice_students = wb_notice[sh_student]
    ws_notice_teachers = wb_notice[sh_teacher]

    # 1) 등록작업파일 학생자료 시트
    ws_reg_students = wb_reg["학생자료"]

    # 1-1) 전입 완료 학생 ID set (학생 안내 시트 색칠용)
    transfer_ids: set[str] = set()
    transfer_dup_ids: set[str] = set()   # 🔸 명부 기준 동명이인 전입 ID만

    for tr in transfer_done_rows:
        uid = tr.get("id")
        if not uid:
            continue
        uid_s = str(uid).strip()
        transfer_ids.add(uid_s)

        if tr.get("dup_with_roster"):    # 1단계에서 붙인 플래그
            transfer_dup_ids.add(uid_s)

    # 1-2) 학생 안내 시트 생성
    _build_notice_student_sheet(
        ws_notice=ws_notice_students,
        register_students_ws=ws_reg_students,
        transfer_ids=transfer_ids,
        transfer_dup_ids=transfer_dup_ids,  # 🔸 추가

    )

    # --- 등록작업파일에서 실제 ID 가져오기 (교사 안내용) ---

    # 2) 학생자료 시트에서 선생님반 학습용 ID
    learn_ids_from_register: Optional[List[str]] = None
    try:
        hm_r = _header_map(ws_reg_students, 1)
        col_r_class = hm_r["수강반"]
        col_r_id    = hm_r["ID"]
        tmp_learn: List[str] = []
        max_row = ws_reg_students.max_row or 1
        for row in range(2, max_row + 1):
            cls_val = ws_reg_students.cell(row=row, column=col_r_class).value
            id_val  = ws_reg_students.cell(row=row, column=col_r_id).value
            if cls_val is None and id_val is None:
                continue
            if str(cls_val).strip() == "선생님반" and id_val:
                tmp_learn.append(str(id_val).strip())
        if tmp_learn:
            learn_ids_from_register = tmp_learn
    except Exception:
        learn_ids_from_register = None

    # 3) 직원정보 시트에서 관리용 ID
    admin_ids_from_register: Optional[List[str]] = None
    try:
        if "직원정보" in wb_reg.sheetnames:
            ws_reg_staff = wb_reg["직원정보"]
            hm_s = _header_map(ws_reg_staff, 1)
            col_s_id = hm_s["아이디"]
            tmp_admin: List[str] = []
            max_row = ws_reg_staff.max_row or 1
            for row in range(2, max_row + 1):
                id_val = ws_reg_staff.cell(row=row, column=col_s_id).value
                if not id_val:
                    continue
                tmp_admin.append(str(id_val).strip())
            if tmp_admin:
                admin_ids_from_register = tmp_admin
    except Exception:
        admin_ids_from_register = None

    # 4) 교사 안내 시트 생성
    teacher_rows = _read_teacher_rows(teacher_file_path) if teacher_file_path else []
    _build_notice_teacher_sheet(
        ws_notice=ws_notice_teachers,
        teacher_rows=teacher_rows,
        learn_ids=learn_ids_from_register,
        admin_ids=admin_ids_from_register,
)

    out_notice_path.parent.mkdir(parents=True, exist_ok=True)
    _backup_if_exists(out_notice_path)

    # 안내 파일도 워크북 공통 규칙 적용
    _clear_format_workbook_from_row(wb_notice, start_row=4)
    _reset_view_to_a1(wb_notice)

    wb_notice.save(out_notice_path)

def render_mail_text(
    mail_template_text: str,
    school_name: str,
    domain: str,
) -> str:
    """
    텍스트 파일 내부:
    - 'OO초'/'OO중'/'OO고' 같은 표현 → school_name
    - 'OOOOO.readinggate.com' → domain
    """
    txt = mail_template_text or ""
    if school_name:
        txt = txt.replace("OO초", school_name).replace("OO중", school_name).replace("OO고", school_name)
    if domain:
        txt = re.sub(r"[A-Za-z0-9\-]+\.readinggate\.com", domain, txt)
    return txt






# ========== L5: orchestrator (scan / execute / run) ==========

def get_project_dirs(work_root: Path) -> Dict[str, Path]:
    """
    작업 폴더(work_root) 구조:

    work_root/
      ├─ ●resources/  (또는 이름에 'resources' 포함된 아무 폴더 1개)
      │    ├─ DB/
      │    ├─ templates/
      │    └─ notices/
      ├─ A초등학교/
      ├─ B중학교/
      └─ ...
    """
    work_root = work_root.resolve()

    # 이름에 'resources' 가 들어간 폴더들을 모두 수집
    candidates = [
        p for p in work_root.iterdir()
        if p.is_dir() and "resources" in p.name.lower()
    ]

    if len(candidates) == 0:
        # 아무것도 없으면 기본값: work_root/resources
        resources_root = work_root / "resources"
    elif len(candidates) == 1:
        resources_root = candidates[0]
    else:
        # 여러 개면 애매하니까 바로 에러
        names = [p.name for p in candidates]
        raise ValueError(
            f"[오류] 작업 폴더 내에 'resources'를 포함한 폴더가 여러 개 있습니다: {names}"
        )

    return {
        "WORK_ROOT": work_root,
        "RESOURCES_ROOT": resources_root,
        "DB": resources_root / "DB",
        "TEMPLATES": resources_root / "templates",
        "NOTICES": resources_root / "notices",
        "SCHOOL_ROOT": work_root,  # 학교 폴더는 work_root 바로 아래
    }

def find_templates(format_dir: Path) -> Tuple[Optional[Path], Optional[Path], List[str]]:
    """
    [templates] 폴더 템플릿 2개 식별:
    - 등록 템플릿: 파일명에 '등록' 포함
    - 안내 템플릿: 파일명에 '안내' 포함
    """
    format_dir = Path(format_dir).resolve()
    if not format_dir.exists():
        return None, None, [f"[오류] [templates] 폴더를 찾을 수 없습니다: {format_dir}"]

    xlsx_files = [
        p for p in format_dir.iterdir()
        if p.is_file() and p.suffix.lower() == ".xlsx" and not p.name.startswith("~$")
    ]
    if not xlsx_files:
        return None, None, [f"[오류] [templates] 폴더에 .xlsx 파일이 없습니다: {format_dir}"]

    reg = [p for p in xlsx_files if "등록" in p.stem]
    notice = [p for p in xlsx_files if "안내" in p.stem]

    errors: List[str] = []
    if len(reg) == 0:
        errors.append("[오류] [templates] 폴더에서 '등록' 템플릿을 찾지 못했습니다. (파일명에 '등록' 포함)")
    elif len(reg) > 1:
        errors.append("[오류] [templates] 폴더에 '등록' 템플릿이 여러 개입니다.")

    if len(notice) == 0:
        errors.append("[오류] [templates] 폴더에서 '안내' 템플릿을 찾지 못했습니다. (파일명에 '안내' 포함)")
    elif len(notice) > 1:
        errors.append("[오류] [templates] 폴더에 '안내' 템플릿이 여러 개입니다.")

    if errors:
        return None, None, errors

    return reg[0], notice[0], []

def scan_work_root(work_root: Path) -> Dict[str, Any]:
    """
    작업 루트에서 resources/DB, resources/templates, resources/notices, 학교 폴더 상태를 점검한다.
    app.py는 여기서 다음 키들을 기대하고 있음:

      - ok: bool
      - errors: List[str]
      - message: str
      - school_folders: List[str]
      - notice_titles: List[str]

      - db_ok: bool
      - errors_db: List[str]
      - db_file: Optional[Path]

      - format_ok: bool
      - errors_format: List[str]
      - register_template: Optional[Path]
      - notice_template: Optional[Path]
    """
    work_root = work_root.resolve()
    dirs = get_project_dirs(work_root)

    # 전체 에러
    errors: List[str] = []

    # -------------------------
    # 0. resources 루트
    # -------------------------
    res_root = dirs["RESOURCES_ROOT"].resolve()

    # 학교 폴더 목록 (resources 폴더 제외)
    school_folders = [
        p.name
        for p in work_root.iterdir()
        if p.is_dir()
        and p.resolve() != res_root
        and not p.name.startswith(".")
    ]
    school_folders.sort()

    # -------------------------
    # 1. DB 폴더 점검
    # -------------------------
    db_ok = False
    errors_db: List[str] = []
    db_file: Optional[Path] = None

    db_dir = dirs["DB"]
    if not db_dir.exists():
        errors_db.append("[오류] resources/DB 폴더가 없습니다.")
    else:
        db_files = [
            p for p in db_dir.glob("*.xlsb")
            if "학교전체명단" in p.stem and not p.name.startswith("~$")
        ]
        if len(db_files) == 0:
            errors_db.append("[오류] DB 폴더에 '학교전체명단' xlsb 파일이 없습니다.")
        elif len(db_files) > 1:
            errors_db.append("[오류] DB 폴더에 '학교전체명단' xlsb 파일이 2개 이상입니다.")
        else:
            db_ok = True
            db_file = db_files[0]

    # -------------------------
    # 2. templates(양식) 폴더 점검
    # -------------------------
    format_ok = False
    errors_format: List[str] = []
    register_template: Optional[Path] = None
    notice_template: Optional[Path] = None

    tpl_dir = dirs["TEMPLATES"]
    if not tpl_dir.exists():
        errors_format.append("[오류] resources/templates 폴더가 없습니다.")
    else:
        reg_files = [
            p for p in tpl_dir.glob("*.xlsx")
            if "등록" in p.stem and not p.name.startswith("~$")
        ]
        notice_files = [
            p for p in tpl_dir.glob("*.xlsx")
            if "안내" in p.stem and not p.name.startswith("~$")
        ]

        if len(reg_files) != 1:
            errors_format.append("templates 폴더에 '등록' 템플릿 파일이 정확히 1개 있어야 합니다.")
        else:
            register_template = reg_files[0]

        if len(notice_files) != 1:
            errors_format.append("templates 폴더에 '안내' 템플릿 파일이 정확히 1개 있어야 합니다.")
        else:
            notice_template = notice_files[0]

        if not errors_format:
            format_ok = True

    # -------------------------
    # 3. notices 폴더 점검
    # -------------------------
    notice_dir = dirs["NOTICES"]
    notice_titles: List[str] = []

    if not notice_dir.exists():
        errors.append("[오류] resources/notices 폴더가 없습니다.")
    else:
        txt_files = [p for p in notice_dir.glob("*.txt") if p.is_file()]
        if not txt_files:
            errors.append("[오류] notices 폴더에 .txt 파일이 없습니다.")
        else:
            notice_titles = sorted({p.stem.strip() for p in txt_files})

    # -------------------------
    # 4. 전체 에러 합치기
    # -------------------------
    errors.extend(errors_db)
    errors.extend(errors_format)

    ok = len(errors) == 0
    message = (
        "[OK] resources(DB/templates/notices)가 정상적으로 준비되었습니다."
        if ok else ""
    )

    return {
        "ok": ok,
        "errors": errors,
        "message": message,
        "school_folders": school_folders,
        "notice_titles": notice_titles,

        # DB 상태 (app.py에서 사용)
        "db_ok": db_ok,
        "errors_db": errors_db,
        "db_file": db_file,

        # 양식 상태 (app.py에서 사용)
        "format_ok": format_ok,
        "errors_format": errors_format,
        "register_template": register_template,
        "notice_template": notice_template,
    }

def find_single_input_file(input_dir: Path, keywords: Sequence[str]) -> Optional[Path]:
    if not input_dir.exists():
        return None

    kw_list: List[str] = []
    for k in keywords:
        k = "" if k is None else str(k).strip()
        if k:
            kw_list.append(k)

    if not kw_list:
        return None

    candidates: List[Path] = []
    for p in input_dir.iterdir():
        if not (p.is_file() and p.suffix.lower() == ".xlsx"):
            continue
        if p.name.startswith("~$"):
            continue
        if any(text_contains(p.name, kw) for kw in kw_list):
            candidates.append(p)

    if len(candidates) == 0:
        return None
    if len(candidates) > 1:
        raise ValueError(f"[오류] {kw_list} 포함 .xlsx 파일이 2개 이상: {[c.name for c in candidates]}")
    return candidates[0]

def choose_template_register(format_dir: Path, year_str: str = "") -> Path:
    reg, notice, errors = find_templates(format_dir)
    if errors:
        raise ValueError(errors[0])
    assert reg is not None
    return reg

def choose_template_notice(format_dir: Path, year_str: str = "") -> Path:
    reg, notice, errors = find_templates(format_dir)
    if errors:
        raise ValueError(errors[-1])
    assert notice is not None
    return notice

def choose_db_xlsb(db_dir: Path) -> Path:
    if not db_dir.exists():
        raise ValueError(f"[오류] DB 폴더가 없습니다: {db_dir}")

    xlsb_files = [
        p for p in db_dir.iterdir()
        if p.is_file() and p.suffix.lower() == ".xlsb" and not p.name.startswith("~$")
    ]
    if not xlsb_files:
        raise ValueError("[오류] DB 폴더에 .xlsb 파일이 없습니다.")
    xlsb_files.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return xlsb_files[0]

def search_schools_in_db(work_root: Path, keyword: str, limit: int = 30) -> List[str]:
    work_root = Path(work_root).resolve()
    dirs = get_project_dirs(work_root)
    db_path = choose_db_xlsb(dirs["DB"])

    kw = (keyword or "").strip()
    if not kw:
        return []

    kw_norm = normalize_text(kw)

    results: List[str] = []
    seen = set()

    with open_xlsb_workbook(str(db_path)) as wb:
        sheet_names = wb.sheets
        if not sheet_names:
            return []
        with wb.get_sheet(sheet_names[0]) as sh:
            for r_idx, row in enumerate(sh.rows()):
                if r_idx < 8:
                    continue
                if len(row) <= 4:
                    continue
                v = row[4].v  # E열
                if v is None:
                    continue
                s = str(v).strip()
                if not s:
                    continue

                if kw_norm and (kw_norm in normalize_text(s)) and s not in seen:
                    seen.add(s)
                    results.append(s)
                    if len(results) >= limit:
                        break

    return results

def school_exists_in_db(db_dir: Path, school_name: str) -> Path:
    db_path = choose_db_xlsb(db_dir)

    target = (school_name or "").strip()
    if not target:
        raise ValueError("[오류] 학교명이 비어 있습니다(DB 검증 불가).")

    target_norm = normalize_text(target)
    found = False

    with open_xlsb_workbook(str(db_path)) as wb:
        sheet_names = wb.sheets
        if not sheet_names:
            raise ValueError("[오류] DB xlsb에 시트가 없습니다.")
        with wb.get_sheet(sheet_names[0]) as sh:
            for r_idx, row in enumerate(sh.rows()):
                if r_idx < 8:
                    continue
                if len(row) <= 4:
                    continue
                v = row[4].v  # E열
                if v is None:
                    continue
                cell = str(v).strip()
                if not cell:
                    continue
                cell_norm = normalize_text(cell)
                if target_norm and cell_norm and (target_norm in cell_norm):
                    found = True
                    break

    if not found:
        raise ValueError(f"[오류] DB(E열 9행~)에서 학교명 '{target}' 포함 항목을 찾지 못했습니다.")

    return db_path

def _normalize_domain(raw: str) -> str:
    if raw is None:
        return ""
    s = str(raw).strip()
    if not s:
        return ""
    s = re.sub(r"^https?://", "", s, flags=re.I)
    s = s.split("/")[0].strip()
    return s

def get_school_domain_from_db(db_dir: Path, school_name: str) -> Optional[str]:
    """
    DB xlsb에서:
    - E열: 학교명 매칭
    - F열: 홈페이지(리딩게이트 전용 도메인) 반환
    없으면 None
    """
    db_path = choose_db_xlsb(db_dir)
    target = (school_name or "").strip()
    if not target:
        return None
    target_norm = normalize_text(target)

    with open_xlsb_workbook(str(db_path)) as wb:
        sheet_names = wb.sheets
        if not sheet_names:
            return None
        with wb.get_sheet(sheet_names[0]) as sh:
            for r_idx, row in enumerate(sh.rows()):
                if r_idx < 8:
                    continue
                if len(row) <= 5:
                    continue
                ev = row[4].v  # E
                if ev is None:
                    continue
                ecell = str(ev).strip()
                if not ecell:
                    continue
                if target_norm and (target_norm in normalize_text(ecell)):
                    fv = row[5].v  # F
                    dom = _normalize_domain("" if fv is None else str(fv))
                    return dom if dom else None
    return None

def load_notice_templates(work_root: Path) -> dict[str, str]:
    dirs = get_project_dirs(work_root)
    notice_dir = dirs["NOTICES"]

    if not notice_dir.exists():
        return {}

    result = {}

    for p in notice_dir.glob("*.txt"):
        if not p.is_file():
            continue
        try:
            text = p.read_text(encoding="utf-8")
        except UnicodeDecodeError:
            text = p.read_text(encoding="utf-8-sig")

        result[p.stem.strip()] = text.strip()

    return result

def domain_missing_message(school_name: str) -> str:
    _, kind_prefix = _school_kind_from_name(school_name)
    kind_disp = kind_prefix if kind_prefix else "학교"
    return f"{kind_disp} (사용자가 작업중인) 의 도메인 주소가 존재하지 않습니다. 학교 전체 명단 파일을 확인하세요."

def scan_pipeline(
    work_root: Path,
    school_name: str,
    school_start_date: date,
    work_date: date,
    roster_basis_date: Optional[date] = None,
) -> ScanResult:
    # 문자열이 아니라 LogEntry로 로그를 쌓는다.
    logs: List[LogEntry] = []

    def log(
        msg: str,
        level: str = "INFO",
        code: str = "TEXT",
        context: Optional[Dict[str, Any]] = None,
    ) -> None:
        logs.append(
            LogEntry(
                level=level,
                code=code,
                message=msg,
                context=context or {},
            )
        )

    work_root = Path(work_root).resolve()
    dirs = get_project_dirs(work_root)

    school_name = (school_name or "").strip()
    year_str = str(school_start_date.year).strip()
    year_int = school_start_date.year

    sr = ScanResult(
        ok=False,
        school_name=school_name,
        year_str=year_str,
        year_int=year_int,
        project_root=work_root,
        input_dir=Path("."),
        output_dir=Path("."),
        template_register=None,
        template_notice=None,
        db_path=None,
        notice_templates_dir=None,
        roster_path=None,
        freshmen_path=None,
        transfer_path=None,
        withdraw_path=None,
        teacher_path=None,
        can_execute_after_input=False,
        logs=logs,  # 🔹 여기서도 LogEntry 리스트
    )

    try:
        if not school_name:
            raise ValueError("[오류] 학교명이 비어 있습니다.")
        year_int = int(year_str)
        sr.year_int = year_int

        db_path = school_exists_in_db(dirs["DB"], school_name)
        sr.db_path = db_path
        log(f"[OK] DB 검증 통과 | 사용 파일: {db_path.name}")

        # 🔹 학교 폴더: 선택한 학교명이 포함된 폴더 찾기 (정규화 포함 매칭)
        root_dir = dirs["SCHOOL_ROOT"]

        kw = (school_name or "").strip()
        if not kw:
            raise ValueError("[오류] 학교명이 비어 있어 학교 폴더를 찾을 수 없습니다.")

        matches = [
            p
            for p in root_dir.iterdir()
            if p.is_dir() and text_contains(p.name, kw)
        ]

        if not matches:
            raise ValueError(
                f"[오류] 작업 폴더 안에서 '{school_name}' 이(가) 포함된 학교 폴더를 찾지 못했습니다."
            )

        if len(matches) > 1:
            raise ValueError(
                f"[오류] '{school_name}' 이(가) 포함된 폴더가 여러 개입니다: "
                + ", ".join(p.name for p in matches)
            )

        school_dir = matches[0]

        # 🔹 로그: 어떤 폴더로 매칭됐는지 명확히 찍어줌
        log(f"[OK] 학교 폴더 매칭: {school_dir.name}")

        input_dir = school_dir
        output_dir = school_dir / "작업"

        sr.input_dir = input_dir
        sr.output_dir = output_dir

        # 🔹 로그: 학교 폴더 안 파일 목록 출력 (안전 처리)
        try:
            file_list = [p.name for p in input_dir.iterdir() if p.is_file()]
            log(f"[DEBUG] input files: {file_list}")
        except Exception as e:
            log(f"[WARN] 학교 폴더 파일 목록 조회 중 오류: {e}")

        # 👉 신입생 파일은 이제 "필수" 아님. None 허용.
        freshmen_file = find_single_input_file(input_dir, FRESHMEN_KEYWORDS)
        teacher_file  = find_single_input_file(input_dir, TEACHER_KEYWORDS)
        transfer_file = find_single_input_file(input_dir, TRANSFER_KEYWORDS)
        withdraw_file = find_single_input_file(input_dir, WITHDRAW_KEYWORDS)

        # 👉 신입/전입/전출/교사 키워드 들어간 파일이 하나도 없으면 에러
        if not any([freshmen_file, teacher_file, transfer_file, withdraw_file]):
            raise ValueError(
                "[오류] 신입생/전입생/전출생/교사 키워드가 들어간 입력 xlsx 파일을 하나도 찾지 못했습니다. "
                "학생명부만 있는 경우에는 전입·전출·교사 중 하나 이상의 파일이 필요합니다."
            )

        sr.freshmen_file = freshmen_file
        sr.teacher_file = teacher_file
        sr.transfer_file = transfer_file
        sr.withdraw_file = withdraw_file

        log(f"[OK] 신입생: {freshmen_file.name}" if freshmen_file else "[SKIP] 신입생 파일 없음 (키워드: 신입생/신입)")
        log(f"[OK] 교사: {teacher_file.name}" if teacher_file else "[SKIP] 교사 파일 없음 (키워드: 교사/교원)")
        log(f"[OK] 전입생: {transfer_file.name}" if transfer_file else "[SKIP] 전입생 파일 없음 (키워드: 전입생/전입)")
        log(f"[OK] 전출생: {withdraw_file.name}" if withdraw_file else "[SKIP] 전출생 파일 없음 (키워드: 전출생/전출)")

        template_register = choose_template_register(dirs["TEMPLATES"], year_str)
        sr.template_register = template_register
        log(f"[OK] 양식(등록): {template_register.name}")

        template_notice = choose_template_notice(dirs["TEMPLATES"], year_str)
        sr.template_notice = template_notice
        log(f"[OK] 양식(안내): {template_notice.name}")

        need_roster = bool(transfer_file) or bool(withdraw_file)
        sr.need_roster = need_roster

        if need_roster:
            roster_ws, roster_path, roster_year = _load_roster_sheet(dirs, school_name)
            sr.roster_path = roster_path
            sr.roster_year = roster_year
            log(f"[OK] 학생명부: {roster_path.name}")

            try:
                modified_date = datetime.fromtimestamp(roster_path.stat().st_mtime).date()

                # 1) 기본값: 파일 마지막 수정일
                auto_basis = modified_date

                # 2) 사용자가 UI에서 바꿔서 내려준 값이 있으면 그것을 우선 사용
                if roster_basis_date is not None and roster_basis_date != auto_basis:
                    sr.roster_basis_date = roster_basis_date
                    log(
                        f"[INFO] 학생명부 마지막 수정일은 {auto_basis.isoformat()} 이지만, "
                        f"사용자가 명부 기준일을 {roster_basis_date.isoformat()} 로 수정했습니다."
                    )
                else:
                    sr.roster_basis_date = auto_basis
                    log(
                        f"[INFO] 학생명부 마지막 수정일({auto_basis.isoformat()})을 "
                        "명부 기준일로 자동 감지했습니다."
                    )

                # 3) 작업일과 다른 경우 참고용 안내
                if sr.roster_basis_date != work_date:
                    log(
                        "[INFO] 현재 작업일과 명부 기준일이 다릅니다. "
                        f"(작업일={work_date.isoformat()}, 명부 기준일={sr.roster_basis_date.isoformat()})"
                    )

            except Exception as e:
                log(f"[WARN] 학생명부 파일 수정일 조회 중 오류: {e}")

                            # 4) 학생명부 내용 분석 (ID prefix 기반 연도/학년 시프트 판정)
            try:
                roster_info = _analyze_roster_once(roster_ws, input_year=year_int)
                sr.roster_info = roster_info

                rt = roster_info.get("roster_time", "unknown")
                shift = roster_info.get("ref_grade_shift", 0)
                log(
                    f"[INFO] 학생명부 분석 결과: roster_time={rt}, ref_grade_shift={shift} "
                    "(ID prefix 기반 판정)"
                )
            except Exception as e:
                log(f"[WARN] 학생명부 분석(analyze_roster_once) 중 오류: {e}")
                sr.roster_info = None

            # 2) ID prefix 기반 학년도 추정 (참고용 안내)
            try:
                expected_year = year_int
                roster_info = _analyze_roster_once(roster_ws, input_year=expected_year)
                id_roster_time = roster_info.get("roster_time")  # this_year / last_year / unknown

                if id_roster_time == "this_year":
                    log(f"[INFO] 학생명부 ID 패턴 기준으로 {expected_year}학년도 명부로 추정됩니다.")
                elif id_roster_time == "last_year":
                    log(f"[INFO] 학생명부 ID 패턴 기준으로 {expected_year-1}학년도 명부로 추정됩니다.")
                else:
                    log("[INFO] 학생명부 ID 패턴 기준 학년도 추정이 불확실합니다(unknown).")

                # 3) '명부 기준일' + 개학일 기준으로 실제 사용할 학년도 결정
                #    - UI에서 사용자가 입력한 값(roster_basis_date)이 있으면 그걸 최우선으로 사용
                basis_date = roster_basis_date or sr.roster_basis_date or work_date
                sr.roster_basis_date = basis_date  # 최종 기준일을 ScanResult에도 반영

                if basis_date < school_start_date:
                    roster_time = "last_year"
                    ref_shift = -1
                else:
                    roster_time = "this_year"
                    ref_shift = 0

                roster_info["roster_time"] = roster_time          # 우리가 실제로 쓸 학년도
                roster_info["ref_grade_shift"] = ref_shift        # g_roster = g_cur + ref_shift
                roster_info["id_roster_time"] = id_roster_time    # ID 패턴 기준 값은 참고용

                sr.roster_info = roster_info

                log(
                    "[INFO] 명부 기준일/개학일 기준으로 "
                    f"'{ '작년' if roster_time == 'last_year' else '올해' } 학년도 명부'로 간주합니다. "
                    f"(ref_grade_shift={ref_shift})"
                )

                # ID 추정값과 실제 사용 학년도가 다르면 경고만
                if id_roster_time in ("this_year", "last_year") and id_roster_time != roster_time:
                    log(
                        "[WARN] 학생명부 ID 패턴 기준 학년도 추정이 "
                        "명부 기준일/개학일 기준 예상 학년도와 다를 수 있습니다. "
                        "명부가 최신인지 한 번 더 확인해 주세요."
                    )
            except Exception as e:
                log(f"[WARN] 학생명부 학년도/ID 패턴 추정 중 오류가 발생했습니다: {e}")
        else:
            log("[SKIP] 전입/전출 파일이 없어 학생명부 로드를 스킵")

        needs_open_date = bool(withdraw_file)
        sr.needs_open_date = needs_open_date
        if needs_open_date:
            log("[INFO] 전출생 파일 감지 → 개학일(퇴원일자 계산용) 입력 필요")
        else:
            log("[INFO] 개학일 입력 불필요")

        # -------------------------
        # 3-3 실행 가능 여부 플래그 정리
        # -------------------------
        missing_fields: List[str] = []

        if sr.db_path is None:
            missing_fields.append("DB 파일")
        if sr.template_register is None:
            missing_fields.append("등록 템플릿")
        if sr.template_notice is None:
            missing_fields.append("안내 템플릿")

        # 🔻 신입생 파일은 선택 사항으로 변경
        #    (입력 xlsx가 아예 없을 때는 위에서 이미 ValueError로 막고 있음)
        # if sr.freshmen_file is None:
        #     missing_fields.append("신입생 명단(.xlsx)")

        if sr.need_roster and sr.roster_path is None:
            missing_fields.append("학생명부(.xlsx)")

        sr.missing_fields = missing_fields
        sr.needs_open_date = bool(sr.withdraw_file)

        # 최종 실행 가능 여부 재계산
        base_ok = True
        if sr.need_roster and sr.roster_path is None:
            base_ok = False

        sr.can_execute_after_input = base_ok
        sr.can_execute = base_ok and (len(sr.missing_fields) == 0)

        sr.ok = True
        log("[DONE] 스캔 완료")
        return sr
    
    except Exception as e:
        log(f"[ERROR] {e}")
        sr.ok = False
        
        sr.logs = logs
        return sr


def _extract_layout(layout_overrides: Dict[str, Any], kind: str, default_header: int):
    """
    layout_overrides[kind]가
      - dict: {"header_row": x, "data_start_row": y, ...}
      - int : y (data_start_row만)
      - None: 자동 감지
    이런 케이스를 모두 처리해서 (header_row, data_start_row) 튜플로 반환.
    """
    info = layout_overrides.get(kind)

    # dict 형태 (detect_input_layout 결과 그대로 들어온 경우)
    if isinstance(info, dict):
        header = info.get("header_row") or default_header
        data_start = info.get("data_start_row")
        return header, data_start

    # 숫자 하나만 들어온 경우 → header는 기본값 유지
    if isinstance(info, (int, float)):
        return default_header, int(info)

    # 아무 것도 없으면 자동 감지
    return default_header, None

def execute_pipeline(
    scan: ScanResult,
    work_date: date,
    school_start_date: Optional[date] = None,
    layout_overrides: Optional[Dict[str, int]] = None,
) -> PipelineResult:
    """
    scan 결과를 기반으로 등록파일 + 안내파일을 한 번에 생성.
    - 신입생만 있어도 동작
    - 전입/전출/교사 파일이 없으면 그 부분은 자동으로 스킵
    - 전출은 학생명부 + 개학일이 모두 있어야 처리
    """
    # 스캔 로그 이어서 사용 (이미 LogEntry 리스트)
    logs: List[LogEntry] = list(scan.logs)

    def log(
        msg: str,
        level: str = "INFO",
        code: str = "TEXT",
        context: Optional[Dict[str, Any]] = None,
    ) -> None:
        logs.append(
            LogEntry(
                level=level,
                code=code,
                message=msg,
                context=context or {},
            )
        )

    layout_overrides = layout_overrides or {}
    outputs: List[Path] = []

    try:
        if not scan.ok:
            raise PipelineError(
                code="SCAN_NOT_OK",
                message="scan.ok=False 상태입니다. 스캔 단계 오류를 먼저 확인해 주세요.",
                context={},
            )
        
        school_name = scan.school_name
        year_str = scan.year_str
        year_int = scan.year_int or int(year_str)

        log(f"[INFO] 실행 시작 | 학교={school_name}, 학년도={year_str}")
        log(f"[INFO] 작업 폴더: {scan.output_dir}")

        # -------------------------------------------------
        # 1) 인풋 파일 경로 정리
        # -------------------------------------------------
        freshmen_path = scan.freshmen_file
        teacher_path = scan.teacher_file
        transfer_path = scan.transfer_file
        withdraw_path = scan.withdraw_file

        # -------------------------------------------------
        # 2) 인풋 읽기 (레이아웃 override 반영)
        # -------------------------------------------------
        # 신입생 (있을 때만)
        freshmen_rows: List[Dict] = []
        if freshmen_path:
            fr_header, fr_start = _extract_layout(layout_overrides, "freshmen", default_header=2)
            log(
                "[DEBUG] 신입생 layout: "
                f"header_row={fr_header}, data_start_row={fr_start if fr_start is not None else 'auto'}"
            )
            freshmen_rows = _read_freshmen_rows(
                freshmen_path,
                header_row=fr_header,
                data_start_row=fr_start,
            )
            log(f"[OK] 신입생 {len(freshmen_rows)}명 로드")
        else:
            log("[INFO] 신입생 파일 없음 → 신입생 등록은 스킵합니다.")

        # 교사
        if teacher_path:
            t_header, t_start = _extract_layout(layout_overrides, "teacher", default_header=3)
            log(
                "[DEBUG] 교사 layout: "
                f"header_row={t_header}, data_start_row={t_start if t_start is not None else 'auto'}"
            )
            teacher_rows = _read_teacher_rows(
                teacher_path,
                header_row=t_header,
                data_start_row=t_start,
            )
            log(f"[OK] 교사 신청 {len(teacher_rows)}건 로드")
        else:
            teacher_rows = []
            log("[INFO] 교사 파일 없음 → 교사 관련 처리는 스킵")

        # 전입
        if transfer_path:
            tr_header, tr_start = _extract_layout(layout_overrides, "transfer", default_header=2)
            log(
                "[DEBUG] 전입생 layout: "
                f"header_row={tr_header}, data_start_row={tr_start if tr_start is not None else 'auto'}"
            )
            transfer_rows = _read_transfer_rows(
                transfer_path,
                header_row=tr_header,
                data_start_row=tr_start,
            )
            log(f"[OK] 전입생 {len(transfer_rows)}명 로드")
        else:
            transfer_rows = []
            log("[INFO] 전입생 파일 없음 → 전입 처리 스킵")

        # 전출
        if withdraw_path:
            wd_header, wd_start = _extract_layout(layout_overrides, "withdraw", default_header=2)
            log(
                "[DEBUG] 전출생 layout: "
                f"header_row={wd_header}, data_start_row={wd_start if wd_start is not None else 'auto'}"
            )
            withdraw_rows = _read_withdraw_rows(
                withdraw_path,
                header_row=wd_header,
                data_start_row=wd_start,
            )
            log(f"[OK] 전출생 {len(withdraw_rows)}명 로드")
        else:
            withdraw_rows = []
            log("[INFO] 전출생 파일 없음 → 전출 처리 스킵")

        # -------------------------------------------------
        # 3) 전입 ID 생성 (학생명부가 있는 경우에만)
        # -------------------------------------------------
        transfer_done_rows: List[Dict] = []
        transfer_hold_rows: List[Dict] = []
        prefix_by_grade: Dict[int, int] = {}

        if transfer_rows:
            if not (scan.roster_path and scan.roster_info):
                raise PipelineError(
                    code="TRANSFER_NEEDS_ROSTER",
                    message="전입생이 있는데 학생명부 정보가 없습니다. 스캔 결과를 확인하세요.",
                    context={
                        "roster_path": str(scan.roster_path) if scan.roster_path else None,
                    },
                )

            # (필요하면 roster_ws 넘기도록 확장 가능)
            # roster_wb = safe_load_workbook(scan.roster_path, data_only=True)
            # roster_ws = roster_wb.worksheets[0]

            transfer_done_rows, transfer_hold_rows, prefix_by_grade = _build_transfer_ids(
                transfer_rows=transfer_rows,
                roster_info=scan.roster_info,
                input_year=year_int,
            )
            log(f"[OK] 전입 ID 매칭 완료 | 완료 {len(transfer_done_rows)}명, 보류 {len(transfer_hold_rows)}명")
        else:
            log("[INFO] 전입생 없음 → 전입 ID 생성 스킵")

        # -------------------------------------------------
        # 4) 전출 퇴원 리스트 생성 (학생명부 + 개학일 + 작업일 필요)
        # -------------------------------------------------
        withdraw_done_rows: List[Dict] = []
        withdraw_hold_rows: List[Dict] = []
        transfer_out_auto_skip: int = 0

        if withdraw_rows:
            if not scan.roster_path:
                raise PipelineError(
                    code="WITHDRAW_NEEDS_ROSTER",
                    message="전출생이 있는데 학생명부 파일 경로가 없습니다. 스캔 결과를 확인하세요.",
                    context={},
                )
            if not scan.roster_info:
                raise ValueError("[오류] 전출생이 있는데 학생명부 정보(roster_info)가 없습니다.")
            if school_start_date is None:
                raise ValueError("[오류] 전출 처리에 필요한 개학일이 입력되지 않았습니다.")

            roster_wb2 = _safe_load_workbook(scan.roster_path, data_only=True)
            sheets2 = roster_wb2.worksheets
            if not sheets2:
                raise ValueError(f"[오류] 학생명부에 시트가 없습니다: {scan.roster_path.name}")
            roster_ws2 = sheets2[0]

            withdraw_done_rows, withdraw_hold_rows = _build_withdraw_outputs(
                roster_ws=roster_ws2,
                withdraw_rows=withdraw_rows,
                school_start_date=school_start_date,
                work_date=work_date,
                roster_info=scan.roster_info,
            )

            transfer_out_auto_skip = sum(
                1 for row in withdraw_hold_rows
                if str(row.get("보류사유", "")).startswith("자동 제외")
            )

            log(
                f"[OK] 전출 퇴원 리스트 생성 | "
                f"퇴원 {len(withdraw_done_rows)}명, 보류 {len(withdraw_hold_rows)}명 "
                f"(자동 제외 {transfer_out_auto_skip}명 포함)"
            )
        else:
            log("[INFO] 전출생 없음 → 퇴원 처리 스킵")

        # -------------------------------------------------
        # 5) 등록작업파일 생성
        # -------------------------------------------------
        if not scan.template_register:
            raise PipelineError(
                code="REGISTER_TEMPLATE_MISSING",
                message="등록 템플릿 경로가 없습니다. 스캔 결과를 확인하세요.",
                context={},
            )

        out_register_path = scan.output_dir / f"★{school_name}_등록작업파일(작업용).xlsx"


        _fill_register(
            template_path=scan.template_register,
            out_path=out_register_path,
            school_name=school_name,
            year=year_str,
            freshmen_rows=freshmen_rows,
            transfer_done_rows=transfer_done_rows,
            teacher_rows=teacher_rows,
            # 리스트 그대로 전달 (비어 있어도 OK)
            transfer_hold_rows=transfer_hold_rows,
            withdraw_done_rows=withdraw_done_rows,
            withdraw_hold_rows=withdraw_hold_rows,
        )
        log(f"[OK] 등록작업파일 생성 완료: {out_register_path.name}")
        

        # -------------------------------------------------
        # 6) 안내파일 생성 (ID/PW)
        # -------------------------------------------------

        if not scan.template_notice:
            raise PipelineError(
                code="NOTICE_TEMPLATE_MISSING",
                message="안내 템플릿 경로가 없습니다. 스캔 결과를 확인하세요.",
                context={},
            )

        # 실제로 작업된 대상만 제목에 포함
        notice_kinds: List[str] = []

        # 신입생: 파일이 있고, 읽어온 행이 1개 이상일 때만 포함
        if freshmen_path and len(freshmen_rows) > 0:
            notice_kinds.append("신입생")

        # 전입생: 실제 ID가 생성된 케이스가 있을 때만 포함
        if len(transfer_done_rows) > 0:
            notice_kinds.append("전입생")

        # 교직원: 신청 데이터가 1개 이상일 때만 포함
        if len(teacher_rows) > 0:
            notice_kinds.append("교직원")

        # 방어: 아무도 없을 일은 거의 없지만, 혹시 모를 경우 대비
        if notice_kinds:
            title_middle = ",".join(notice_kinds) + "_ID,PW안내"
        else:
            title_middle = "ID,PW안내"

        out_notice_path = scan.output_dir / f"☆{school_name}_{title_middle}.xlsx"

        _build_notice_file(
            template_notice_path=scan.template_notice,
            out_notice_path=out_notice_path,
            out_register_path=out_register_path,
            teacher_file_path=teacher_path,
            transfer_done_rows=transfer_done_rows,
        )
        log(f"[OK] 안내파일 생성 완료: {out_notice_path.name}")

        # -------------------------------------------------
        # 7) 결과 정리
        # -------------------------------------------------
        pr = PipelineResult(
            ok=True,
            outputs=[out_register_path, out_notice_path],
            logs=logs,
        )
        pr.transfer_in_done = len(transfer_done_rows)
        pr.transfer_in_hold = len(transfer_hold_rows)
        pr.transfer_out_done = len(withdraw_done_rows)
        pr.transfer_out_hold = len(withdraw_hold_rows)
        pr.transfer_out_auto_skip = transfer_out_auto_skip

        log("[DONE] 실행 완료")
        return pr

    except PipelineError as e:
            # 의도된 도메인 오류 (우리가 명시적으로 raise한 것)
            log(
                e.message,
                level="ERROR",
                code=e.code,
                context=e.context,
            )
            return PipelineResult(
                ok=False,
                outputs=[],
                logs=logs,
            )

    except Exception as e:
        # 예상 못 한 예외
        log(
            f"실행 중 예외 발생: {e}",
            level="ERROR",
            code="UNEXPECTED_ERROR",
        )
        return PipelineResult(
            ok=False,
            outputs=[],
            logs=logs,
        )

def run_pipeline(
    work_root: Path,
    school_name: str,
    school_start_date: date,
    work_date: date,
    layout_overrides: Optional[Dict[str, Dict[str, int]]] = None,
    roster_basis_date: Optional[date] = None,
) -> PipelineResult:
    """
    Streamlit에서 부르는 실제 실행 함수.

    - 1) scan_pipeline으로 인풋 상태 점검
    - 2) 문제 없으면 execute_pipeline으로 등록/안내 엑셀 파일 생성
    - 3) 안내 메일/문자 텍스트는 generate_notice_mail_text에서만 처리 (txt 파일 생성 없음)
    """
    logs: List[LogEntry] = []

    def log(
        msg: str,
        level: str = "INFO",
        code: str = "TEXT",
        context: Optional[Dict[str, Any]] = None,
    ) -> None:
        logs.append(
            LogEntry(
                level=level,
                code=code,
                message=msg,
                context=context or {},
            )
        )

    work_root = Path(work_root).resolve()
    school_name = (school_name or "").strip()

    if not school_name:
        log(
            "학교명이 비어 있습니다.",
            level="ERROR",
            code="RUN_SCHOOL_NAME_EMPTY",
        )
        return PipelineResult(ok=False, outputs=[], logs=logs)
    
    try:
        # 1) 사전 점검 (scan)
        scan = scan_pipeline(
            work_root=work_root,
            school_name=school_name,
            school_start_date=school_start_date,
            work_date=work_date,
            roster_basis_date=roster_basis_date,
        )
        logs.extend(scan.logs)

        if not scan.ok:
            log(
                "스캔 단계에서 오류가 발생하여 실행을 중단합니다.",
                level="ERROR",
                code="SCAN_NOT_OK",
            )
            return PipelineResult(ok=False, outputs=[], logs=logs)

        if not scan.can_execute:
            msg = ", ".join(scan.missing_fields) if scan.missing_fields else "필수 파일 누락"
            log(
                f"실행 불가 상태입니다. ({msg})",
                level="ERROR",
                code="SCAN_CANNOT_EXECUTE",
                context={"missing": scan.missing_fields},
            )
            return PipelineResult(ok=False, outputs=[], logs=logs)

        # 2) 실제 실행은 execute_pipeline에 위임
        result = execute_pipeline(
            scan=scan,
            work_date=work_date,
            school_start_date=school_start_date,
            layout_overrides=layout_overrides,
        )

        # scan 단계 로그 + 실행 로그 합치기
        full_logs = logs + result.logs
        return PipelineResult(
            ok=result.ok,
            outputs=result.outputs,
            logs=full_logs,
            transfer_in_done=result.transfer_in_done,
            transfer_in_hold=result.transfer_in_hold,
            transfer_out_done=result.transfer_out_done,
            transfer_out_hold=result.transfer_out_hold,
            transfer_out_auto_skip=result.transfer_out_auto_skip,
        )

    except PipelineError as e:
        log(
            e.message,
            level="ERROR",
            code=e.code,
            context=e.context,
        )
        return PipelineResult(
            ok=False,
            outputs=[],
            logs=logs,
        )

    except Exception as e:
        log(
            f"실행 중 예외 발생: {e}",
            level="ERROR",
            code="UNEXPECTED_ERROR",
        )
        return PipelineResult(
            ok=False,
            outputs=[],
            logs=logs,
        )

def run_pipeline_partial(
    work_root: Path,
    school_name: str,
    open_date: date,
    mode: str,
) -> PipelineResult:
    """
    UI의 '부분 실행' 버튼용.
    현재는 안정성을 위해 전체 파이프라인을 재생성하는 방식으로 동작.
    mode: 'freshmen'|'teacher'|'transfer'|'withdraw' (지금은 아직 구분하지 않음)
    """
    # 부분 실행 분기 로직은 나중에 진짜 나누고,
    # 지금은 전체 run_pipeline을 그대로 돌린다.
    return run_pipeline(
        work_root=work_root,
        school_name=school_name,
        school_start_date=open_date,
        work_date=open_date,
        roster_basis_date=None,
    )

 