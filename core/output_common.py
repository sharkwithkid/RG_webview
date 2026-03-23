# core/output_common.py

"""
출력 파일 생성 시 공통으로 사용하는 유틸 모듈.

책임 범위:
  - 기존 파일 백업
  - 셀 값 쓰기
  - 출력 영역 초기화
  - 워크북 열기 위치 / 시트 보기 상태 정리
  - 출력 파일 후처리 공통 기능

등록파일, 안내문, diff 결과 파일 등 출력물 생성 시 공통으로 사용한다.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any, Optional

from openpyxl.styles import PatternFill, Border
from openpyxl.worksheet.views import Selection


def backup_if_exists(out_path: Path) -> Optional[Path]:
    """기존 파일이 있으면 _backup 폴더로 이동."""
    out_path = Path(out_path)
    if not out_path.exists():
        return None
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_dir = out_path.parent / "_backup"
    backup_dir.mkdir(parents=True, exist_ok=True)
    dest = backup_dir / f"{out_path.stem}_{ts}{out_path.suffix}"
    out_path.replace(dest)
    return dest


def write_text_cell(ws, row: int, col: int, value: Any):
    cell = ws.cell(row=row, column=col)
    cell.value = "" if value is None else str(value)
    cell.data_type = "s"
    cell.number_format = "@"
    return cell


def clear_format_workbook_from_row(wb, start_row: int = 2):
    for ws in wb.worksheets:
        last_data_row = 0
        max_row = ws.max_row
        max_col = ws.max_column or 1

        for r in range(start_row, max_row + 1):
            for c in range(1, max_col + 1):
                if ws.cell(row=r, column=c).value not in (None, ""):
                    last_data_row = r
                    break

        if not last_data_row:
            continue

        for r in range(last_data_row + 1, max_row + 1):
            for c in range(1, max_col + 1):
                cell = ws.cell(r, c)
                cell.fill = PatternFill(fill_type=None)
                cell.border = Border()


def reset_view_to_a1(wb):
    for ws in wb.worksheets:
        sv = ws.sheet_view
        sv.topLeftCell = "A1"
        sv.activeCell = "A2"
        sv.selection = [Selection(activeCell="A2", sqref="A2")]
        ws.freeze_panes = "A2"
        if hasattr(sv, "tabSelected"):
            sv.tabSelected = False

    first_ws = wb.worksheets[0]
    if hasattr(first_ws.sheet_view, "tabSelected"):
        first_ws.sheet_view.tabSelected = True
    wb.active = 0

    if getattr(wb, "views", None) and wb.views:
        wb.views[0].activeTab = 0
        wb.views[0].firstSheet = 0