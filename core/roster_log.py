# roster_log.py
"""
학교 전체 명단 파일(xlsx) 읽기/쓰기 모듈.

열 위치는 앱에서 사용자가 직접 지정한 col_map을 받아서 사용합니다.
하드코딩된 열 번호 없음.

col_map 구조 (1-based 열 번호):
  {
    "sheet":       "학교명단",
    "header_row":  7,
    "data_start":  8,
    "col_school":  5,
    "col_email_arr": 10,
    "col_email_snt": 11,
    "col_worker":  12,
    "col_freshmen": 13,
    "col_transfer": 14,
    "col_withdraw": 15,
    "col_teacher":  16,
  }

공개 API:
  find_school_row(ws, school_name, col_school, data_start) -> int | None
  write_work_result(xlsx_path, school_name, worker, kind_flags,
                    email_arrived_date, col_map) -> (bool, str)
  write_email_sent(xlsx_path, school_name, sent_date, col_map) -> (bool, str)
"""
from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path
from typing import Optional, Dict, Any

from openpyxl import load_workbook

DONE_TEXT = "완료"

# 하위 호환용 기본값 (col_map 없이 호출될 때 fallback)
_DEFAULT_COL_MAP = {
    "sheet":        "학교명단",
    "header_row":   7,
    "data_start":   8,
    "col_school":   5,
    "col_domain":   None,  # 도메인(홈페이지) 열 — 미지정 시 None
    "col_email_arr": 10,
    "col_email_snt": 11,
    "col_worker":   12,
    "col_freshmen": 13,
    "col_transfer": 14,
    "col_withdraw": 15,
    "col_teacher":  16,
    "col_seq":      None,  # 자료실 순번 열 — 미지정 시 None
}


def _resolve_col_map(col_map: Optional[Dict]) -> Dict:
    """col_map이 없거나 불완전하면 기본값으로 채움."""
    base = dict(_DEFAULT_COL_MAP)
    if col_map:
        for k, v in col_map.items():
            if v:  # 0이나 빈값이 아닌 경우만 덮어씀
                base[k] = v
    return base


def _normalize(s) -> str:
    if s is None:
        return ""
    return re.sub(r"\s+", "", str(s)).strip()


def find_school_row(
    ws,
    school_name: str,
    col_school: int,
    data_start: int,
) -> Optional[int]:
    """
    학교명이 포함된 행 번호 반환 (없으면 None).
    exact match 우선, 없으면 contains 검색.
    """
    key = _normalize(school_name)
    if not key:
        return None

    exact_row = None
    contain_row = None

    for r in range(data_start, ws.max_row + 1):
        cell_val = _normalize(ws.cell(r, col_school).value)
        if not cell_val:
            continue
        if cell_val == key:
            exact_row = r
            break
        if contain_row is None and key in cell_val:
            contain_row = r

    return exact_row or contain_row


def _open_with_retry(xlsx_path: Path):
    try:
        wb = load_workbook(str(xlsx_path))
        return wb, None
    except PermissionError:
        return None, "파일이 다른 프로그램에서 열려 있습니다. 닫은 후 다시 시도해 주세요."
    except Exception as e:
        return None, str(e)


def write_work_result(
    xlsx_path: Path,
    school_name: str,
    worker: str,
    kind_flags: Dict[str, bool],
    email_arrived_date: Optional[date] = None,
    col_map: Optional[Dict[str, Any]] = None,
    seq_no: Optional[int] = None,
) -> tuple[bool, str]:
    """실행 완료 후 명단 파일에 작업 결과 기록."""
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        return False, f"명단 파일을 찾을 수 없습니다: {xlsx_path.name}"
    if xlsx_path.suffix.lower() != ".xlsx":
        return False, "명단 파일은 xlsx 형식이어야 합니다. xlsb는 xlsx로 변환 후 사용하세요."

    cm = _resolve_col_map(col_map)
    sheet_name  = cm["sheet"]
    data_start  = int(cm["data_start"])
    col_school  = int(cm["col_school"])
    col_worker  = int(cm["col_worker"])
    col_email_arr = int(cm["col_email_arr"])

    kind_col_map = {
        "신입생": int(cm["col_freshmen"]),
        "전입생": int(cm["col_transfer"]),
        "전출생": int(cm["col_withdraw"]),
        "교직원": int(cm["col_teacher"]),
    }

    wb, err = _open_with_retry(xlsx_path)
    if err:
        return False, err

    try:
        if sheet_name not in wb.sheetnames:
            return False, f"'{sheet_name}' 시트를 찾을 수 없습니다."

        ws = wb[sheet_name]
        row = find_school_row(ws, school_name, col_school, data_start)
        if row is None:
            return False, f"명단에서 '{school_name}' 학교를 찾을 수 없습니다."

        col_seq_raw = cm.get("col_seq")
        col_seq = int(col_seq_raw) if col_seq_raw else None

        if worker and col_worker:
            ws.cell(row, col_worker).value = worker

        if seq_no is not None and col_seq:
            ws.cell(row, col_seq).value = seq_no

        for kind, flag in kind_flags.items():
            col = kind_col_map.get(kind)
            if not col:
                continue
            if flag:
                existing = ws.cell(row, col).value
                if existing != DONE_TEXT:
                    ws.cell(row, col).value = DONE_TEXT
            else:
                # 이번 작업에 포함되지 않은 종류는 기존 값 초기화
                ws.cell(row, col).value = None

        if email_arrived_date is not None and col_email_arr:
            cell = ws.cell(row, col_email_arr)
            cell.value = f"{email_arrived_date.month}/{email_arrived_date.day}"

        wb.save(str(xlsx_path))
        return True, f"'{school_name}' 명단 기록 완료"

    except PermissionError:
        return False, "파일 저장 중 권한 오류 — 파일이 열려 있는지 확인하세요."
    except Exception as e:
        return False, f"명단 기록 중 오류: {e}"
    finally:
        try:
            wb.close()
        except Exception:
            pass


def write_email_sent(
    xlsx_path: Path,
    school_name: str,
    sent_date: Optional[date],
    col_map: Optional[Dict[str, Any]] = None,
) -> tuple[bool, str]:
    """완료 이메일 발송일 기록."""
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        return False, f"명단 파일을 찾을 수 없습니다: {xlsx_path.name}"

    cm = _resolve_col_map(col_map)
    sheet_name  = cm["sheet"]
    data_start  = int(cm["data_start"])
    col_school  = int(cm["col_school"])
    col_email_snt = int(cm["col_email_snt"])

    wb, err = _open_with_retry(xlsx_path)
    if err:
        return False, err

    try:
        if sheet_name not in wb.sheetnames:
            return False, f"'{sheet_name}' 시트를 찾을 수 없습니다."

        ws = wb[sheet_name]
        row = find_school_row(ws, school_name, col_school, data_start)
        if row is None:
            return False, f"명단에서 '{school_name}' 학교를 찾을 수 없습니다."

        if col_email_snt:
            if sent_date is not None:
                cell = ws.cell(row, col_email_snt)
                cell.value = f"{sent_date.month}/{sent_date.day}"
            else:
                ws.cell(row, col_email_snt).value = None

        wb.save(str(xlsx_path))
        label = sent_date.strftime("%Y-%m-%d") if sent_date else "비움(보류)"
        return True, f"완료 이메일 발송일 기록: {label}"

    except PermissionError:
        return False, "파일 저장 중 권한 오류 — 파일이 열려 있는지 확인하세요."
    except Exception as e:
        return False, f"명단 기록 중 오류: {e}"
    finally:
        try:
            wb.close()
        except Exception:
            pass

