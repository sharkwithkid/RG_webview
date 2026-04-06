# core/run_main.py

"""
메인 반이동 파이프라인의 실행 전용 모듈.

책임 범위:
  - scan_main의 스캔 결과를 입력으로 받음
  - 등록 작업 파일 생성
  - 안내문 / 출력 파일 생성
  - 신입생 / 전입생 / 전출생 / 교직원 데이터 반영
  - 실행 결과 집계 및 반환

이 모듈은 실제 산출물 생성과 실행 로직을 담당.

공개 API:
  PipelineResult
  execute_pipeline(scan, work_date, school_start_date, layout_overrides, school_kind_override) -> PipelineResult
  run_pipeline(work_root, school_name, school_start_date, work_date, ..., school_kind_override) -> PipelineResult
  run_pipeline_partial(work_root, school_name, open_date, mode) -> PipelineResult
  scan_work_root(work_root) -> Dict[str, Any]
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from core.events import (
    transfer_in_hold as _evt_transfer_in_hold,
    transfer_out_hold as _evt_transfer_out_hold,
    freshmen_transfer_dup as _evt_freshmen_transfer_dup,
    roster_duplicate_transfer as _evt_roster_duplicate_transfer,
    roster_not_found_at_run as _evt_roster_not_found_at_run,
    template_register_not_found as _evt_template_register_not_found,
    template_notice_not_found as _evt_template_notice_not_found,
    db_file_error as _evt_db_file_error,
    open_date_required as _evt_open_date_required,
)
from typing import Dict, List, Optional, Tuple, Any
from collections import Counter

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Font, Alignment

from core.scan_main import (
    ScanResult,
    scan_pipeline,
    FRESHMEN_HEADER_SLOTS,
    TRANSFER_HEADER_SLOTS,
    WITHDRAW_HEADER_SLOTS,
    TEACHER_HEADER_SLOTS,
    detect_header_row_freshmen,
    detect_header_row_transfer,
    detect_header_row_withdraw,
    detect_header_row_teacher,
    detect_example_and_data_start,
    validate_input_sheet_structure,
)

from core.common import (
    get_project_dirs,
    safe_load_workbook,
    ensure_xlsx_only,
    get_first_sheet_with_warning,
    header_map,
    normalize_header_cell,
    _build_header_slot_map,
    normalize_name,
    normalize_name_key,
    dedup_suffix_letters,
    apply_suffix_for_duplicates,
    notice_name_key,
    parse_class_str,
    school_kind_from_name,
    school_profile_from_name,
    apply_school_kind_override,
    resolve_school_kind_by_grade,
    resolve_transfer_name_conflicts,
    RosterInfo,
)

from core.output_common import (
    backup_if_exists,
    write_text_cell,
    clear_format_workbook_from_row,
    reset_view_to_a1,
)


# =========================
# Result types
# =========================
@dataclass
class PipelineResult:
    ok: bool
    outputs: List[Path]
    logs: List[str]

    transfer_in_done: int = 0
    transfer_in_hold: int = 0
    transfer_out_done: int = 0
    transfer_out_hold: int = 0
    transfer_out_auto_skip: int = 0
    notice_dup_rows: List[int] = field(default_factory=list)
    notice_teacher_dup_rows: List[int] = field(default_factory=list)

    audit_summary: Dict[str, Any] = field(default_factory=dict)

    # 구조화된 판정 결과 — bridge/UI는 이것만 참조
    events:    List[Any] = field(default_factory=list)  # List[CoreEvent]
    row_marks: List[Any] = field(default_factory=list)  # List[RowMark]

# =========================
# L1. Input readers
# =========================
def normalize_withdraw_class(raw, grade: int) -> str:
    if raw is None:
        return ""

    if isinstance(raw, (int, float)):
        class_no = int(raw)
        return f"{grade}-{class_no}반"

    s = str(raw).strip()
    if not s:
        return ""

    s = s.replace("\u3000", " ").replace("\u00A0", " ")
    s = re.sub(r"\s+", "", s)

    if re.fullmatch(r"\d+\.0+", s):
        class_no = int(float(s))
        return f"{grade}-{class_no}반"

    nums = re.findall(r"\d+", s)
    if not nums:
        return s

    class_no = int(nums[-1])
    return f"{grade}-{class_no}반"


def _parse_freshmen_grade_meta(raw_grade: Any, input_year: int, school_name: str = "") -> Dict[str, Any]:
    s = "" if raw_grade is None else str(raw_grade).strip()
    s_norm = re.sub(r"\s+", "", s)

    if s_norm in {"유치원", "유치원반", "7세", "6세", "5세", "7세반", "6세반", "5세반"}:
        grade_label = s_norm.replace("반", "")
        if not re.search(r"\d", grade_label):

            # "유치원"처럼 나이 숫자 없는 경우 → 학교 종류 기반 1학년으로
            from core.common import school_kind_from_name, school_profile_from_name, apply_school_kind_override, resolve_school_kind_by_grade
            _, kind_prefix = school_kind_from_name(school_name)
            prefix = kind_prefix if kind_prefix else "초"
            grade_label = f"{prefix}1"

        return {
            "grade": 0,
            "grade_label": grade_label,
            "id_year": None,
            "register_grade": None,
            "register_class_name": None,
            "group_name": "기타그룹",
            "group_class_name": None,
            "is_kindergarten": True,
        }

    m = re.search(r"\d+", s_norm)
    if m:
        grade_i = int(m.group(0))
        return {
            "grade": grade_i,
            "grade_label": grade_i,
            "id_year": input_year,
            "register_grade": grade_i,
            "register_class_name": None,
            "group_name": None,
            "group_class_name": None,
            "is_kindergarten": False,
        }

    raise ValueError(f"[ERROR] 신입생 파일에서 학년 값을 인식할 수 없습니다: {s!r}")


def build_freshmen_prefix_map(
    freshmen_rows: List[Dict[str, Any]],
    input_year: int,
    roster_info: Optional[Dict[str, Any]] = None,
    manual_grade_year_map: Optional[Dict[int, int]] = None,
) -> Dict[int, int]:
    grades = sorted({
        int(r["grade"])
        for r in freshmen_rows
        if not r.get("is_kindergarten") and int(r["grade"]) > 0
    })

    if not grades:
        return {}

    manual_grade_year_map = {
        int(k): int(v)
        for k, v in (manual_grade_year_map or {}).items()
        if str(k).strip() != "" and str(v).strip() != ""
    }

    # 1) 기본값: 입학년도 기준 역산
    result: Dict[int, int] = {
        g: input_year - (g - 1)
        for g in grades
    }

    # 2) 명부 정보 없으면 기본값 + 수동 입력값만 반영
    if not roster_info:
        for g, y in manual_grade_year_map.items():
            if g in result:
                result[g] = y
        return result

    prefix_mode = (
        getattr(roster_info, "prefix_mode_by_roster_grade", None)
        if not isinstance(roster_info, dict)
        else roster_info.get("prefix_mode_by_roster_grade", {})
    ) or {}
    shift = int((
        getattr(roster_info, "ref_grade_shift", 0)
        if not isinstance(roster_info, dict)
        else roster_info.get("ref_grade_shift", 0)
    ) or 0)

    # 3) 명부 학년 -> 현재 학년 변환
    direct_current_grade_map: Dict[int, int] = {}
    for g_roster, pref in prefix_mode.items():
        try:
            g_roster_i = int(g_roster)
            pref_i = int(pref)
        except Exception:
            continue

        g_cur = g_roster_i - shift
        if g_cur <= 0:
            continue

        direct_current_grade_map[g_cur] = pref_i

    # 4) 직접값 우선 반영
    for g in grades:
        if g in direct_current_grade_map:
            result[g] = direct_current_grade_map[g]

    # 5) 없는 학년은 다른 학년 기준 역산/순산
    known_current_grades = sorted(direct_current_grade_map.keys())
    if known_current_grades:
        for g in grades:
            if g in direct_current_grade_map:
                continue

            anchor_g = min(known_current_grades, key=lambda x: abs(x - g))
            anchor_pref = direct_current_grade_map[anchor_g]
            result[g] = anchor_pref + (anchor_g - g)

    # 6) 수동 입력값이 있으면 최우선 적용
    for g, y in manual_grade_year_map.items():
        if g in result:
            result[g] = y

    return result


def read_freshmen_rows(
    xlsx_path: Path,
    input_year: int,
    header_row: Optional[int] = None,
    data_start_row: Optional[int] = None,
    roster_info: Optional[Dict[str, Any]] = None,
    school_name: str = "",
    manual_grade_year_map: Optional[Dict[int, int]] = None,
) -> List[Dict]:

    ensure_xlsx_only(xlsx_path)
    wb = safe_load_workbook(xlsx_path, data_only=True)
    try:
        ws = get_first_sheet_with_warning(wb, xlsx_path.name)

        if header_row is None:
            header_row = detect_header_row_freshmen(ws)

        slot_cols = _build_header_slot_map(ws, header_row, FRESHMEN_HEADER_SLOTS)
        col_grade = slot_cols.get("grade")
        col_class = slot_cols.get("class")
        col_num   = slot_cols.get("num") or slot_cols.get("no")
        col_name  = slot_cols.get("name")

        missing = []
        if col_grade is None:
            missing.append("학년")
        if col_class is None:
            missing.append("반")
        if col_name is None:
            missing.append("성명/이름")

        if missing:
            raise ValueError(
                "[ERROR] 신입생 파일 헤더에서 " + ", ".join(missing)
                + " 열을 찾을 수 없습니다. '학년', '반', '이름' 헤더를 확인해 주세요."
            )

        if data_start_row is None:
            _, data_start_row = detect_example_and_data_start(
                ws, header_row=header_row, name_col=col_name
            )

        issues, _, _, _ = validate_input_sheet_structure(
            ws=ws,
            kind="신입생",
            header_row=header_row,
            data_start_row=data_start_row,
            required_cols={"grade": col_grade, "class": col_class, "name": col_name},
            allow_blank_class_for_kindergarten=True,
        )
        if issues:
            raise ValueError("\n".join(issues[:20]))

        out: List[Dict[str, Any]] = []
        row = data_start_row
        
        while True:
            grade = ws.cell(row=row, column=col_grade).value
            cls   = ws.cell(row=row, column=col_class).value
            num   = ws.cell(row=row, column=col_num).value if col_num is not None else None
            name  = ws.cell(row=row, column=col_name).value

            grade_s = "" if grade is None else str(grade).strip()
            cls_s   = "" if cls is None else str(cls).strip()
            name_s  = "" if name is None else str(name).strip()

            # 필수열(학년/반/이름) 기준 빈 행이면 종료
            if not grade_s and not cls_s and not name_s:
                break

            grade_norm = re.sub(r"\s+", "", grade_s)
            cls_norm   = re.sub(r"\s+", "", cls_s)

            is_kindergarten = (
                grade_norm in {"유치원", "유치원반", "5세", "6세", "7세", "5세반", "6세반", "7세반"}
                or (not grade_s and cls_norm in {"유치원", "유치원반"})
                or cls_norm in {"유치원", "유치원반"}
            )

            check_vals = [name] if is_kindergarten else [grade, cls, name]
            if any(v is None or str(v).strip() == "" for v in check_vals):
                raise ValueError(f"[ERROR] 신입생 파일 {row}행에 학년/반/이름 중 빈 값이 있습니다.")

            grade_meta = _parse_freshmen_grade_meta(
                grade, input_year=input_year, school_name=school_name
            )

            name_n = normalize_name(name)
            if not name_n:
                raise ValueError(f"[ERROR] 신입생 파일 {row}행 이름을 인식할 수 없습니다.")

            out.append({
                "grade": grade_meta["grade"],
                "grade_label": grade_meta["grade_label"],
                "class": None if (cls is None or str(cls).strip() == "") else str(cls).strip(),
                "number": "" if (num is None or str(num).strip() == "") else str(num).strip(),
                "name": name_n,
                "id_year": grade_meta["id_year"],
                "register_grade": grade_meta["register_grade"],
                "register_class_name": grade_meta["register_class_name"],
                "group_name": grade_meta["group_name"],
                "group_class_name": grade_meta["group_class_name"],
                "is_kindergarten": grade_meta["is_kindergarten"],
            })
            row += 1

        prefix_map = build_freshmen_prefix_map(
            out,
            input_year=input_year,
            roster_info=roster_info,
            manual_grade_year_map=manual_grade_year_map,
        )

        for r in out:
            if not r.get("is_kindergarten") and int(r["grade"]) in prefix_map:
                r["id_year"] = prefix_map[int(r["grade"])]

        def _safe_int(x):
            try:
                return (0, int(x))
            except Exception:
                return (1, str(x))

        def _sort_grade(r):
            if r.get("is_kindergarten"):
                m = re.search(r"\d+", str(r.get("grade_label", "")))
                return (0, int(m.group(0)) if m else 999)
            return (1, int(r["grade"]))

        def _sort_key(r):
            sg = _sort_grade(r)
            if r.get("is_kindergarten"):
                cls = r.get("class")
                if cls and str(cls).strip():
                    return (sg, (0, str(cls).strip()), r["name"])
                return (sg, (1, ""), r["name"])
            return (sg, _safe_int(r["class"]), r["name"])

        out.sort(key=_sort_key)
        return out

    finally:
        wb.close()


def read_transfer_rows(
    xlsx_path: Path,
    header_row: Optional[int] = None,
    data_start_row: Optional[int] = None,
) -> List[Dict]:
    ensure_xlsx_only(xlsx_path)
    wb = safe_load_workbook(xlsx_path, data_only=True)
    try:
        ws = get_first_sheet_with_warning(wb, xlsx_path.name)

        if header_row is None:
            header_row = detect_header_row_transfer(ws)

        slot_cols = _build_header_slot_map(ws, header_row, TRANSFER_HEADER_SLOTS)
        col_grade = slot_cols.get("grade")
        col_class = slot_cols.get("class")
        col_num   = slot_cols.get("number")
        col_name  = slot_cols.get("name")

        missing = []
        if col_grade is None:
            missing.append("학년")
        if col_class is None:
            missing.append("반")
        if col_name is None:
            missing.append("이름")
        if missing:
            raise ValueError("[ERROR] 전입생 파일 헤더에서 " + ", ".join(missing) + " 열을 찾을 수 없습니다.")

        if data_start_row is None:
            _, data_start_row = detect_example_and_data_start(ws, header_row=header_row, name_col=col_name)

        issues, _, _, _ = validate_input_sheet_structure(
            ws=ws,
            kind="전입생",
            header_row=header_row,
            data_start_row=data_start_row,
            required_cols={"grade": col_grade, "class": col_class, "name": col_name},
            allow_blank_class_for_kindergarten=True,
        )

        errors = [i for i in issues if "[ERROR]" in i]
        warns  = [i for i in issues if "[WARN]" in i]

        if errors or warns:
            raise ValueError("\n".join(errors + warns))
        out: List[Dict[str, Any]] = []
        row = data_start_row
        while True:
            grade = ws.cell(row=row, column=col_grade).value
            cls   = ws.cell(row=row, column=col_class).value
            num   = ws.cell(row=row, column=col_num).value if col_num is not None else None
            name  = ws.cell(row=row, column=col_name).value

            if all(v is None or str(v).strip() == "" for v in [grade, cls, num, name]):
                break

            grade_s = "" if grade is None else str(grade).strip()
            cls_s_raw = "" if cls is None else str(cls).strip()
            grade_norm = re.sub(r"\s+", "", grade_s)
            cls_norm = re.sub(r"\s+", "", cls_s_raw)
            KINDER = {"유치원", "유치원반"}
            is_kindergarten = (
                grade_norm in {"유치원", "유치원반", "5세", "6세", "7세", "5세반", "6세반", "7세반"}
                or (not grade_s and cls_norm in KINDER)
                or cls_norm in KINDER
            )
            if is_kindergarten:
                name_n = normalize_name(name) if name else ""
                if not name_n:
                    raise ValueError(f"[ERROR] 전입생 파일 {row}행 이름을 인식할 수 없습니다.")
                out.append({
                    "grade": 0,
                    "class": "유치원반",
                    "number": "" if (num is None or str(num).strip() == "") else str(num).strip(),
                    "name": name_n,
                    "is_kindergarten": True,
                })
                row += 1
                continue

            if any(v is None or str(v).strip() == "" for v in [grade, cls, name]):
                raise ValueError(f"[ERROR] 전입생 파일 {row}행에 학년/반/이름 중 빈 값이 있습니다.")

            grade_s = str(grade).strip()
            m = re.search(r"\d+", grade_s)
            if not m:
                raise ValueError(f"[ERROR] 전입생 파일 {row}행에서 학년 값을 인식할 수 없습니다: {grade_s!r}")

            out.append({
                "grade": int(m.group(0)),
                "class": str(cls).strip(),
                "number": "" if (num is None or str(num).strip() == "") else str(num).strip(),
                "name": normalize_name(name),
            })
            row += 1

        return out
    finally:
        wb.close()


def read_teacher_rows(
    xlsx_path: Path,
    header_row: Optional[int] = None,
    data_start_row: Optional[int] = None,
) -> List[Dict]:
    ensure_xlsx_only(xlsx_path)
    wb = safe_load_workbook(xlsx_path, data_only=True)
    try:
        ws = get_first_sheet_with_warning(wb, xlsx_path.name)

        if header_row is None:
            header_row = detect_header_row_teacher(ws)

        slot_cols = _build_header_slot_map(ws, header_row, TEACHER_HEADER_SLOTS)
        col_pos   = slot_cols.get("position")
        col_name  = slot_cols.get("name")
        col_learn = slot_cols.get("learn")
        col_admin = slot_cols.get("admin")

        if col_name is None:
            raise ValueError("[ERROR] 교사 파일 헤더에서 이름 열을 찾을 수 없습니다.")

        if data_start_row is None:
            _, data_start_row = detect_example_and_data_start(ws, header_row=header_row, name_col=col_name)

        issues, _, _, _ = validate_input_sheet_structure(
            ws=ws,
            kind="교사",
            header_row=header_row,
            data_start_row=data_start_row,
            required_cols={"name": col_name},
        )
        if issues:
            raise ValueError("\n".join(issues[:20]))

        out: List[Dict[str, Any]] = []
        row = data_start_row
        while True:
            def _get(col_idx):
                return None if col_idx is None else ws.cell(row=row, column=col_idx).value

            pos = _get(col_pos)
            name = _get(col_name)
            v_learn = _get(col_learn)
            v_admin = _get(col_admin)

            if all(v is None or str(v).strip() == "" for v in [pos, name, v_learn, v_admin]):
                break
            if name is None or str(name).strip() == "":
                row += 1
                continue

            name_n = normalize_name(name)
            if not name_n:
                row += 1
                continue

            out.append({
                "position": "" if pos is None else str(pos).strip(),
                "name": name_n,
                "learn_apply": col_learn is not None and not (v_learn is None or str(v_learn).strip() == ""),
                "admin_apply": col_admin is not None and not (v_admin is None or str(v_admin).strip() == ""),
            })
            row += 1

        return out
    finally:
        wb.close()


def read_withdraw_rows(
    xlsx_path: Path,
    header_row: Optional[int] = None,
    data_start_row: Optional[int] = None,
) -> List[Dict]:
    ensure_xlsx_only(xlsx_path)
    wb = safe_load_workbook(xlsx_path, data_only=True)
    try:
        ws = get_first_sheet_with_warning(wb, xlsx_path.name)

        if header_row is None:
            header_row = detect_header_row_withdraw(ws)

        slot_cols = _build_header_slot_map(ws, header_row, WITHDRAW_HEADER_SLOTS)
        col_grade = slot_cols.get("grade")
        col_class = slot_cols.get("class")
        col_name  = slot_cols.get("name")

        missing = []
        if col_grade is None:
            missing.append("학년")
        if col_class is None:
            missing.append("반")
        if col_name is None:
            missing.append("성명/이름")
        if missing:
            raise ValueError("[ERROR] 전출생 파일 헤더에서 " + ", ".join(missing) + " 열을 찾을 수 없습니다.")

        if data_start_row is None:
            _, data_start_row = detect_example_and_data_start(ws, header_row=header_row, name_col=col_name)

        issues, _, _, _ = validate_input_sheet_structure(
            ws=ws,
            kind="전출생",
            header_row=header_row,
            data_start_row=data_start_row,
            required_cols={"grade": col_grade, "class": col_class, "name": col_name},
            allow_blank_class_for_kindergarten=True,
        )
        if issues:
            raise ValueError("\n".join(issues[:20]))

        out: List[Dict[str, Any]] = []
        row = data_start_row
        while True:
            grade = ws.cell(row=row, column=col_grade).value
            cls   = ws.cell(row=row, column=col_class).value
            name  = ws.cell(row=row, column=col_name).value

            if all(v is None or str(v).strip() == "" for v in [grade, cls, name]):
                break

            grade_s = "" if grade is None else str(grade).strip()
            cls_s_raw = "" if cls is None else str(cls).strip()
            grade_norm = re.sub(r"\s+", "", grade_s)
            cls_norm = re.sub(r"\s+", "", cls_s_raw)

            KINDER = {"유치원", "유치원반"}
            is_kindergarten = (
                grade_norm in {"유치원", "유치원반", "5세", "6세", "7세", "5세반", "6세반", "7세반"}
                or (not grade_s and cls_norm in KINDER)
                or cls_norm in KINDER
            )
            if is_kindergarten:
                name_n = normalize_name(name) if name else ""
                if not name_n:
                    raise ValueError(f"[ERROR] 전출생 파일 {row}행 이름을 인식할 수 없습니다.")
                out.append({"grade": 0, "class": "유치원반", "name": name_n, "is_kindergarten": True})
                row += 1
                continue

            if any(v is None or str(v).strip() == "" for v in [grade, cls, name]):
                raise ValueError(f"[ERROR] 전출생 파일 {row}행에 학년/반/이름 중 빈 값이 있습니다.")

            m = re.search(r"\d+", grade_s)
            if not m:
                raise ValueError(f"[ERROR] 전출생 파일 {row}행에서 학년 값을 인식할 수 없습니다: {grade_s!r}")
            grade_i = int(m.group(0))

            cls_s = normalize_withdraw_class(cls, grade_i)
            if not cls_s:
                raise ValueError(f"[ERROR] 전출생 파일 {row}행 반 값을 인식할 수 없습니다.")

            name_n = normalize_name(name)
            if not name_n:
                raise ValueError(f"[ERROR] 전출생 파일 {row}행 이름을 인식할 수 없습니다.")

            out.append({"grade": grade_i, "class": cls_s, "name": name_n})
            row += 1

        return out
    finally:
        wb.close()


# =========================
# L2. ID generation
# =========================
def build_transfer_ids(
    transfer_rows: List[Dict],
    roster_info: Dict,
    input_year: int,
    freshmen_rows: Optional[List[Dict]] = None,
    manual_grade_year_map: Optional[Dict[int, int]] = None,
) -> Tuple[List[Dict], List[Dict], Dict[int, int]]:

    freshmen_rows = freshmen_rows or []
    manual_grade_year_map = {
        int(k): int(v)
        for k, v in (manual_grade_year_map or {}).items()
        if str(k).strip() != "" and str(v).strip() != ""
    }

    resolved = resolve_transfer_name_conflicts(
        transfer_rows=transfer_rows,
        roster_info=roster_info,
    )

    freshmen_prefix_map = build_freshmen_prefix_map(
        freshmen_rows,
        input_year=input_year,
        roster_info=roster_info,
        manual_grade_year_map=manual_grade_year_map,
    )

    shift = (
        roster_info.get("ref_grade_shift", 0)
        if isinstance(roster_info, dict)
        else getattr(roster_info, "ref_grade_shift", 0)
    ) or 0
    prefix_mode = (
        roster_info.get("prefix_mode_by_roster_grade", {})
        if isinstance(roster_info, dict)
        else getattr(roster_info, "prefix_mode_by_roster_grade", {})
    ) or {}

    done: List[Dict] = []
    hold: List[Dict] = []
    final_prefix_by_current_grade: Dict[int, int] = {}

    for tr, rr in zip(transfer_rows, resolved):
        g_cur = tr["grade"]
        name_out = rr["name_out"]
        dup_with_roster = rr["dup_with_roster"]
        needs_highlight = rr["needs_highlight"]

        if g_cur in manual_grade_year_map:
            pref = manual_grade_year_map[g_cur]
            final_prefix_by_current_grade[g_cur] = pref
            done.append({
                **tr,
                "name": name_out,
                "id": f"{pref}{name_out}",
                "dup_with_roster": dup_with_roster,
                "needs_highlight": needs_highlight,
            })
            continue

        if g_cur in freshmen_prefix_map:
            pref = freshmen_prefix_map[g_cur]
            final_prefix_by_current_grade[g_cur] = pref
            done.append({
                **tr,
                "name": name_out,
                "id": f"{pref}{name_out}",
                "dup_with_roster": dup_with_roster,
                "needs_highlight": needs_highlight,
            })
            continue

        g_roster = g_cur + shift
        pref = prefix_mode.get(g_roster)
        if pref is None:
            hold.append({
                **tr,
                "보류사유": f"명부 학년({g_roster})에서 ID prefix 최빈값 산출 불가"
            })
            continue

        final_prefix_by_current_grade[g_cur] = pref

        done.append({
            **tr,
            "name": name_out,
            "id": f"{pref}{name_out}",
            "dup_with_roster": dup_with_roster,
            "needs_highlight": needs_highlight,
        })

    def _safe_int(x):
        try:
            return (0, int(x))
        except Exception:
            return (1, str(x))

    done.sort(key=lambda r: (r["grade"], _safe_int(r["class"]), _safe_int(r["number"]), r["name"]))
    hold.sort(key=lambda r: (r["grade"], _safe_int(r["class"]), _safe_int(r["number"]), r["name"]))

    return done, hold, final_prefix_by_current_grade


def split_transfer_dup_against_freshmen(
    freshmen_rows: List[Dict],
    transfer_done_rows: List[Dict],
) -> Tuple[List[Dict], List[Dict]]:
    """
    신입생 명단과 전입 완료 명단을 교차 비교해
    학년/반/이름이 동일한 전입생을 보류(hold) 목록으로 분리한다.
    엑셀 스펙: 신입생-전입생 중복 → hold (자동 제외 아님)
    """
    if not freshmen_rows or not transfer_done_rows:
        return transfer_done_rows, []

    def _norm_class(v: Any) -> str:
        if v is None:
            return ""
        return re.sub(r"\s+", "", str(v).strip())

    freshmen_keys = {
        (
            int(r.get("grade", 0) or 0),
            _norm_class(r.get("class", "")),
            normalize_name_key(r.get("name", "")),
        )
        for r in freshmen_rows
        if normalize_name_key(r.get("name", ""))
    }

    kept: List[Dict] = []
    dup_hold: List[Dict] = []
    for row in transfer_done_rows:
        key = (
            int(row.get("grade", 0) or 0),
            _norm_class(row.get("class", "")),
            normalize_name_key(row.get("name", "")),
        )
        if key in freshmen_keys:
            dup_hold.append({**row, "보류사유": "신입생 명단과 학년/반/이름이 동일합니다."})
        else:
            kept.append(row)

    return kept, dup_hold


# =========================
# L3. Withdraw outputs
# =========================
def build_withdraw_outputs(
    roster_ws,
    withdraw_rows: List[Dict],
    school_start_date: date,
    work_date: date,
    roster_info: Optional[Dict] = None,
) -> Tuple[List[Dict], List[Dict]]:
    done: List[Dict] = []
    hold: List[Dict] = []

    eff = school_start_date if work_date < school_start_date else work_date

    hm = _require_headers(
        roster_ws,
        "학생명부",
        ["현재반", "이전반", "학생이름", "아이디"],
        header_row=1,
    )

    col_now  = hm[normalize_header_cell("현재반")]
    col_prev = hm[normalize_header_cell("이전반")]
    col_name = hm[normalize_header_cell("학생이름")]
    col_id   = hm[normalize_header_cell("아이디")]

    roster_map: Dict[str, List[Dict]] = {}
    roster_by_grade_name: Dict[str, List[Dict]] = {}
    seen_grade_name_ids: set = set()

    def _norm_class(c: str) -> str:
        """유치원/유치원반 등 유치원 계열 반 이름을 통일."""
        s = c.strip()
        if re.sub(r"\s+", "", s) in {"유치원반", "유치원"}:
            return "유치원"
        return s

    def _strip_sfx(nk):
        return re.sub(r"[A-Za-z]+$", "", nk) if nk else ""

    def _sfx_cands(grade, base_nk):
        if not grade or not base_nk:
            return []
        return [
            r for rows in roster_by_grade_name.items()
            if rows[0].startswith(f"{grade}|") and _strip_sfx(rows[0].split("|", 1)[1]) == base_nk
            for r in rows[1]
        ]

    def _idx_class(cv, now_cv, nk, idv, nd):
        if cv is None: return
        c = _norm_class(str(cv).strip())
        if not c: return
        cn = "" if now_cv is None else str(now_cv).strip()
        roster_map.setdefault(f"{c}|{nk}", []).append(
            {"class": c, "now_class": cn, "name_key": nk, "name_disp": nd,
             "id": "" if idv is None else str(idv).strip()}
        )

    def _idx_grade(cv, now_cv, nk, idv, nd):
        if cv is None: return
        c = str(cv).strip()
        if not c: return
        parsed = parse_class_str(c)
        if parsed is None: return
        g = parsed[0]
        id_str = "" if idv is None else str(idv).strip()
        dk = (g, nk, id_str)
        if dk in seen_grade_name_ids: return
        seen_grade_name_ids.add(dk)
        cn = "" if now_cv is None else str(now_cv).strip()
        roster_by_grade_name.setdefault(f"{g}|{nk}", []).append(
            {"class": c, "now_class": cn, "name_key": nk, "name_disp": nd, "id": id_str, "grade": g}
        )

    # iter_rows + 연속 빈 행 조기종료 (max_row 오염 대응)
    MAX_BLANK = 20
    blank_streak = 0
    max_col_r = max(col_name, col_id, col_now, col_prev)
    for row_tuple in roster_ws.iter_rows(min_row=2, max_col=max_col_r, values_only=True):
        nmv  = row_tuple[col_name - 1] if col_name - 1 < len(row_tuple) else None
        idv  = row_tuple[col_id   - 1] if col_id   - 1 < len(row_tuple) else None
        nowv = row_tuple[col_now  - 1] if col_now  - 1 < len(row_tuple) else None
        prevv= row_tuple[col_prev - 1] if col_prev - 1 < len(row_tuple) else None
        if nmv is None and nowv is None and prevv is None:
            blank_streak += 1
            if blank_streak >= MAX_BLANK:
                break
            continue
        blank_streak = 0
        if nmv is None: continue
        nd = normalize_name(nmv)
        nk = normalize_name_key(nmv)
        if not nk: continue
        _idx_class(nowv,  nowv, nk, idv, nd)
        _idx_class(prevv, nowv, nk, idv, nd)
        _idx_grade(nowv or prevv, nowv, nk, idv, nd)

    for w in withdraw_rows:
        g_cur = w["grade"]
        wnd = w["name"]
        wnk = normalize_name_key(wnd)
        if not wnk:
            hold.append({"학년": g_cur, "반": w["class"], "성명": wnd, "보류사유": "성명 정규화(키) 결과가 비어 있음"})
            continue

        key = f"{_norm_class(str(w['class']))}|{wnk}"
        matches = roster_map.get(key, [])

        if not matches:
            c0 = roster_by_grade_name.get(f"{g_cur}|{wnk}", [])
            c1 = roster_by_grade_name.get(f"{g_cur+1}|{wnk}", [])
            cand = c0 + c1
            if len(cand) == 1:
                matches = cand
            else:
                base_nk = _strip_sfx(wnk)
                sc0 = _sfx_cands(g_cur,   base_nk)
                sc1 = _sfx_cands(g_cur+1, base_nk)
                if   len(sc0) == 1: matches = sc0
                elif len(sc0) >= 2:
                    hold.append({"학년": g_cur, "반": w["class"], "성명": wnd,
                                 "보류사유": "보류: 학생명부에서 동명이인(A,B,C 등)으로 구분된 이름 – 수동 확인이 필요합니다."})
                    continue
                elif len(sc1) == 1: matches = sc1
                elif len(sc1) >= 2:
                    hold.append({"학년": g_cur, "반": w["class"], "성명": wnd,
                                 "보류사유": "보류: 학생명부에서 동명이인(A,B,C 등)으로 구분된 이름 – 수동 확인이 필요합니다."})
                    continue
                else:
                    reason = (
                        "자동 제외: 학생명부에 존재하지 않는 학생 – 서버 미등록/학년 불일치 등으로 추정됩니다."
                        if not cand else
                        f"보류: 학년+이름 후보가 2건 이상({len(cand)}건) – 수동 확인 필요."
                    )
                    hold.append({"학년": g_cur, "반": w["class"], "성명": wnd, "보류사유": reason})
                    continue

        if len(matches) > 1:
            hold.append({"학년": g_cur, "반": w["class"], "성명": wnd,
                         "보류사유": f"중복 매칭({len(matches)}건)"})
            continue

        m = matches[0]
        done.append({
            "퇴원반명": m.get("now_class") or m.get("class") or w.get("class"),
            "학생이름": wnd,
            "아이디": m["id"],
            "퇴원일자": eff,
        })

    return done, hold


# =========================
# L3. Output builders (workbook write utils)
# =========================

def _normalized_header_map(ws, header_row: int = 1) -> Dict[str, int]:
    raw = header_map(ws, header_row)
    return {normalize_header_cell(k): v for k, v in raw.items()}


def _require_headers(ws, sheet_label: str, required_labels: List[str], header_row: int = 1) -> Dict[str, int]:
    hm = _normalized_header_map(ws, header_row)
    for label in required_labels:
        key = normalize_header_cell(label)
        if key not in hm:
            raise ValueError(f"[ERROR] {sheet_label}에 '{label}' 열이 없습니다.")
    return hm


def find_last_data_row(ws, key_col: int, start_row: int) -> int:
    last = start_row - 1
    MAX_BLANK = 20
    blank_streak = 0
    for i, row_tuple in enumerate(ws.iter_rows(min_row=start_row, min_col=key_col,
                                                max_col=key_col, values_only=True)):
        v = row_tuple[0] if row_tuple else None
        if v is not None and str(v).strip() != "":
            last = start_row + i
            blank_streak = 0
        else:
            blank_streak += 1
            if blank_streak >= MAX_BLANK:
                break
    return last


def clear_sheet_rows(ws, start_row=2):
    if ws.max_row >= start_row:
        ws.delete_rows(start_row, ws.max_row - start_row + 1)


def move_sheet_after(wb, sheet_name: str, after_name: str):
    if sheet_name not in wb.sheetnames or after_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]
    wb._sheets.remove(ws)
    wb._sheets.insert(wb.sheetnames.index(after_name) + 1, ws)


def delete_rows_below(ws, last_keep_row: int):
    if ws.max_row > last_keep_row:
        ws.delete_rows(last_keep_row + 1, ws.max_row - last_keep_row)


def write_withdraw_to_register(wb, done_rows: List[Dict], hold_rows: List[Dict]):
    ws_done = wb["퇴원"] if "퇴원" in wb.sheetnames else wb.create_sheet("퇴원")
    done_rows = sorted(done_rows, key=lambda r: (str(r.get("퇴원반명", "")), str(r.get("학생이름", ""))))
    clear_sheet_rows(ws_done, 2)

    r = 2
    for row in done_rows:
        write_text_cell(ws_done, r, 1, row["퇴원반명"])
        write_text_cell(ws_done, r, 2, row["학생이름"])
        write_text_cell(ws_done, r, 3, row["아이디"])
        _d = row["퇴원일자"]
        write_text_cell(ws_done, r, 4, _d.strftime("%Y-%m-%d") if hasattr(_d, "strftime") else str(_d))
        r += 1

    ws_hold = None
    if hold_rows:
        hold_rows = sorted(hold_rows, key=lambda r: (str(r.get("학년", "")), str(r.get("반", "")), str(r.get("성명", ""))))
        ws_hold = wb["퇴원_보류"] if "퇴원_보류" in wb.sheetnames else wb.create_sheet("퇴원_보류")
        ws_hold.delete_rows(1, ws_hold.max_row)
        for i, h in enumerate(["학년", "반", "성명", "보류사유"], 1):
            write_text_cell(ws_hold, 1, i, h)
        r = 2
        for row in hold_rows:
            write_text_cell(ws_hold, r, 1, row.get("학년", ""))
            write_text_cell(ws_hold, r, 2, row.get("반", ""))
            write_text_cell(ws_hold, r, 3, row.get("성명", ""))
            write_text_cell(ws_hold, r, 4, row.get("보류사유", ""))
            r += 1
        move_sheet_after(wb, "퇴원_보류", "퇴원")
    else:
        if "퇴원_보류" in wb.sheetnames:
            wb.remove(wb["퇴원_보류"])

    def _fmt(ws):
        for rr in range(1, ws.max_row + 1):
            for cc in range(1, ws.max_column + 1):
                cell = ws.cell(rr, cc)
                cell.font = Font(size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center")

    _fmt(ws_done)
    if ws_hold:
        _fmt(ws_hold)


def write_transfer_hold_sheet(wb, hold_rows: List[Dict]):
    sn = "전입_보류"
    ws = wb[sn] if sn in wb.sheetnames else wb.create_sheet(sn)
    ws.delete_rows(1, ws.max_row)
    for i, h in enumerate(["학년", "반", "번호", "성명", "보류사유"], 1):
        write_text_cell(ws, 1, i, h)
    for i, row in enumerate(hold_rows, 2):
        write_text_cell(ws, i, 1, row.get("grade", ""))
        write_text_cell(ws, i, 2, row.get("class", ""))
        write_text_cell(ws, i, 3, row.get("number", ""))
        write_text_cell(ws, i, 4, row.get("name", ""))
        write_text_cell(ws, i, 5, row.get("보류사유", ""))


def make_kindergarten_class_name(class_value: Any) -> str:
    if class_value is None:
        return "유치원반"
    s = str(class_value).strip()
    if not s:
        return "유치원반"
    return s if s.endswith("반") else f"{s}반"


def make_register_class_name(grade_i: int, class_value: Any) -> str:
    if class_value is None:
        return ""
    s = str(class_value).strip()
    if not s:
        return ""
    m = re.match(r"^\s*0*(\d+)\s*-\s*0*(\d+)\s*$", s)
    if m:
        return f"{int(m.group(1))}-{int(m.group(2))}"
    m2 = re.match(r"^\s*0*(\d+)\s*$", s)
    if m2:
        return f"{grade_i}-{int(m2.group(1))}"
    return f"{grade_i}-{s}"


def fill_register(
    template_path: Path,
    out_path: Path,
    school_name: str,
    year: str,
    freshmen_rows: List[Dict],
    transfer_done_rows: List[Dict],
    teacher_rows: List[Dict],
    transfer_hold_rows: Optional[List[Dict]] = None,
    withdraw_done_rows: Optional[List[Dict]] = None,
    withdraw_hold_rows: Optional[List[Dict]] = None,
    school_kind_override: Optional[str] = None,
) -> None:
    ensure_xlsx_only(template_path)

    wb = load_workbook(template_path)

    try:
        ws_students = wb["학생자료"]
        ws_staff    = wb["직원정보"]
        ws_groups   = wb["그룹반정보"]

        hm = _require_headers(
            ws_students,
            "등록 템플릿 [학생자료]",
            ["No", "학생이름", "ID", "학교구분", "학교", "학년", "수강반"],
            header_row=1,
        )

        col_no     = hm[normalize_header_cell("No")]
        col_name   = hm[normalize_header_cell("학생이름")]
        col_id     = hm[normalize_header_cell("ID")]
        col_kind   = hm[normalize_header_cell("학교구분")]
        col_school = hm[normalize_header_cell("학교")]
        col_grade  = hm[normalize_header_cell("학년")]
        col_class  = hm[normalize_header_cell("수강반")]

        for r in range(2, ws_students.max_row + 1):
            for c in [col_no, col_name, col_id, col_kind, col_school, col_grade, col_class]:
                ws_students.cell(row=r, column=c).value = None

        school_profile = school_profile_from_name(school_name)
        school_profile = apply_school_kind_override(school_profile, school_kind_override)

        def write_student_row(r, no, name, uid, grade_i, cls_name, school_kind_text=None):
            write_text_cell(ws_students, r, col_no, no)
            write_text_cell(ws_students, r, col_name, name)
            write_text_cell(ws_students, r, col_id, uid)
            resolved_kind_full, resolved_kind_prefix = resolve_school_kind_by_grade(school_profile, grade_i)
            write_text_cell(
                ws_students,
                r,
                col_kind,
                school_kind_text if school_kind_text is not None else (resolved_kind_full or ""),
            )
            write_text_cell(ws_students, r, col_school, school_name)
            grade_text = grade_i if isinstance(grade_i, str) else (f"{resolved_kind_prefix}{grade_i}" if resolved_kind_prefix else "")
            write_text_cell(ws_students, r, col_grade, grade_text)
            write_text_cell(ws_students, r, col_class, cls_name)

        write_row = 2
        running_no = 1

        def _fn_dup_key(row):
            if row.get("is_kindergarten"):
                return ("KINDER", normalize_name_key(row.get("name", "")))
            return (int(row.get("grade", 0)), normalize_name_key(row.get("name", "")))

        total = Counter(_fn_dup_key(r) for r in freshmen_rows)
        seen = Counter()
        fn_names_sfx = []
        for fr in freshmen_rows:
            base = str(fr.get("name", "")).strip()
            key = _fn_dup_key(fr)
            if total[key] <= 1:
                fn_names_sfx.append(base)
            else:
                seen[key] += 1
                fn_names_sfx.append(f"{base}{dedup_suffix_letters(seen[key])}")

        fn_ids = [
            nm if fr.get("is_kindergarten") else f"{fr.get('id_year', int(year))}{nm}"
            for fr, nm in zip(freshmen_rows, fn_names_sfx)
        ]

        for i, fr in enumerate(freshmen_rows):
            if fr.get("is_kindergarten"):
                reg_grade = fr.get("grade_label", "유치부")
                reg_cls = make_kindergarten_class_name(fr.get("class"))
                skt = "유치부"
            else:
                reg_grade = fr.get("register_grade", fr["grade"])
                reg_cls = fr.get("register_class_name") or make_register_class_name(fr["grade"], fr["class"])
                skt = None
            write_student_row(write_row + i, running_no, fn_names_sfx[i], fn_ids[i], reg_grade, reg_cls, skt)
            running_no += 1
        write_row += len(freshmen_rows)

        for tr in transfer_done_rows:
            write_student_row(
                write_row,
                running_no,
                tr["name"],
                tr["id"],
                tr["grade"],
                make_register_class_name(tr["grade"], tr["class"]),
            )
            running_no += 1
            write_row += 1

        teachers_learn = [t for t in teacher_rows if t["learn_apply"]]
        t_names_sfx = apply_suffix_for_duplicates([t["name"] for t in teachers_learn])
        for j, t in enumerate(teachers_learn):
            nm_sfx = t_names_sfx[j]
            write_student_row(write_row + j, running_no, nm_sfx, f"{nm_sfx}1", 1, "선생님반")
            running_no += 1
        write_row += len(teachers_learn)

        hm2 = _require_headers(
            ws_staff,
            "등록 템플릿 [직원정보]",
            ["No", "이름", "아이디", "권한부여"],
            header_row=1,
        )

        col_staff_no   = hm2[normalize_header_cell("No")]
        col_staff_name = hm2[normalize_header_cell("이름")]
        col_staff_id   = hm2[normalize_header_cell("아이디")]
        col_staff_auth = hm2[normalize_header_cell("권한부여")]

        for r in range(2, ws_staff.max_row + 1):
            for c in [col_staff_no, col_staff_name, col_staff_id, col_staff_auth]:
                ws_staff.cell(row=r, column=c).value = None

        teachers_admin = [t for t in teacher_rows if t["admin_apply"]]
        a_names_sfx = apply_suffix_for_duplicates([t["name"] for t in teachers_admin])

        for i, t in enumerate(teachers_admin):
            nm = a_names_sfx[i]
            write_text_cell(ws_staff, 2 + i, col_staff_no, i + 1)
            write_text_cell(ws_staff, 2 + i, col_staff_name, nm)
            write_text_cell(ws_staff, 2 + i, col_staff_id, nm)
            write_text_cell(ws_staff, 2 + i, col_staff_auth, "선생님")

        hm_g = _require_headers(
            ws_groups,
            "등록 템플릿 [그룹반정보]",
            ["그룹명", "반명", "수강료", "담임명", "FullMode"],
            header_row=1,
        )

        col_group_name = hm_g[normalize_header_cell("그룹명")]
        col_group_cls  = hm_g[normalize_header_cell("반명")]
        col_group_fee  = hm_g[normalize_header_cell("수강료")]
        col_group_tch  = hm_g[normalize_header_cell("담임명")]
        col_group_full = hm_g[normalize_header_cell("FullMode")]

        for r in range(2, ws_groups.max_row + 1):
            for c in [col_group_name, col_group_cls, col_group_fee, col_group_tch, col_group_full]:
                ws_groups.cell(row=r, column=c).value = None

        class_set = set()
        last_r = find_last_data_row(ws_students, key_col=col_no, start_row=2)
        for r in range(2, last_r + 1):
            v = ws_students.cell(row=r, column=col_class).value
            if v and str(v).strip():
                class_set.add(str(v).strip())

        def _parse_gc(cn):
            m = re.match(r"^\s*(\d+)\s*-\s*(\d+)\s*$", str(cn))
            if m:
                return int(m.group(1)), int(m.group(2))
            m = re.match(r"^\s*(\d+)\s*-\s*(.+)\s*$", str(cn))
            if m:
                return int(m.group(1)), None
            return None, None

        def _group(cn):
            s = str(cn).strip()
            if s == "유치원반":
                return "기타그룹"
            g, _ = _parse_gc(s)
            return f"{g}학년" if g else "기타그룹"

        def _sort_key(cn):
            s = str(cn).strip()
            if s == "유치원반":
                return (1, 0, 0, s)
            if s == "선생님반":
                return (2, 0, 0, s)
            g, c = _parse_gc(s)
            if g is None:
                return (1, 1, 0, s)
            return (0, g, c if c is not None else 9999, s)

        for i, cn in enumerate(sorted(class_set, key=_sort_key), 2):
            write_text_cell(ws_groups, i, col_group_name, _group(cn))
            write_text_cell(ws_groups, i, col_group_cls, cn)
            ws_groups.cell(i, col_group_fee).value = None
            write_text_cell(ws_groups, i, col_group_tch, "선생님")
            write_text_cell(ws_groups, i, col_group_full, "Y")

        if transfer_hold_rows:
            write_transfer_hold_sheet(wb, transfer_hold_rows)
        if withdraw_done_rows is not None and withdraw_hold_rows is not None:
            write_withdraw_to_register(wb, withdraw_done_rows, withdraw_hold_rows)

        clear_format_workbook_from_row(wb, start_row=2)
        reset_view_to_a1(wb)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        backup_if_exists(out_path)
        wb.save(out_path)

    finally:
        wb.close()


# =========================
# L3.5 Notice helpers
# =========================
def _parse_grade_class_from_register(raw: Any) -> Tuple[Optional[int], str]:
    if raw is None:
        return None, ""
    s = raw.strftime("%m-%d") if isinstance(raw, (datetime, date)) else str(raw).strip()
    if not s:
        return None, ""
    m = re.match(r"^\s*0*(\d+)\s*-\s*0*(\d+)\s*$", s)
    if not m:
        return None, s
    return int(m.group(1)), str(int(m.group(2)))


FILL_TRANSFER = PatternFill("solid", fgColor="F8CBAD")
FILL_DUP      = PatternFill("solid", fgColor="FFFF00")
FILL_GREY     = PatternFill("solid", fgColor="D9D9D9")


def build_notice_student_sheet(ws_notice, register_students_ws, transfer_ids: set, transfer_highlight_ids: set) -> List[int]:
    """동명이인 행 번호(1-based, running_no 기준) 리스트를 반환한다."""

    hm_r = _require_headers(
        register_students_ws,
        "등록작업파일 [학생자료]",
        ["No", "학생이름", "ID", "수강반"],
        header_row=1,
    )

    c_r_no   = hm_r[normalize_header_cell("No")]
    c_r_name = hm_r[normalize_header_cell("학생이름")]
    c_r_id   = hm_r[normalize_header_cell("ID")]
    c_r_cls  = hm_r[normalize_header_cell("수강반")]

    start_row = 4

    if ws_notice.max_row >= start_row:
        for r in range(start_row, ws_notice.max_row + 1):
            for c in range(1, 7):
                ws_notice.cell(row=r, column=c).value = None

    tmp_rows: List[Dict[str, Any]] = []
    name_counter: Counter = Counter()

    last_r = find_last_data_row(register_students_ws, key_col=c_r_no, start_row=2)
    for r in range(2, last_r + 1):
        nm  = register_students_ws.cell(r, c_r_name).value
        uid = register_students_ws.cell(r, c_r_id).value
        cls = register_students_ws.cell(r, c_r_cls).value
        cls_str = "" if cls is None else str(cls).strip()
        if cls_str == "선생님반": continue
        if (nm is None or str(nm).strip() == "") and (uid is None or str(uid).strip() == ""): continue
        uid_s = "" if uid is None else str(uid).strip()
        if not uid_s: continue
        nm_s = "" if nm is None else str(nm).strip()
        grade, cls_only = _parse_grade_class_from_register(cls)
        g_disp = grade if grade is not None else ""
        cls_disp = cls_only if grade is not None else cls_str
        nk = notice_name_key(nm_s)
        key = (grade, nk)
        tmp_rows.append({"grade": g_disp, "class_disp": cls_disp, "name": nm_s, "id": uid_s,
                          "key": key, "is_transfer": uid_s in transfer_ids,
                          "is_transfer_highlight": uid_s in transfer_highlight_ids})
        if grade is not None and nk:
            name_counter[key] += 1

    cur_row = start_row; running_no = 1
    dup_row_numbers: List[int] = []
    for rec in tmp_rows:
        dup_flag = name_counter.get(rec["key"], 0) >= 2
        if rec["is_transfer"] and rec.get("is_transfer_highlight"):
            dup_flag = True

        cells = [
            write_text_cell(ws_notice, cur_row, 1, running_no),
            write_text_cell(ws_notice, cur_row, 2, rec["grade"]),
            write_text_cell(ws_notice, cur_row, 3, rec["class_disp"]),
            write_text_cell(ws_notice, cur_row, 4, rec["name"]),
            write_text_cell(ws_notice, cur_row, 5, rec["id"]),
            write_text_cell(ws_notice, cur_row, 6, "1234"),
        ]
        if rec["is_transfer"]:
            for c in cells: c.fill = FILL_TRANSFER
        if dup_flag:
            for c in cells: c.fill = FILL_DUP
            dup_row_numbers.append(running_no)
        running_no += 1; cur_row += 1
    return dup_row_numbers


def build_notice_teacher_sheet(ws_notice, teacher_rows, learn_ids=None, admin_ids=None) -> List[int]:
    """동명이인 행 번호(1-based, no 기준) 리스트를 반환한다."""
    start_row = 4
    try: ws_notice.column_dimensions["B"].width = 16.6
    except Exception: pass

    # 교사 동명이인: 입력 명단 기준으로 같은 이름이 2번 이상이면 동명이인
    from collections import Counter
    name_counter: Counter = Counter(
        str(t.get("name") or "").strip()
        for t in teacher_rows
        if str(t.get("name") or "").strip()
    )

    admin_total = sum(1 for t in teacher_rows if t.get("admin_apply"))
    learn_total = sum(1 for t in teacher_rows if t.get("learn_apply"))
    a_ids = admin_ids or []; l_ids = learn_ids or []
    use_admin = admin_total > 0 and len(a_ids) >= admin_total
    use_learn = learn_total > 0 and len(l_ids) >= learn_total
    ia = il = 0; r_out = start_row; no = 1
    dup_row_numbers: List[int] = []

    for t in teacher_rows:
        pos = "" if t.get("position") is None else str(t["position"]).strip()
        nm  = "" if t.get("name")     is None else str(t["name"]).strip()
        if not nm and not pos and not t.get("learn_apply") and not t.get("admin_apply"):
            continue
        aa = bool(t.get("admin_apply")); la = bool(t.get("learn_apply"))
        admin_id = (a_ids[ia] if use_admin else nm) if aa else ""; ia += aa
        admin_pw = "t1234" if admin_id else ""
        learn_id = (l_ids[il] if use_learn else f"{nm}1") if la else ""; il += la
        learn_pw = "1234" if learn_id else ""
        write_text_cell(ws_notice, r_out, 1, no)
        write_text_cell(ws_notice, r_out, 2, pos)
        write_text_cell(ws_notice, r_out, 3, nm)
        write_text_cell(ws_notice, r_out, 5, admin_id)
        write_text_cell(ws_notice, r_out, 6, admin_pw)
        write_text_cell(ws_notice, r_out, 8, learn_id)
        write_text_cell(ws_notice, r_out, 9, learn_pw)
        if not aa:
            for c in [5, 6]: ws_notice.cell(r_out, c).fill = FILL_GREY
        if not la:
            for c in [8, 9]: ws_notice.cell(r_out, c).fill = FILL_GREY
        if nm and name_counter[nm] >= 2:
            for c in range(1, 10): ws_notice.cell(r_out, c).fill = FILL_DUP
            dup_row_numbers.append(no)
        no += 1; r_out += 1

    delete_rows_below(ws_notice, r_out - 1)
    return dup_row_numbers


def build_notice_file(
    template_notice_path: Path,
    out_notice_path: Path,
    out_register_path: Path,
    teacher_rows: Optional[List[Dict]],
    transfer_done_rows: List[Dict],
) -> List[int]:
    
    ensure_xlsx_only(template_notice_path)
    ensure_xlsx_only(out_register_path)

    wb_notice = safe_load_workbook(template_notice_path, data_only=False)
    wb_reg    = load_workbook(out_register_path)

    try:
        if "학생자료" not in wb_reg.sheetnames:
            raise ValueError("[ERROR] 등록작업파일에 '학생자료' 시트가 없습니다.")

        def _norm(s):
            return re.sub(r"\s+", "", (s or "").replace("\u00A0", " "))

        def _pick(wb, kws):
            ks = [_norm(k) for k in kws]
            for n in wb.sheetnames:
                if all(k in _norm(n) for k in ks):
                    return n
            raise ValueError(f"[ERROR] 안내 템플릿에서 시트를 찾을 수 없습니다. 템플릿 파일을 확인해 주세요.")

        ws_ns = wb_notice[_pick(wb_notice, ["학생", "PW", "학습용"])]
        ws_nt = wb_notice[_pick(wb_notice, ["선생님", "PW"])]
        ws_rs = wb_reg["학생자료"]

        transfer_ids: set = {
            str(tr["id"]).strip()
            for tr in transfer_done_rows
            if tr.get("id")
        }

        transfer_highlight_ids: set = {
            str(tr["id"]).strip()
            for tr in transfer_done_rows
            if tr.get("id") and tr.get("needs_highlight")
        }

        dup_row_numbers = build_notice_student_sheet(ws_ns, ws_rs, transfer_ids, transfer_highlight_ids)

        learn_ids_from_reg: Optional[List[str]] = None
        try:
            hm_r = _require_headers(
                ws_rs,
                "등록작업파일 [학생자료]",
                ["수강반", "ID"],
                header_row=1,
            )
            cc = hm_r[normalize_header_cell("수강반")]
            ci = hm_r[normalize_header_cell("ID")]

            tmp = [
                str(ws_rs.cell(r, ci).value).strip()
                for r in range(2, (ws_rs.max_row or 1) + 1)
                if str(ws_rs.cell(r, cc).value or "").strip() == "선생님반"
                and ws_rs.cell(r, ci).value
            ]
            if tmp:
                learn_ids_from_reg = tmp
        except Exception:
            pass

        admin_ids_from_reg: Optional[List[str]] = None
        try:
            if "직원정보" in wb_reg.sheetnames:
                ws_s = wb_reg["직원정보"]

                hm_s = _require_headers(
                    ws_s,
                    "등록작업파일 [직원정보]",
                    ["아이디"],
                    header_row=1,
                )
                ci = hm_s[normalize_header_cell("아이디")]

                tmp = [
                    str(ws_s.cell(r, ci).value).strip()
                    for r in range(2, (ws_s.max_row or 1) + 1)
                    if ws_s.cell(r, ci).value
                ]
                if tmp:
                    admin_ids_from_reg = tmp
        except Exception:
            pass

        teacher_rows = teacher_rows or []
        teacher_dup_rows = build_notice_teacher_sheet(
            ws_nt,
            teacher_rows,
            learn_ids=learn_ids_from_reg,
            admin_ids=admin_ids_from_reg,
        )

        out_notice_path.parent.mkdir(parents=True, exist_ok=True)
        backup_if_exists(out_notice_path)
        clear_format_workbook_from_row(wb_notice, start_row=4)
        reset_view_to_a1(wb_notice)
        wb_notice.save(out_notice_path)
        return dup_row_numbers, teacher_dup_rows

    finally:
        wb_reg.close()
        wb_notice.close()


def render_mail_text(mail_template_text: str, school_name: str, domain: str) -> str:
    txt = mail_template_text or ""
    if school_name:
        txt = txt.replace("OO초", school_name).replace("OO중", school_name).replace("OO고", school_name)
    if domain:
        txt = re.sub(r"[A-Za-z0-9\-]+\.readinggate\.com", domain, txt)
    return txt


# =========================
# L4. Execute helpers + pipeline
def _extract_layout(
    layout_overrides: Dict[str, Any],
    kind: str,
    default_header: Optional[int] = None,
    scan_meta: Optional[Dict[str, Any]] = None,
):
    info = layout_overrides.get(kind)

    # 1) 사용자 수동 수정
    if isinstance(info, dict):
        h = info.get("header_row")
        s = info.get("data_start_row")

        # 둘 다 있으면 그대로 사용
        if h is not None:
            return h, s

        # data_start_row만 수정한 경우:
        # header_row는 scan_meta에서 보강하고, 없으면 default 사용
        if s is not None:
            if scan_meta and isinstance(scan_meta, dict):
                h_scan = scan_meta.get("header_row")
                if h_scan is not None:
                    return h_scan, s
            return default_header, s

    # 2) 스캔 자동 감지
    if scan_meta and isinstance(scan_meta, dict):
        h = scan_meta.get("header_row")
        s = scan_meta.get("data_start_row")
        if h is not None:
            return h, s

    # 3) 정말 아무것도 없을 때만 fallback
    return default_header, None

def execute_pipeline(
    scan: ScanResult,
    work_date: date,
    school_start_date: Optional[date] = None,
    layout_overrides: Optional[Dict[str, Dict[str, int]]] = None,
    school_kind_override: Optional[str] = None,
) -> PipelineResult:
    logs: List[str] = []  # 실행 로그만 담음 (스캔 로그는 scan.logs에 별도 보존)

    def log(msg):
        from datetime import datetime as _dt
        logs.append(f"[{_dt.now().strftime('%H:%M:%S')}] {msg}")

    layout_overrides = layout_overrides or {}

    # grade_year_map 오버라이드 — UI에서 수동 설정한 학년별 학년도
    # 우선순위: 수동 입력 > 명부 최빈값
    grade_year_map: Dict[int, int] = {
        int(k): int(v)
        for k, v in (layout_overrides.get("grade_year_map", {}) or {}).items()
        if str(k).strip() != "" and str(v).strip() != ""
    }
    if grade_year_map and not scan.roster_info:
        scan.roster_info = RosterInfo(
            roster_time="manual",
            ref_grade_shift=0,
            prefix_mode_by_roster_grade={},
        )

    try:
        if not scan.ok:
            raise ValueError("[ERROR] 스캔 결과가 유효하지 않아 실행할 수 없습니다.")

        school_name = scan.school_name
        year_str    = scan.year_str
        year_int    = scan.year_int or int(year_str)

        log(f"[INFO] 실행 시작 | 학교={school_name}, 학년도={year_str}")
        log(f"[INFO] 작업 폴더: {scan.output_dir}")

        freshmen_path = scan.freshmen_file
        teacher_path  = scan.teacher_file
        transfer_path = scan.transfer_file
        withdraw_path = scan.withdraw_file

        # 신입생
        freshmen_rows: List[Dict] = []
        if freshmen_path:
            h, s = _extract_layout(layout_overrides, "freshmen", 2, scan_meta=scan.freshmen)
            log(f"[DEBUG] 신입생 layout: header_row={h}, data_start_row={s or 'auto'}")

            freshmen_rows = read_freshmen_rows(
            freshmen_path,
            input_year=year_int,
            header_row=h,
            data_start_row=s,
            roster_info=scan.roster_info,
            school_name=scan.school_name,
                manual_grade_year_map=grade_year_map,
            )
            
            log(f"[OK] 신입생 {len(freshmen_rows)}명 로드")
        else:
            log("[INFO] 신입생 파일 없음 → 신입생 등록은 스킵합니다.")

        # 교사
        if teacher_path:
            h, s = _extract_layout(
                    layout_overrides,
                    "teacher",
                    default_header=2,
                    scan_meta=scan.teachers,
                )
            
            log(f"[DEBUG] 교사 layout: header_row={h}, data_start_row={s or 'auto'}")
            teacher_rows = read_teacher_rows(teacher_path, header_row=h, data_start_row=s)
            log(f"[OK] 교사 신청 {len(teacher_rows)}건 로드")
            _teacher_no_id_warn = teacher_rows and not any(
                r.get("admin_apply") or r.get("learn_apply") for r in teacher_rows
            )
        else:
            teacher_rows = []; log("[INFO] 교사 파일 없음 → 교사 관련 처리는 스킵")
            _teacher_no_id_warn = False

        # 전입
        if transfer_path:
            h, s = _extract_layout(layout_overrides, "transfer", 2, scan_meta=scan.transfer_in)
            log(f"[DEBUG] 전입생 layout: header_row={h}, data_start_row={s or 'auto'}")
            transfer_rows = read_transfer_rows(transfer_path, header_row=h, data_start_row=s)
            log(f"[OK] 전입생 {len(transfer_rows)}명 로드")
        else:
            transfer_rows = []; log("[INFO] 전입생 파일 없음 → 전입 처리 스킵")

        # 전출
        if withdraw_path:
            h, s = _extract_layout(layout_overrides, "withdraw", 2, scan_meta=scan.transfer_out)
            log(f"[DEBUG] 전출생 layout: header_row={h}, data_start_row={s or 'auto'}")
            withdraw_rows = read_withdraw_rows(withdraw_path, header_row=h, data_start_row=s)
            log(f"[OK] 전출생 {len(withdraw_rows)}명 로드")
        else:
            withdraw_rows = []; log("[INFO] 전출생 파일 없음 → 전출 처리 스킵")

        # 전입 ID 생성
        transfer_done_rows: List[Dict] = []
        transfer_hold_rows: List[Dict] = []
        if transfer_rows:
            if not scan.roster_info:
                raise ValueError("[ERROR] 전입생 처리를 위해 학생명부가 필요합니다. 학생명부를 준비해 주세요.")
            transfer_done_rows, transfer_hold_rows, _ = build_transfer_ids(
                transfer_rows=transfer_rows, roster_info=scan.roster_info,
                input_year=year_int, freshmen_rows=freshmen_rows,
            )
            # 명부에 이미 존재하는 전입생 → hold로 분리
            # roster_match 유형에 따라 보류사유 메시지 구분
            _roster_dup_rows = [r for r in transfer_done_rows if r.get("dup_with_roster")]
            transfer_done_rows = [r for r in transfer_done_rows if not r.get("dup_with_roster")]
            for _rd in _roster_dup_rows:
                if _rd.get("roster_match") == "exact":
                    _rd["보류사유"] = "학생명부에 이미 존재하는 학생입니다."
                else:
                    _rd["보류사유"] = "학생명부에 동일인으로 의심되는 학생이 있습니다. 수동 확인이 필요합니다."
            transfer_hold_rows = transfer_hold_rows + _roster_dup_rows
            log(f"[OK] 전입 ID 매칭 완료 | 완료 {len(transfer_done_rows)}명, 보류 {len(transfer_hold_rows)}명")
        else:
            log("[INFO] 전입생 없음 → 전입 ID 생성 스킵")

        transfer_freshmen_dup_rows: List[Dict] = []
        if freshmen_rows and transfer_done_rows:
            transfer_done_rows, transfer_freshmen_dup_rows = split_transfer_dup_against_freshmen(
                freshmen_rows=freshmen_rows,
                transfer_done_rows=transfer_done_rows,
            )
            if transfer_freshmen_dup_rows:
                log(f"[INFO] 신입생 명단과 학년/반/이름이 동일한 전입생 {len(transfer_freshmen_dup_rows)}명 → 보류 처리")

        # 전출 퇴원 리스트
        withdraw_done_rows: List[Dict] = []
        withdraw_hold_rows: List[Dict] = []
        transfer_out_auto_skip: int = 0

        if withdraw_rows:
            if not scan.roster_path:
                raise ValueError("[ERROR] 전출생 처리에는 학생명부 파일이 필요합니다. 학생명부를 준비해 주세요.")
            if not scan.roster_info:
                raise ValueError("[ERROR] 전출생 처리에 필요한 학생명부 정보를 읽지 못했습니다.")
            if school_start_date is None:
                raise ValueError("[ERROR] 전출 처리에 필요한 개학일이 입력되지 않았습니다.")

            roster_wb2 = safe_load_workbook(scan.roster_path, data_only=True)
            try:
                roster_ws2 = get_first_sheet_with_warning(
                    roster_wb2, scan.roster_path.name, logs=logs
                )

                withdraw_done_rows, withdraw_hold_rows = build_withdraw_outputs(
                    roster_ws=roster_ws2,
                    withdraw_rows=withdraw_rows,
                    school_start_date=school_start_date,
                    work_date=work_date,
                    roster_info=scan.roster_info,
                )
            finally:
                roster_wb2.close()

            transfer_out_auto_skip = sum(
                1 for r in withdraw_hold_rows if str(r.get("보류사유", "")).startswith("자동 제외")
            )
            log(
                f"[OK] 전출 퇴원 리스트 생성 | 퇴원 {len(withdraw_done_rows)}명, "
                f"보류 {len(withdraw_hold_rows)}명 (자동 제외 {transfer_out_auto_skip}명 포함)"
            )
        else:
            log("[INFO] 전출생 없음 → 퇴원 처리 스킵")

        # 등록작업파일
        if not scan.template_register:
            raise ValueError("[ERROR] 등록 템플릿 파일을 찾을 수 없습니다.")

        out_register_path = scan.output_dir / f"★{school_name}_등록작업파일(작업용).xlsx"
        fill_register(
            template_path=scan.template_register,
            out_path=out_register_path,
            school_name=school_name, year=year_str,
            freshmen_rows=freshmen_rows,
            transfer_done_rows=transfer_done_rows,
            teacher_rows=teacher_rows,
            transfer_hold_rows=transfer_hold_rows,
            withdraw_done_rows=withdraw_done_rows,
            withdraw_hold_rows=withdraw_hold_rows,
            school_kind_override=school_kind_override,
        )
        log(f"[OK] 등록작업파일 생성 완료: {out_register_path.name}")

        # 안내파일
        if not scan.template_notice:
            raise ValueError("[ERROR] 안내 템플릿 파일을 찾을 수 없습니다.")

        notice_kinds = (
            (["신입생"] if freshmen_path and freshmen_rows else [])
            + (["전입생"] if transfer_done_rows else [])
            + (["교직원"] if teacher_rows else [])
        )
        title_middle = (",".join(notice_kinds) + "_ID,PW안내") if notice_kinds else "ID,PW안내"
        out_notice_path = scan.output_dir / f"☆{school_name}_{title_middle}.xlsx"

        _notice_result = build_notice_file(
            template_notice_path=scan.template_notice,
            out_notice_path=out_notice_path,
            out_register_path=out_register_path,
            teacher_rows=teacher_rows,
            transfer_done_rows=transfer_done_rows,
        )
        notice_dup_rows, notice_teacher_dup_rows = _notice_result if _notice_result else ([], [])

        log(f"[OK] 안내파일 생성 완료: {out_notice_path.name}")

        pr = PipelineResult(ok=True, outputs=[out_register_path, out_notice_path], logs=logs)
        pr.notice_dup_rows         = notice_dup_rows
        pr.notice_teacher_dup_rows = notice_teacher_dup_rows
        transfer_hold_rows = transfer_hold_rows + transfer_freshmen_dup_rows
        pr.transfer_in_done   = len(transfer_done_rows)
        pr.transfer_in_hold   = len(transfer_hold_rows)
        pr.transfer_out_done  = len(withdraw_done_rows)
        pr.transfer_out_hold  = len(withdraw_hold_rows)
        pr.transfer_out_auto_skip = transfer_out_auto_skip

                # ===== 검수 요약 추가 =====
        transfer_input_count = len(transfer_rows)
        transfer_done_count = len(transfer_done_rows)
        transfer_hold_count = len(transfer_hold_rows)

        withdraw_input_count = len(withdraw_rows)
        withdraw_done_count = len(withdraw_done_rows)
        withdraw_hold_count = len(withdraw_hold_rows)

        transfer_total_match = (
            transfer_input_count == (transfer_done_count + transfer_hold_count)
        )
        withdraw_total_match = (
            withdraw_input_count == (withdraw_done_count + withdraw_hold_count)
        )

        pr.audit_summary = {
            "school_name": scan.school_name,
            "year_str": scan.year_str,

            "input_counts": {
                "freshmen": len(freshmen_rows),
                "teacher": len(teacher_rows),
                "transfer": transfer_input_count,
                "withdraw": withdraw_input_count,
            },

            "result_counts": {
                "transfer_done": transfer_done_count,
                "transfer_hold": transfer_hold_count,
                "withdraw_done": withdraw_done_count,
                "withdraw_hold": withdraw_hold_count,
                "withdraw_auto_skip": transfer_out_auto_skip,
            },

            "checks": {
                "transfer_total_match": transfer_total_match,
                "withdraw_total_match": withdraw_total_match,
            },
        }

        # ===== events 생성 (hold/dup → CoreEvent) =====
        # bridge/UI는 events만 참조하므로 여기서 모두 생성
        # 신입생-전입생 중복 보류 이벤트 (hold)
        for row in transfer_freshmen_dup_rows:
            name  = str(row.get("name", ""))
            grade = int(row.get("grade", 0) or 0)
            cls   = str(row.get("class", ""))
            pr.events.append(_evt_freshmen_transfer_dup(name, grade, cls))

        # 명부 매칭 실패 / 명부 존재 전입생 보류 이벤트
        for row in transfer_hold_rows:
            if row in transfer_freshmen_dup_rows:
                continue  # freshmen_transfer_dup으로 이미 추가됨
            name   = str(row.get("name", ""))
            reason = str(row.get("보류사유", ""))
            if "학생명부에 이미 존재" in reason:
                pr.events.append(_evt_roster_duplicate_transfer(name, reason))
            else:
                pr.events.append(_evt_transfer_in_hold(name, reason))

        for row in withdraw_hold_rows:
            name = str(row.get("성명", ""))
            reason = str(row.get("보류사유", ""))
            if not reason.startswith("자동 제외"):
                pr.events.append(_evt_transfer_out_hold(name, reason))

        dup_count = len(notice_dup_rows)
        if dup_count > 0:
            from core.events import duplicate_name as _evt_dup
            pr.events.append(_evt_dup(dup_count))

        if _teacher_no_id_warn:
            from core.events import no_teacher_id_request as _evt_no_tid
            pr.events.append(_evt_no_tid())

        log("[DONE] 실행 완료")
        return pr

    except Exception as e:
        import traceback as _tb
        _tb_str = _tb.format_exc()
        log(f"[ERROR] {e}")
        log(f"[DEBUG] {_tb_str}")
        try:
            import os as _os
            _app_dir = Path(_os.environ.get("RG_APP_DIR") or Path(__file__).parent.parent)
            _dump = _app_dir / "run_error.log"
            _dump.write_text(_tb_str, encoding="utf-8")
        except Exception:
            pass
        _err_str = str(e)
        _pr_fail = PipelineResult(ok=False, outputs=[], logs=logs)
        if "등록 템플릿" in _err_str:
            _pr_fail.events.append(_evt_template_register_not_found())
        elif "안내 템플릿" in _err_str:
            _pr_fail.events.append(_evt_template_notice_not_found())
        elif "학생명부" in _err_str:
            _pr_fail.events.append(_evt_roster_not_found_at_run())
        elif "개학일" in _err_str:
            _pr_fail.events.append(_evt_open_date_required())
        return _pr_fail


# =========================
# L5. Run wrappers
# =========================
def scan_work_root(work_root: Path) -> Dict[str, Any]:
    """
    작업 루트 점검. app.py가 기대하는 키:
      ok, errors, message, school_folders, notice_titles,
      db_ok, errors_db, db_file,
      format_ok, errors_format, register_template, notice_template
    """
    work_root = work_root.resolve()
    dirs = get_project_dirs(work_root)
    errors: List[str] = []

    res_root = dirs["RESOURCES_ROOT"].resolve()
    school_folders = sorted(
        p.name for p in work_root.iterdir()
        if p.is_dir() and p.resolve() != res_root and not p.name.startswith(".")
    )

    db_ok = False; errors_db: List[str] = []; db_file: Optional[Path] = None
    db_dir = dirs["DB"]
    if not db_dir.exists():
        errors_db.append("[ERROR] resources/DB 폴더가 없습니다.")
    else:
        db_files = [p for p in db_dir.glob("*.xlsb") if "학교전체명단" in p.stem and not p.name.startswith("~$")]
        if   len(db_files) == 0: errors_db.append("[ERROR] DB 폴더에 '학교전체명단' xlsb 파일이 없습니다.")
        elif len(db_files) > 1:  errors_db.append("[ERROR] DB 폴더에 '학교전체명단' xlsb 파일이 2개 이상 있습니다.")
        else: db_ok = True; db_file = db_files[0]

    format_ok = False; errors_format: List[str] = []
    register_template: Optional[Path] = None; notice_template: Optional[Path] = None
    tpl_dir = dirs["TEMPLATES"]
    if not tpl_dir.exists():
        errors_format.append("[ERROR] resources/templates 폴더가 없습니다.")
    else:
        reg_f = [p for p in tpl_dir.glob("*.xlsx") if "등록" in p.stem and not p.name.startswith("~$")]
        ntc_f = [p for p in tpl_dir.glob("*.xlsx") if "안내" in p.stem and not p.name.startswith("~$")]

        if len(reg_f) != 1: 
            errors_format.append("[ERROR] templates 폴더에 '등록' 템플릿 파일이 정확히 1개 있어야 합니다.")
        else: register_template = reg_f[0]

        if len(ntc_f) != 1: 
            errors_format.append("[ERROR] templates 폴더에 '안내' 템플릿 파일이 정확히 1개 있어야 합니다.")
            
        else: notice_template = ntc_f[0]
        if not errors_format: format_ok = True

    notice_titles: List[str] = []
    notice_dir = dirs["NOTICES"]
    if not notice_dir.exists():
        errors.append("[ERROR] resources/notices 폴더가 없습니다.")
    else:
        txt_files = [p for p in notice_dir.glob("*.txt") if p.is_file()]
        if not txt_files: errors.append("[ERROR] notices 폴더에 .txt 파일이 없습니다.")
        else: notice_titles = sorted({p.stem.strip() for p in txt_files})

    errors.extend(errors_db)
    errors.extend(errors_format)
    ok = not errors

    return {
        "ok": ok,
        "errors": errors,
        "message": "[OK] resources(DB/templates/notices)가 정상적으로 준비되었습니다." if ok else "",
        "school_folders": school_folders,
        "notice_titles": notice_titles,
        "db_ok": db_ok, "errors_db": errors_db, "db_file": db_file,
        "format_ok": format_ok, "errors_format": errors_format,
        "register_template": register_template, "notice_template": notice_template,
    }


def run_pipeline(
    work_root: Path,
    school_name: str,
    school_start_date: date,
    work_date: date,
    layout_overrides: Optional[Dict[str, Dict[str, int]]] = None,
    roster_basis_date: Optional[date] = None,
    school_kind_override: Optional[str] = None,
) -> PipelineResult:
    logs: List[str] = []

    def log(msg):
        logs.append(msg)

    work_root = Path(work_root).resolve()
    school_name = (school_name or "").strip()

    if not school_name:
        log("[ERROR] 학교명을 입력해 주세요.")
        return PipelineResult(ok=False, outputs=[], logs=logs)

    try:
        scan = scan_pipeline(
            work_root=work_root,
            school_name=school_name,
            school_start_date=school_start_date,
            work_date=work_date,
            roster_basis_date=roster_basis_date,
        )

        logs.extend(scan.logs)

        if not scan.ok:
            return PipelineResult(ok=False, outputs=[], logs=logs)

        if not scan.can_execute:
            msg = ", ".join(scan.missing_fields) if scan.missing_fields else "필수 파일 누락"
            log(f"[ERROR] 실행할 수 없습니다. ({msg})")
            return PipelineResult(ok=False, outputs=[], logs=logs)

        result = execute_pipeline(
            scan=scan,
            work_date=work_date,
            school_start_date=school_start_date,
            layout_overrides=layout_overrides,
            school_kind_override=school_kind_override,
        )

        result.logs = logs + [m for m in result.logs if m not in logs]
        return result

    except Exception as e:
        log(f"[ERROR] 실행 중 문제가 발생했습니다.: {e}")
        return PipelineResult(ok=False, outputs=[], logs=logs)


def run_pipeline_partial(
    work_root: Path,
    school_name: str,
    open_date: date,
    mode: str,
) -> PipelineResult:
    """
    UI의 '부분 실행' 버튼용.
    현재는 전체 파이프라인을 재생성하는 방식으로 동작.
    mode: 'freshmen'|'teacher'|'transfer'|'withdraw' (아직 구분 안 함)
    """
    return run_pipeline(
        work_root=work_root,
        school_name=school_name,
        school_start_date=open_date,
        work_date=open_date,
        roster_basis_date=None,
    )