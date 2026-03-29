# core/xlsx_db.py
"""
학교 전체 명단 xlsx 파일 기반 학교 검색 모듈.

주의:
- 이 모듈의 대상 파일은 작업 이력/발송일 등을 다시 써야 하는 전체 명단이다.
- 따라서 읽기 단계부터 .xlsx만 허용한다.

기존 db.py(pyxlsb/.xlsb 의존)를 대체한다.
열 위치는 col_map(roster_log.py와 동일 구조)으로 받으며,
resources/DB/ 폴더나 pyxlsb 패키지가 없어도 동작한다.

col_map 구조 (1-based 열 번호, roster_log._DEFAULT_COL_MAP과 동일):
  {
    "sheet":        "학교명단",
    "header_row":   7,
    "data_start":   8,
    "col_school":   5,
    "col_domain":   None,   # 도메인 열 (선택)
    "col_email_arr": 10,
    ...
  }

공개 API:
  load_school_names_from_xlsx(xlsx_path, col_map)      -> List[str]
  search_schools_in_xlsx(xlsx_path, keyword, col_map)  -> List[str]
  get_school_domain_from_xlsx(xlsx_path, school_name, col_map) -> Optional[str]
  school_exists_in_xlsx(xlsx_path, school_name, col_map) -> bool
"""
from __future__ import annotations

import re
from functools import lru_cache
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook

from core.common import ensure_xlsx_only

# =========================
# 기본 col_map (roster_log._DEFAULT_COL_MAP 과 동기화)
# =========================
_DEFAULT_COL_MAP: Dict[str, Any] = {
    "sheet":         "학교명단",
    "header_row":    7,
    "data_start":    8,
    "col_school":    5,
    "col_domain":    None,  # 도메인 열: 없으면 None
    "col_email_arr": 10,
    "col_email_snt": 11,
    "col_worker":    12,
    "col_freshmen":  13,
    "col_transfer":  14,
    "col_withdraw":  15,
    "col_teacher":   16,
}

BLOCK_END_BLANK_STREAK = 3
MAX_SCAN_ROWS = 5000

EXCLUDE_VALUES = {"-", "—", "–", "학교명", "학교이름", "기관명"}


# =========================
# 내부 헬퍼
# =========================
def _resolve_col_map(col_map: Optional[Dict]) -> Dict[str, Any]:
    base = dict(_DEFAULT_COL_MAP)
    if col_map:
        for k, v in col_map.items():
            # 0은 의미없는 값이므로 None 취급, 나머지는 덮어씀
            if v is not None and v != 0:
                base[k] = v
    return base


def _normalize(s: Any) -> str:
    if s is None:
        return ""
    return re.sub(r"\s+", "", str(s)).strip()


def _is_header_like(val: str) -> bool:
    """학교명 열 헤더 행인지 판별 (데이터로 취급 안 함)."""
    norm = _normalize(val)
    return norm in {"학교명", "학교이름", "기관명", "기관이름", "schoolname"}


def _open_wb(xlsx_path: Path):
    xlsx_path = Path(xlsx_path)
    ensure_xlsx_only(xlsx_path)
    try:
        return load_workbook(str(xlsx_path), data_only=True, read_only=True)
    except Exception as e:
        raise IOError(f"명단 파일을 열 수 없습니다: {xlsx_path.name} — {e}") from e


# =========================
# 학교명 목록 로드
# =========================
def load_school_names_from_xlsx(
    xlsx_path: Path,
    col_map: Optional[Dict[str, Any]] = None,
) -> List[str]:
    """
    명단 xlsx에서 학교명 전체 목록을 읽어 반환.
    빈 값·헤더·구분선은 제외.
    """
    xlsx_path = Path(xlsx_path)
    ensure_xlsx_only(xlsx_path)
    if not xlsx_path.exists():
        raise FileNotFoundError(f"명단 파일을 찾을 수 없습니다: {xlsx_path}")

    cm = _resolve_col_map(col_map)
    sheet_name = cm["sheet"]
    data_start  = int(cm["data_start"])
    col_school  = int(cm["col_school"])  # 1-based

    wb = _open_wb(xlsx_path)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"'{sheet_name}' 시트를 찾을 수 없습니다: {xlsx_path.name}")

        ws = wb[sheet_name]
        schools: List[str] = []
        seen: set = set()
        blank_streak = 0

        for r_idx, row in enumerate(ws.iter_rows(min_row=data_start, max_row=data_start + MAX_SCAN_ROWS), start=data_start):
            cells = list(row)
            if col_school - 1 >= len(cells):
                blank_streak += 1
                if blank_streak >= BLOCK_END_BLANK_STREAK:
                    break
                continue

            val = cells[col_school - 1].value
            norm = _normalize(val)

            if not norm:
                blank_streak += 1
                if blank_streak >= BLOCK_END_BLANK_STREAK:
                    break
                continue

            blank_streak = 0

            # 헤더 행 또는 구분선 제외
            if _is_header_like(norm):
                continue
            if all(ch in "-—–" for ch in norm):
                continue
            if norm in {_normalize(x) for x in EXCLUDE_VALUES}:
                continue

            raw = str(val).strip()
            if raw and raw not in seen:
                seen.add(raw)
                schools.append(raw)

        return schools

    finally:
        wb.close()


# =========================
# 학교명 검색 (자동완성용)
# =========================
def search_schools_in_xlsx(
    xlsx_path: Path,
    keyword: str,
    col_map: Optional[Dict[str, Any]] = None,
    limit: int = 30,
) -> List[str]:
    """키워드로 학교명 검색. exact match 우선, 이후 contains."""
    kw = (keyword or "").strip()
    if not kw or not xlsx_path:
        return []

    schools = load_school_names_from_xlsx(xlsx_path, col_map)
    kw_norm = _normalize(kw)

    exacts: List[str] = []
    partials: List[str] = []

    for s in schools:
        snorm = _normalize(s)
        if snorm == kw_norm:
            exacts.append(s)
        elif kw_norm in snorm:
            partials.append(s)

    return (exacts + partials)[:limit]


# =========================
# 도메인 조회
# =========================
def get_school_domain_from_xlsx(
    xlsx_path: Path,
    school_name: str,
    col_map: Optional[Dict[str, Any]] = None,
) -> Optional[str]:
    """
    학교명으로 도메인 값 조회.
    col_map에 col_domain이 없거나 None이면 None 반환.
    """
    target = (school_name or "").strip()
    if not target or not xlsx_path:
        return None

    cm = _resolve_col_map(col_map)
    col_domain = cm.get("col_domain")
    if not col_domain:
        return None

    col_domain = int(col_domain)
    sheet_name = cm["sheet"]
    data_start  = int(cm["data_start"])
    col_school  = int(cm["col_school"])

    xlsx_path = Path(xlsx_path)
    ensure_xlsx_only(xlsx_path)
    if not xlsx_path.exists():
        return None

    target_norm = _normalize(target)
    fallback: Optional[str] = None

    wb = _open_wb(xlsx_path)
    try:
        if sheet_name not in wb.sheetnames:
            return None

        ws = wb[sheet_name]

        for row in ws.iter_rows(min_row=data_start, max_row=data_start + MAX_SCAN_ROWS):
            cells = list(row)

            school_val = cells[col_school - 1].value if col_school - 1 < len(cells) else None
            snorm = _normalize(school_val)
            if not snorm or _is_header_like(snorm):
                continue

            domain_raw = cells[col_domain - 1].value if col_domain - 1 < len(cells) else None
            dom = _normalize_domain(domain_raw)

            if snorm == target_norm:
                return dom or None

            if fallback is None and target_norm in snorm:
                fallback = dom or None

        return fallback

    finally:
        wb.close()


def _normalize_domain(raw: Any) -> str:
    if not raw:
        return ""
    s = str(raw).strip()
    s = re.sub(r"^https?://", "", s, flags=re.I)
    s = s.split("/")[0].strip().lower()
    if s in {"y", "n", "yes", "no", "사용", "미사용", "-", ""}:
        return ""
    return s


# =========================
# 존재 여부 확인
# =========================
def school_exists_in_xlsx(
    xlsx_path: Path,
    school_name: str,
    col_map: Optional[Dict[str, Any]] = None,
) -> bool:
    """학교명이 명단 xlsx에 존재하는지 확인."""
    target = (school_name or "").strip()
    if not target or not xlsx_path:
        return False

    schools = load_school_names_from_xlsx(xlsx_path, col_map)
    target_norm = _normalize(target)

    for s in schools:
        if _normalize(s) == target_norm:
            return True
    # contains fallback
    for s in schools:
        if target_norm in _normalize(s):
            return True
    return False
