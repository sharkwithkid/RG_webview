# core/run_diff.py
"""

재학생 명렬 비교(diff) 파이프라인의 실행 전용 모듈.

책임 범위:
  - scan_diff의 스캔 결과를 입력으로 받음
  - 학생명렬표 비교 수행
  - 전입생 / 전출생 / 보류 대상 산출
  - diff 결과 출력 파일 생성
  - 실행 결과 집계 및 반환

이 모듈은 diff 작업의 실제 실행과 산출물 생성을 담당.

공개 API:
  DiffPipelineResult
  execute_diff_pipeline(scan) -> DiffPipelineResult
  run_diff_pipeline(work_root, school_name, target_year, school_start_date, work_date, roster_basis_date)
  -> DiffPipelineResult
"""
from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook

from core.common import (
    get_project_dirs,
    load_roster_sheet,
    safe_load_workbook,
    ensure_xlsx_only,
)

from core.output_common import (
    backup_if_exists,
    write_text_cell,
    clear_format_workbook_from_row,
    reset_view_to_a1,
)

from core.scan_diff import (
    DiffScanResult,
    scan_diff_pipeline,
    build_diff_rows,
    read_roster_compare_rows,
    read_compare_rows,
    collect_text_only_classes_from_roster,
    TARGET_GRADES,
)


# =========================
# Result types
# =========================
@dataclass
class DiffPipelineResult:
    ok: bool
    outputs: List[Path]
    logs: List[str]

    # 명단 비교 관점 (상위 개념)
    compare_only_count: int = 0   # 학교 명단에만 있음
    roster_only_count: int = 0    # 명부에만 있음
    matched_count: int = 0        # 양쪽 일치 (정상 재학생)
    unresolved_count: int = 0     # 자동 판정 불가

    # 자동 분류 결과 (하위 개념)
    transfer_in_done: int = 0
    transfer_in_hold: int = 0
    transfer_out_done: int = 0
    transfer_out_hold: int = 0

    roster_only_rows: List[Dict[str, Any]] = None
    matched_rows: List[Dict[str, Any]] = None
    compare_only_rows: List[Dict[str, Any]] = None
    unresolved_rows: List[Dict[str, Any]] = None


# =========================
# L3. Compare output builders
# =========================
def _ensure_hold_sheet(wb, base_sheet_name: str = "보류"):
    if "보류" in wb.sheetnames:
        return wb["보류"]

    ws_hold = wb.create_sheet("보류")
    return ws_hold


def build_compare_result_workbook(
    out_path: Path,
    school_name: str,
    roster_only_rows: List[Dict[str, Any]],
    matched_rows: List[Dict[str, Any]],
    compare_only_rows: List[Dict[str, Any]],
    unresolved_rows: List[Dict[str, Any]],
):
    """명단 비교 결과를 단일 엑셀 파일 1개로 저장한다."""
    backup_if_exists(out_path)

    wb = Workbook()
    try:
        ws_summary = wb.active
        ws_summary.title = "비교요약"
        ws_summary.append(["구분", "인원"])
        ws_summary.append(["명부에만 있음", len(roster_only_rows or [])])
        ws_summary.append(["공통 학생", len(matched_rows or [])])
        ws_summary.append(["재학생 명단에만 있음", len(compare_only_rows or [])])
        ws_summary.append(["판정 불가", len(unresolved_rows or [])])

        def add_simple_sheet(title: str, rows: List[Dict[str, Any]]):
            ws = wb.create_sheet(title)
            ws.append(["학년", "반", "이름"])
            for rec in rows or []:
                ws.append([
                    rec.get("grade", ""),
                    rec.get("class", ""),
                    rec.get("name", ""),
                ])
            return ws

        def add_unresolved_sheet(title: str, rows: List[Dict[str, Any]]):
            ws = wb.create_sheet(title)
            ws.append(["학년", "반", "이름", "보류사유"])
            for rec in rows or []:
                ws.append([
                    rec.get("grade", ""),
                    rec.get("class", ""),
                    rec.get("name", ""),
                    rec.get("hold_reason", rec.get("remark", "")),
                ])
            return ws

        add_simple_sheet("명부에만 있음", roster_only_rows)
        add_simple_sheet("공통 학생", matched_rows)
        add_simple_sheet("재학생 명단에만 있음", compare_only_rows)
        add_unresolved_sheet("판정불가", unresolved_rows)

        for ws in wb.worksheets:
            reset_view_to_a1(wb)
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    val = "" if cell.value is None else str(cell.value)
                    if len(val) > max_len:
                        max_len = len(val)
                ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 28)

        wb.save(out_path)
    finally:
        wb.close()


def fill_transfer_in_workbook(
    template_path: Path,
    out_path: Path,
    done_rows: List[Dict[str, Any]],
    hold_rows: List[Dict[str, Any]],
):
    ensure_xlsx_only(template_path)
    backup_if_exists(out_path)

    wb = safe_load_workbook(template_path, data_only=False)
    try:
        ws = wb.worksheets[0]

        clear_format_workbook_from_row(wb, start_row=3)

        # 본 시트: A no / B 학년 / C 반 / D 번호(빈칸) / E 이름 / F 비고
        cur_row = 3
        for idx, rec in enumerate(done_rows, start=1):
            write_text_cell(ws, cur_row, 1, idx)
            write_text_cell(ws, cur_row, 2, rec.get("grade", ""))
            write_text_cell(ws, cur_row, 3, rec.get("class", ""))
            write_text_cell(ws, cur_row, 4, "")  # 번호는 사용 안 함
            write_text_cell(ws, cur_row, 5, rec.get("name", ""))
            write_text_cell(ws, cur_row, 6, rec.get("remark", ""))
            cur_row += 1

        if hold_rows:
            ws_hold = _ensure_hold_sheet(wb)

            write_text_cell(ws_hold, 1, 1, "no")
            write_text_cell(ws_hold, 1, 2, "학년")
            write_text_cell(ws_hold, 1, 3, "반")
            write_text_cell(ws_hold, 1, 4, "이름")
            write_text_cell(ws_hold, 1, 5, "보류사유")

            hold_row = 2
            for idx, rec in enumerate(hold_rows, start=1):
                write_text_cell(ws_hold, hold_row, 1, idx)
                write_text_cell(ws_hold, hold_row, 2, rec.get("grade", ""))
                write_text_cell(ws_hold, hold_row, 3, rec.get("class", ""))
                write_text_cell(ws_hold, hold_row, 4, rec.get("name", ""))
                write_text_cell(ws_hold, hold_row, 5, rec.get("hold_reason", ""))
                hold_row += 1

        reset_view_to_a1(wb)
        wb.save(out_path)

    finally:
        wb.close()


def fill_transfer_out_workbook(
    template_path: Path,
    out_path: Path,
    done_rows: List[Dict[str, Any]],
    hold_rows: List[Dict[str, Any]],
):
    ensure_xlsx_only(template_path)
    backup_if_exists(out_path)

    wb = safe_load_workbook(template_path, data_only=False)
    try:
        ws = wb.worksheets[0]

        clear_format_workbook_from_row(wb, start_row=3)

        # 본 시트: A no / B 학년 / C 반 / D 이름 / E 비고
        cur_row = 3
        for idx, rec in enumerate(done_rows, start=1):
            write_text_cell(ws, cur_row, 1, idx)
            write_text_cell(ws, cur_row, 2, rec.get("grade", ""))
            write_text_cell(ws, cur_row, 3, rec.get("class", ""))
            write_text_cell(ws, cur_row, 4, rec.get("name", ""))
            write_text_cell(ws, cur_row, 5, rec.get("remark", ""))
            cur_row += 1

        if hold_rows:
            ws_hold = _ensure_hold_sheet(wb)

            write_text_cell(ws_hold, 1, 1, "no")
            write_text_cell(ws_hold, 1, 2, "학년")
            write_text_cell(ws_hold, 1, 3, "반")
            write_text_cell(ws_hold, 1, 4, "이름")
            write_text_cell(ws_hold, 1, 5, "보류사유")

            hold_row = 2
            for idx, rec in enumerate(hold_rows, start=1):
                write_text_cell(ws_hold, hold_row, 1, idx)
                write_text_cell(ws_hold, hold_row, 2, rec.get("grade", ""))
                write_text_cell(ws_hold, hold_row, 3, rec.get("class", ""))
                write_text_cell(ws_hold, hold_row, 4, rec.get("name", ""))
                write_text_cell(ws_hold, hold_row, 5, rec.get("hold_reason", ""))
                hold_row += 1

        reset_view_to_a1(wb)
        wb.save(out_path)

    finally:
        wb.close()

# =========================
# L4. Execute
# =========================
def execute_diff_pipeline(
    scan: DiffScanResult,
) -> DiffPipelineResult:
    logs: List[str] = list(scan.logs)

    def log(msg: str):
        logs.append(msg)

    try:
        if not scan.ok:
            raise ValueError("[ERROR] 스캔 결과가 유효하지 않아 실행할 수 없습니다.")
        if not scan.can_execute:
            missing = ", ".join(scan.missing_fields) if scan.missing_fields else "필수 파일 누락"
            raise ValueError(f"[ERROR] 실행할 수 없습니다. ({missing})")

        school_name = scan.school_name
        year_int    = scan.year_int

        log(f"[INFO] 비교 실행 시작 | 학교={school_name}, 학년도={year_int}")
        log(
            f"[DEBUG] 명부 기준일={getattr(scan, 'roster_basis_date', None)} / "
            f"작업일={getattr(scan, 'work_date', None)} / ref_grade_shift={getattr(scan, 'ref_grade_shift', 0)}"
        )

        if not scan.compare_file or not scan.roster_path:
            raise ValueError("[ERROR] 실행에 필요한 파일 경로 정보가 없습니다.")

        # 명부 다시 로드
        dirs = get_project_dirs(scan.project_root)
        roster_wb, roster_ws, roster_path, _ = load_roster_sheet(dirs, school_name)
        try:
            text_only_classes = collect_text_only_classes_from_roster(
                roster_ws,
                target_grades=TARGET_GRADES,
                ref_grade_shift=scan.ref_grade_shift,
            )

            if text_only_classes:
                log(
                    "[WARN] 학생명부에서 반을 확인해 주세요.: "
                    + ", ".join(text_only_classes)
                )

            roster_rows = read_roster_compare_rows(
                roster_ws,
                target_grades=TARGET_GRADES,
                ref_grade_shift=scan.ref_grade_shift,
            )
        finally:
            roster_wb.close()

        compare_rows = read_compare_rows(
            scan.compare_file,
            header_row=scan.compare_layout["header_row"] if scan.compare_layout else None,
            data_start_row=scan.compare_layout["data_start_row"] if scan.compare_layout else None,
        )

        log(f"[INFO] 명부 비교 대상 수: {len(roster_rows)}명")
        log(f"[INFO] 재학생 명렬표 대상 수: {len(compare_rows)}명")

        roster_keys  = {(r["grade"], r["name_key"]) for r in roster_rows}
        compare_keys = {(r["grade"], r["name_key"]) for r in compare_rows}

        compare_only_sample = [
            r for r in compare_rows
            if (r["grade"], r["name_key"]) not in roster_keys
        ][:10]

        roster_only_sample = [
            r for r in roster_rows
            if (r["grade"], r["name_key"]) not in compare_keys
        ][:10]

        diff = build_diff_rows(roster_rows, compare_rows)

        log(f"[DEBUG] compare only 샘플 후보: {compare_only_sample}")
        log(f"[DEBUG] roster only 샘플 후보: {roster_only_sample}")

        matched_rows      = diff["matched_rows"]
        compare_only_rows = diff["compare_only_rows"]
        roster_only_rows  = diff["roster_only_rows"]
        unresolved_rows   = diff["unresolved_rows"]
        transfer_in_done  = diff["transfer_in_done"]
        transfer_in_hold  = diff["transfer_in_hold"]
        transfer_out_done = diff["transfer_out_done"]
        transfer_out_hold = diff["transfer_out_hold"]

        log(f"[INFO] 비교 요약 | 일치 {len(matched_rows)}명 / "
            f"학교명단에만 {len(compare_only_rows)}명 / "
            f"명부에만 {len(roster_only_rows)}명 / "
            f"판정불가 {len(unresolved_rows)}명")
        log(f"[DEBUG] 전입 완료 {len(transfer_in_done)} / 보류 {len(transfer_in_hold)}")
        log(f"[DEBUG] 전출 완료 {len(transfer_out_done)} / 보류 {len(transfer_out_hold)}")

        out_compare = scan.output_dir / f"{school_name}_명단비교 결과.xlsx"

        build_compare_result_workbook(
            out_path=out_compare,
            school_name=school_name,
            roster_only_rows=roster_only_rows,
            matched_rows=matched_rows,
            compare_only_rows=compare_only_rows,
            unresolved_rows=unresolved_rows,
        )
        log(f"[OK] 명단 비교 결과 생성 완료: {out_compare.name}")

        pr = DiffPipelineResult(
            ok=True,
            outputs=[out_compare],
            logs=logs,
            compare_only_count=len(compare_only_rows),
            roster_only_count=len(roster_only_rows),
            matched_count=len(matched_rows),
            unresolved_count=len(unresolved_rows),
            transfer_in_done=len(transfer_in_done),
            transfer_in_hold=len(transfer_in_hold),
            transfer_out_done=len(transfer_out_done),
            transfer_out_hold=len(transfer_out_hold),
            roster_only_rows=roster_only_rows,
            matched_rows=matched_rows,
            compare_only_rows=compare_only_rows,
            unresolved_rows=unresolved_rows,
        )

        log("[DONE] 비교 실행 완료")
        return pr

    except Exception as e:
        import traceback
        if not isinstance(e, ValueError):
            log(f"[DEBUG] {traceback.format_exc()}")
        log(f"[ERROR] {e}")
        return DiffPipelineResult(ok=False, outputs=[], logs=logs)


# =========================
# L5. Run wrapper
# =========================
def run_diff_pipeline(
    work_root: Path,
    school_name: str,
    target_year: Optional[int],
    school_start_date: date,
    work_date: date,
    roster_basis_date: Optional[date] = None,
    roster_xlsx: Optional[Path] = None,
    col_map: Optional[dict] = None,
    layout_overrides: Optional[dict] = None,
) -> DiffPipelineResult:

    logs: List[str] = []

    def log(msg: str):
        logs.append(msg)

    work_root   = Path(work_root).resolve()
    school_name = (school_name or "").strip()
    target_year = int(target_year) if target_year is not None else int(school_start_date.year)

    if not school_name:
        log("[ERROR] 학교명을 입력해 주세요.")
        return DiffPipelineResult(ok=False, outputs=[], logs=logs)

    try:
        scan = scan_diff_pipeline(
            work_root=work_root,
            school_name=school_name,
            target_year=target_year,
            school_start_date=school_start_date,
            work_date=work_date,
            roster_basis_date=roster_basis_date,
            roster_xlsx=roster_xlsx,
            col_map=col_map,
            layout_overrides=layout_overrides,
        )

        logs.extend(scan.logs)

        if not scan.ok:
            log("[ERROR] 스캔 단계에서 오류가 발생했습니다.")
            return DiffPipelineResult(ok=False, outputs=[], logs=logs)

        if not scan.can_execute:
            msg = ", ".join(scan.missing_fields) if scan.missing_fields else "필수 파일 누락"
            log(f"[ERROR] 실행할 수 없습니다. ({msg})")
            return DiffPipelineResult(ok=False, outputs=[], logs=logs)

        result = execute_diff_pipeline(scan=scan)
        result.logs = list(dict.fromkeys(logs + list(result.logs or [])))
        return result

    except Exception as e:
        import traceback
        if not isinstance(e, ValueError):
            log(f"[DEBUG] {traceback.format_exc()}")
        log(f"[ERROR] 실행 중 문제가 발생했습니다: {e}")
        return DiffPipelineResult(ok=False, outputs=[], logs=logs)